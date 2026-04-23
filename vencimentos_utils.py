"""Apoio do app de vencimentos do comparador."""

import io

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@st.cache_data
def carregar_dados_vencimentos(arquivo):
    """Carrega a base e já descarta as linhas que vieram completamente vazias."""
    try:
        if arquivo.name.endswith(".csv"):
            df = pd.read_csv(arquivo, encoding="utf-8", sep=";", decimal=",", header=0)
        else:
            df = pd.read_excel(arquivo, header=0)
        return df.dropna(how="all")
    except Exception as exc:
        st.error(f"Erro ao carregar arquivo: {exc}")
        return None


def _serie_vencimento_datetime(df, coluna_vencimento):
    """Converte a coluna de vencimento para data do jeito mais direto possível."""
    if df[coluna_vencimento].dtype == "datetime64[ns]":
        return df[coluna_vencimento]
    return pd.to_datetime(
        df[coluna_vencimento],
        errors="coerce",
        dayfirst=True,
    )


def remover_duplicados_manter_mais_antiga(df, coluna_auto, coluna_vencimento):
    """Quando o auto repete, fica com o vencimento mais antigo."""
    df_resultado = df.copy()
    try:
        df_resultado["_VENCIMENTO_DT"] = _serie_vencimento_datetime(df_resultado, coluna_vencimento)
        df_resultado = df_resultado.sort_values(
            by=[coluna_auto, "_VENCIMENTO_DT"],
            ascending=[True, True],
            na_position="last",
        )
        return df_resultado.drop_duplicates(subset=[coluna_auto], keep="first").drop(columns=["_VENCIMENTO_DT"])
    except Exception as exc:
        st.error(f"Erro ao remover duplicados: {exc}")
        return df


def extrair_ano_vencimento(df, coluna_vencimento):
    """Cria a coluna de ano de vencimento para filtros e resumo do app."""
    df_resultado = df.copy()
    try:
        df_resultado["_VENCIMENTO_DT"] = _serie_vencimento_datetime(df_resultado, coluna_vencimento)
        df_resultado["ANO_VENCIMENTO"] = df_resultado["_VENCIMENTO_DT"].dt.year
        return df_resultado.drop(columns=["_VENCIMENTO_DT"])
    except Exception as exc:
        st.error(f"Erro ao extrair ano: {exc}")
        return df


def gerar_excel_vencimentos_formatado(dados_df, nome_aba, nome_arquivo):
    """Gera o Excel já com o layout usado nas exportações desse app."""
    del nome_arquivo
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            dados_df.to_excel(writer, sheet_name=nome_aba, index=False, header=True)
            worksheet = writer.sheets[nome_aba]
            num_colunas = len(dados_df.columns)

            for col_idx, col in enumerate(dados_df.columns):
                col_letter = chr(65 + col_idx)
                if col == "IDENTIFICADOR DE DÉBITO":
                    worksheet.column_dimensions[col_letter].width = 25
                elif col == "Nº DE PROCESSO":
                    worksheet.column_dimensions[col_letter].width = 20
                elif col in {"DATA DE VENCIMENTO", "CNPJ", "MODAL"}:
                    worksheet.column_dimensions[col_letter].width = 18
                else:
                    worksheet.column_dimensions[col_letter].width = 15

            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            idx_por_coluna = {nome: idx + 1 for idx, nome in enumerate(dados_df.columns)}
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                for cell in row:
                    cell.border = thin_border
                    if cell.row <= 1:
                        continue
                    if cell.column == idx_por_coluna.get("CNPJ"):
                        cell.number_format = "@"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif cell.column == idx_por_coluna.get("IDENTIFICADOR DE DÉBITO"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif cell.column == idx_por_coluna.get("Nº DE PROCESSO"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif cell.column == idx_por_coluna.get("DATA DE VENCIMENTO"):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.number_format = "@"
                    elif cell.column == idx_por_coluna.get("MODAL"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.number_format = "@"

            worksheet.freeze_panes = "A2"

        buffer.seek(0)
        return buffer.getvalue()
    except Exception:
        buffer.close()
        raise
