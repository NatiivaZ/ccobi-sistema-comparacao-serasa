"""Geração de arquivos Excel formatados para o Sistema de Comparação SERASA."""

import io
from datetime import date, datetime

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


COLUNAS_DATA_EXCEL = {"Data de Vencimento", "Data Infração", "Data Pagamento"}
COLUNAS_MOEDA_EXCEL = {"Valor", "Valor (R$)"}
FORMATO_MOEDA_EXCEL = '"R$" #,##0.00'


def _converter_valor_data_excel(valor):
    """Converte valores comuns de planilha em datas reais do Excel."""
    if pd.isna(valor):
        return pd.NaT

    if isinstance(valor, pd.Timestamp):
        return valor.normalize()

    if isinstance(valor, (datetime, date)):
        return pd.Timestamp(valor).normalize()

    if isinstance(valor, (int, float, np.integer, np.floating)):
        valor_float = float(valor)
        if 1 <= valor_float <= 2958465:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(valor_float, unit="D")).normalize()

    texto = str(valor).strip()
    if not texto:
        return pd.NaT

    return pd.to_datetime(texto, errors="coerce", dayfirst=True)


def _preparar_colunas_data_para_excel(dados_df):
    """Transforma colunas de data em datetime para o Excel agrupar corretamente."""
    dados_excel = dados_df.copy()

    for coluna in COLUNAS_DATA_EXCEL.intersection(dados_excel.columns):
        serie_original = dados_excel[coluna]
        mascara_preenchida = serie_original.notna() & serie_original.astype("string").str.strip().ne("")
        if not mascara_preenchida.any():
            continue

        serie_convertida = serie_original.map(_converter_valor_data_excel)
        if serie_convertida.notna().sum() != mascara_preenchida.sum():
            continue

        dados_excel[coluna] = serie_convertida

    return dados_excel


def _converter_valor_moeda_excel(valor):
    """Converte textos monetários em número real para o Excel."""
    if pd.isna(valor):
        return np.nan

    if isinstance(valor, (int, float, np.integer, np.floating)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return np.nan

    texto = (
        texto.replace("R$", "")
        .replace(".", "")
        .replace(",", ".")
        .strip()
    )

    try:
        return float(texto)
    except ValueError:
        return np.nan


def _preparar_colunas_moeda_para_excel(dados_df):
    """Transforma colunas monetárias em número para o Excel aplicar moeda real."""
    dados_excel = dados_df.copy()

    for coluna in COLUNAS_MOEDA_EXCEL.intersection(dados_excel.columns):
        serie_original = dados_excel[coluna]
        mascara_preenchida = serie_original.notna() & serie_original.astype("string").str.strip().ne("")
        if not mascara_preenchida.any():
            continue

        serie_convertida = serie_original.map(_converter_valor_moeda_excel)
        if serie_convertida.notna().sum() != mascara_preenchida.sum():
            continue

        dados_excel[coluna] = serie_convertida

    return dados_excel


def gerar_excel_formatado(dados_df, nome_aba, nome_arquivo):
    """Gera arquivo Excel formatado a partir de um DataFrame.
    Remove linhas totalmente vazias e reindexa para evitar linhas em branco entre os dados."""
    if dados_df is None or dados_df.empty:
        return None
    dados_df = dados_df.replace('', np.nan).dropna(how='all').reset_index(drop=True)
    if dados_df.empty:
        return None
    dados_excel = _preparar_colunas_data_para_excel(dados_df)
    dados_excel = _preparar_colunas_moeda_para_excel(dados_excel)
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            dados_excel.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False,
                header=True
            )

            worksheet = writer.sheets[nome_aba]

            num_colunas = len(dados_excel.columns)
            if num_colunas == 5:
                worksheet.column_dimensions['A'].width = 25
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 18
                worksheet.column_dimensions['D'].width = 18
                worksheet.column_dimensions['E'].width = 15
            elif num_colunas == 4:
                if 'Número de Protocolo' in dados_excel.columns:
                    worksheet.column_dimensions['A'].width = 25
                    worksheet.column_dimensions['B'].width = 20
                    worksheet.column_dimensions['C'].width = 18
                    worksheet.column_dimensions['D'].width = 15
                else:
                    worksheet.column_dimensions['A'].width = 25
                    worksheet.column_dimensions['B'].width = 18
                    worksheet.column_dimensions['C'].width = 18
                    worksheet.column_dimensions['D'].width = 15
            else:
                worksheet.column_dimensions['A'].width = 25
                worksheet.column_dimensions['B'].width = 18
                worksheet.column_dimensions['C'].width = 15

            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left_center = Alignment(horizontal="left", vertical="center")
            align_center_center = Alignment(horizontal="center", vertical="center")
            align_right_center = Alignment(horizontal="right", vertical="center")

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            num_colunas = len(dados_excel.columns)
            tem_protocolo = 'Número de Protocolo' in dados_excel.columns
            tem_data_venc = 'Data de Vencimento' in dados_excel.columns
            tem_data_infracao = 'Data Infração' in dados_excel.columns
            tem_modais = 'Modais' in dados_excel.columns

            col_names = list(dados_excel.columns)
            idx_auto = 1
            idx_protocolo = 2 if tem_protocolo else None
            idx_data_venc = None
            idx_data_infracao = None
            idx_modais = None
            idx_cpf = None
            idx_valor = col_names.index('Valor') + 1 if 'Valor' in col_names else None
            tem_valor_r = 'Valor (R$)' in col_names
            idx_valor_r = col_names.index('Valor (R$)') + 1 if tem_valor_r else None
            tem_situacao_divida = 'Situação Dívida' in col_names
            idx_situacao_divida = col_names.index('Situação Dívida') + 1 if tem_situacao_divida else None
            tem_situacao_congelamento = 'Situação Congelamento' in col_names
            idx_situacao_congelamento = col_names.index('Situação Congelamento') + 1 if tem_situacao_congelamento else None
            tem_data_pagamento = 'Data Pagamento' in col_names
            idx_data_pagamento = col_names.index('Data Pagamento') + 1 if tem_data_pagamento else None
            tem_nome_autuado = 'Nome Autuado' in col_names
            idx_nome_autuado = col_names.index('Nome Autuado') + 1 if tem_nome_autuado else None
            tem_classificacao_autuado = 'Classificação Autuado' in col_names
            idx_classificacao_autuado = col_names.index('Classificação Autuado') + 1 if tem_classificacao_autuado else None
            tem_motivo_classificacao = 'Motivo Classificação' in col_names
            idx_motivo_classificacao = col_names.index('Motivo Classificação') + 1 if tem_motivo_classificacao else None
            tem_termo_identificado = 'Termo Identificado' in col_names
            idx_termo_identificado = col_names.index('Termo Identificado') + 1 if tem_termo_identificado else None
            tem_situacao_decadente = 'Situação decadente' in col_names
            idx_situacao_decadente = col_names.index('Situação decadente') + 1 if tem_situacao_decadente else None

            col_idx = 1
            if tem_protocolo:
                col_idx += 1
            if tem_data_venc:
                idx_data_venc = col_idx
                col_idx += 1
            if tem_data_infracao:
                idx_data_infracao = col_idx
                col_idx += 1
            if tem_modais:
                idx_modais = col_idx
                col_idx += 1
            idx_cpf = col_idx

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == idx_cpf and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_center_center
                    elif idx_valor and cell.column == idx_valor and cell.row > 1:
                        if cell.value is not None:
                            cell.number_format = FORMATO_MOEDA_EXCEL
                            cell.alignment = align_right_center
                    elif cell.column == idx_auto and cell.row > 1:
                        cell.alignment = align_left_center
                    elif idx_protocolo and cell.column == idx_protocolo and cell.row > 1:
                        cell.alignment = align_left_center
                    elif idx_situacao_divida and cell.column == idx_situacao_divida and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_left_center
                    elif idx_situacao_congelamento and cell.column == idx_situacao_congelamento and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_left_center
                    elif idx_data_pagamento and cell.column == idx_data_pagamento and cell.row > 1:
                        cell.number_format = 'dd/mm/yyyy'
                        cell.alignment = align_center_center
                    elif idx_nome_autuado and cell.column == idx_nome_autuado and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_left_center
                    elif idx_classificacao_autuado and cell.column == idx_classificacao_autuado and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_left_center
                    elif idx_motivo_classificacao and cell.column == idx_motivo_classificacao and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_left_center
                    elif idx_termo_identificado and cell.column == idx_termo_identificado and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_left_center
                    elif idx_data_venc and cell.column == idx_data_venc and cell.row > 1:
                        cell.alignment = align_center_center
                        cell.number_format = 'dd/mm/yyyy'
                    elif idx_data_infracao and cell.column == idx_data_infracao and cell.row > 1:
                        cell.alignment = align_center_center
                        cell.number_format = 'dd/mm/yyyy'
                    elif idx_modais and cell.column == idx_modais and cell.row > 1:
                        cell.alignment = align_left_center
                        cell.number_format = '@'
                    elif idx_valor_r and cell.column == idx_valor_r and cell.row > 1:
                        cell.number_format = FORMATO_MOEDA_EXCEL
                        cell.alignment = align_right_center
                    elif idx_situacao_decadente and cell.column == idx_situacao_decadente and cell.row > 1:
                        cell.number_format = '@'
                        cell.alignment = align_left_center

            if tem_data_infracao and idx_data_infracao is not None and idx_data_infracao <= 26:
                worksheet.column_dimensions[chr(64 + idx_data_infracao)].width = 18
            if tem_situacao_divida and idx_situacao_divida is not None and idx_situacao_divida <= 26:
                worksheet.column_dimensions[chr(64 + idx_situacao_divida)].width = 22
            if tem_situacao_congelamento and idx_situacao_congelamento is not None and idx_situacao_congelamento <= 26:
                worksheet.column_dimensions[chr(64 + idx_situacao_congelamento)].width = 22
            if tem_data_pagamento and idx_data_pagamento is not None and idx_data_pagamento <= 26:
                worksheet.column_dimensions[chr(64 + idx_data_pagamento)].width = 18
            if tem_nome_autuado and idx_nome_autuado is not None and idx_nome_autuado <= 26:
                worksheet.column_dimensions[chr(64 + idx_nome_autuado)].width = 40
            if tem_classificacao_autuado and idx_classificacao_autuado is not None and idx_classificacao_autuado <= 26:
                worksheet.column_dimensions[chr(64 + idx_classificacao_autuado)].width = 28
            if tem_motivo_classificacao and idx_motivo_classificacao is not None and idx_motivo_classificacao <= 26:
                worksheet.column_dimensions[chr(64 + idx_motivo_classificacao)].width = 38
            if tem_termo_identificado and idx_termo_identificado is not None and idx_termo_identificado <= 26:
                worksheet.column_dimensions[chr(64 + idx_termo_identificado)].width = 28
            if tem_valor_r and idx_valor_r is not None and idx_valor_r <= 26:
                worksheet.column_dimensions[chr(64 + idx_valor_r)].width = 15
            if tem_situacao_decadente and idx_situacao_decadente is not None and idx_situacao_decadente <= 26:
                worksheet.column_dimensions[chr(64 + idx_situacao_decadente)].width = 30

            worksheet.freeze_panes = 'A2'

        buffer.seek(0)
        return buffer.getvalue()
    except Exception as e:
        buffer.close()
        raise e
