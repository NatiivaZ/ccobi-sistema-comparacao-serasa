import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configuração da página
st.set_page_config(
    page_title="Sistema de Análise por Ano de Vencimento",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f4e79;
        text-align: center;
        padding: 1rem 0;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f4e79;
    }
    .stButton>button {
        width: 100%;
        background-color: #1f4e79;
        color: white;
        font-weight: bold;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Título principal
st.markdown('<div class="main-header">📅 Sistema de Análise por Ano de Vencimento</div>', unsafe_allow_html=True)
st.markdown('<p style="text-align: center; color: #666; font-size: 1.1rem;">Filtre e visualize autos de infração por ano de vencimento</p>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image("https://via.placeholder.com/200x80/1f4e79/ffffff?text=ANTT", use_container_width=True)
    st.markdown("### 📁 Upload de Arquivo")
    
    arquivo = st.file_uploader(
        "Selecione a planilha",
        type=['xlsx', 'xls', 'csv'],
        key='arquivo_vencimentos'
    )
    
    st.markdown("---")
    st.markdown("### ⚙️ Configurações")
    
    st.markdown("#### 🔑 Colunas Obrigatórias")
    # Campo para identificar Auto de Infração
    coluna_auto = st.text_input(
        "Nome da coluna Auto de Infração",
        value="Identificador do Débito",
        help="⚠️ OBRIGATÓRIO: Digite o nome exato da coluna que contém os Autos de Infração"
    )
    
    # Campo para identificar data de vencimento
    coluna_vencimento = st.text_input(
        "Nome da coluna Vencimento",
        value="Data do Vencimento",
        help="⚠️ OBRIGATÓRIO: Digite o nome exato da coluna que contém as datas de vencimento"
    )
    
    st.markdown("---")
    st.markdown("#### 📋 Colunas Opcionais")
    # Campo para identificar número de protocolos (opcional)
    coluna_protocolo = st.text_input(
        "Nome da coluna Nº do Processo (Opcional)",
        value="Nº do Processo",
        help="Digite o nome exato da coluna que contém os números de protocolos (opcional)"
    )
    
    # Campo para identificar Subtipo de Débito (Modal) (opcional)
    coluna_modal = st.text_input(
        "Nome da coluna Subtipo de Débito (Opcional)",
        value="Subtipo de Débito",
        help="Digite o nome exato da coluna que contém os modais (opcional)"
    )
    
    # Campo para identificar CPF/CNPJ (opcional)
    coluna_cpf_cnpj = st.text_input(
        "Nome da coluna CPF/CNPJ (Opcional)",
        value="CPF/CNPJ",
        help="Digite o nome exato da coluna que contém CPF/CNPJ (opcional)"
    )

# Função para carregar dados
@st.cache_data
def carregar_dados(arquivo):
    try:
        if arquivo.name.endswith('.csv'):
            df = pd.read_csv(arquivo, encoding='utf-8', sep=';', decimal=',', header=0)
        else:
            df = pd.read_excel(arquivo, header=0)
        
        # Remover linhas completamente vazias
        df = df.dropna(how='all')
        
        return df
    except Exception as e:
        st.error(f"Erro ao carregar arquivo: {str(e)}")
        return None

# Função para remover duplicados mantendo data mais antiga
def remover_duplicados_manter_mais_antiga(df, coluna_auto, coluna_vencimento):
    """Remove duplicados baseado em Auto de Infração, mantendo a data de vencimento mais antiga"""
    df_resultado = df.copy()
    
    # Converter data de vencimento para datetime
    try:
        if df_resultado[coluna_vencimento].dtype != 'datetime64[ns]':
            df_resultado['_VENCIMENTO_DT'] = pd.to_datetime(
                df_resultado[coluna_vencimento],
                errors='coerce',
                dayfirst=True,
                infer_datetime_format=True
            )
        else:
            df_resultado['_VENCIMENTO_DT'] = df_resultado[coluna_vencimento]
        
        # Ordenar por Auto de Infração e depois por data (mais antiga primeiro)
        df_resultado = df_resultado.sort_values(
            by=[coluna_auto, '_VENCIMENTO_DT'],
            ascending=[True, True],  # Mais antiga primeiro
            na_position='last'
        )
        
        # Remover duplicados mantendo a primeira ocorrência (que será a mais antiga)
        df_resultado = df_resultado.drop_duplicates(
            subset=[coluna_auto],
            keep='first'
        ).copy()
        
        # Remover coluna auxiliar
        if '_VENCIMENTO_DT' in df_resultado.columns:
            df_resultado = df_resultado.drop(columns=['_VENCIMENTO_DT'])
        
        return df_resultado
    except Exception as e:
        st.error(f"Erro ao remover duplicados: {str(e)}")
        return df

# Função para extrair ano da data de vencimento
def extrair_ano_vencimento(df, coluna_vencimento):
    """Extrai o ano da coluna de vencimento"""
    df_resultado = df.copy()
    
    # Converter para datetime
    try:
        if df_resultado[coluna_vencimento].dtype != 'datetime64[ns]':
            df_resultado['_VENCIMENTO_DT'] = pd.to_datetime(
                df_resultado[coluna_vencimento],
                errors='coerce',
                dayfirst=True,
                infer_datetime_format=True
            )
        else:
            df_resultado['_VENCIMENTO_DT'] = df_resultado[coluna_vencimento]
        
        # Extrair ano
        df_resultado['ANO_VENCIMENTO'] = df_resultado['_VENCIMENTO_DT'].dt.year
        
        # Remover coluna auxiliar
        if '_VENCIMENTO_DT' in df_resultado.columns:
            df_resultado = df_resultado.drop(columns=['_VENCIMENTO_DT'])
        
        return df_resultado
    except Exception as e:
        st.error(f"Erro ao extrair ano: {str(e)}")
        return df

# Função para formatar CPF/CNPJ no formato brasileiro
def formatar_cpf_cnpj_brasileiro(valor):
    """Formata CPF/CNPJ no formato brasileiro"""
    if pd.isna(valor) or valor == '' or valor is None:
        return ''
    
    # Remove caracteres não numéricos
    valor_str = str(valor).replace('.', '').replace('-', '').replace('/', '').strip()
    
    # Se estiver vazio, retorna vazio
    if not valor_str or not valor_str.isdigit():
        return str(valor)
    
    # Formatar CPF (11 dígitos)
    if len(valor_str) == 11:
        return f"{valor_str[0:3]}.{valor_str[3:6]}.{valor_str[6:9]}-{valor_str[9:11]}"
    
    # Formatar CNPJ (14 dígitos)
    elif len(valor_str) == 14:
        return f"{valor_str[0:2]}.{valor_str[2:5]}.{valor_str[5:8]}/{valor_str[8:12]}-{valor_str[12:14]}"
    
    # Se não tiver 11 ou 14 dígitos, retorna o valor original
    return str(valor)

# Função para gerar Excel formatado (mesmo estilo do app.py)
def gerar_excel_formatado(dados_df, nome_aba, nome_arquivo):
    """Gera arquivo Excel formatado a partir de um DataFrame"""
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            dados_df.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False,
                header=True
            )
            
            worksheet = writer.sheets[nome_aba]
            
            # Aplicar formatação completa
            num_colunas = len(dados_df.columns)
            tem_protocolo = 'Nº DE PROCESSO' in dados_df.columns
            tem_data_venc = 'DATA DE VENCIMENTO' in dados_df.columns
            tem_cpf = 'CNPJ' in dados_df.columns
            tem_modal = 'MODAL' in dados_df.columns
            
            # Ajustar larguras das colunas
            col_idx = 0
            for col in dados_df.columns:
                col_letter = chr(65 + col_idx)  # A, B, C, etc.
                if col == 'IDENTIFICADOR DE DÉBITO':
                    worksheet.column_dimensions[col_letter].width = 25
                elif col == 'Nº DE PROCESSO':
                    worksheet.column_dimensions[col_letter].width = 20
                elif col == 'DATA DE VENCIMENTO':
                    worksheet.column_dimensions[col_letter].width = 18
                elif col == 'CNPJ':
                    worksheet.column_dimensions[col_letter].width = 18
                elif col == 'MODAL':
                    worksheet.column_dimensions[col_letter].width = 18
                else:
                    worksheet.column_dimensions[col_letter].width = 15
                col_idx += 1
            
            # Formatação do cabeçalho
            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Formatar colunas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Calcular índices das colunas
            idx_auto = None
            idx_protocolo = None
            idx_data_venc = None
            idx_cpf = None
            idx_modal = None
            
            col_idx = 1
            for col in dados_df.columns:
                if col == 'IDENTIFICADOR DE DÉBITO':
                    idx_auto = col_idx
                elif col == 'Nº DE PROCESSO':
                    idx_protocolo = col_idx
                elif col == 'DATA DE VENCIMENTO':
                    idx_data_venc = col_idx
                elif col == 'CNPJ':
                    idx_cpf = col_idx
                elif col == 'MODAL':
                    idx_modal = col_idx
                col_idx += 1
            
            # Aplicar formatação nas células
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Não formatar cabeçalho
                        if idx_cpf and cell.column == idx_cpf:  # CNPJ
                            cell.number_format = '@'
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif idx_auto and cell.column == idx_auto:  # Identificador de Débito
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif idx_protocolo and cell.column == idx_protocolo:  # Nº de Processo
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif idx_data_venc and cell.column == idx_data_venc:  # Data de Vencimento
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.number_format = '@'  # Formato texto para manter formato DD/MM/YYYY
                        elif idx_modal and cell.column == idx_modal:  # Modal
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.number_format = '@'  # Formato texto
            
            worksheet.freeze_panes = 'A2'
        
        buffer.seek(0)
        excel_data = buffer.getvalue()
        return excel_data
    except Exception as e:
        buffer.close()
        raise e

# Main
if arquivo:
    st.markdown("---")
    
    # Carregar dados
    with st.spinner("Carregando planilha..."):
        df = carregar_dados(arquivo)
    
    if df is not None:
        # Mostrar preview
        st.markdown("### 📋 Preview da Planilha")
        st.dataframe(df.head(), use_container_width=True)
        st.caption(f"Total de registros: {len(df):,}")
        st.caption(f"Colunas: {', '.join(df.columns.tolist()[:10])}...")
        
        st.markdown("---")
        
        # Verificar se as colunas obrigatórias existem
        if coluna_auto not in df.columns:
            st.error(f"⚠️ Coluna '{coluna_auto}' não encontrada na planilha. Esta coluna é OBRIGATÓRIA!")
        elif coluna_vencimento not in df.columns:
            st.error(f"⚠️ Coluna '{coluna_vencimento}' não encontrada na planilha. Esta coluna é OBRIGATÓRIA!")
        else:
            # Botão de análise
            if st.button("🚀 Analisar por Ano de Vencimento", type="primary", use_container_width=True):
                with st.spinner("Analisando dados por ano de vencimento..."):
                    # PASSO 1: Remover duplicados ANTES de analisar por ano (manter data mais antiga)
                    df_sem_duplicados = remover_duplicados_manter_mais_antiga(df, coluna_auto, coluna_vencimento)
                    
                    # PASSO 2: Extrair ano de vencimento
                    df_com_ano = extrair_ano_vencimento(df_sem_duplicados, coluna_vencimento)
                    
                    # Remover linhas sem ano válido
                    df_com_ano = df_com_ano[df_com_ano['ANO_VENCIMENTO'].notna()].copy()
                    
                    # Agrupar por ano
                    anos_encontrados = sorted(df_com_ano['ANO_VENCIMENTO'].unique())
                    anos_encontrados = [int(ano) for ano in anos_encontrados if not pd.isna(ano)]
                    
                    # Calcular estatísticas por ano
                    stats_por_ano = {}
                    for ano in anos_encontrados:
                        df_ano = df_com_ano[df_com_ano['ANO_VENCIMENTO'] == ano].copy()
                        stats_por_ano[ano] = {
                            'quantidade': len(df_ano),
                            'dataframe': df_ano
                        }
                    
                    # Calcular quantos duplicados foram removidos
                    total_original = len(df)
                    total_sem_duplicados = len(df_sem_duplicados)
                    duplicados_removidos = total_original - total_sem_duplicados
                    
                    # Armazenar no session_state
                    st.session_state['df_com_ano'] = df_com_ano
                    st.session_state['anos_encontrados'] = anos_encontrados
                    st.session_state['stats_por_ano'] = stats_por_ano
                    st.session_state['coluna_auto'] = coluna_auto
                    st.session_state['coluna_vencimento'] = coluna_vencimento
                    st.session_state['coluna_cpf_cnpj'] = coluna_cpf_cnpj
                    st.session_state['coluna_modal'] = coluna_modal
                    st.session_state['coluna_protocolo'] = coluna_protocolo
                    st.session_state['duplicados_removidos'] = duplicados_removidos
                    st.session_state['total_original'] = total_original
                    
                    st.success("✅ Análise concluída com sucesso!")
                    if duplicados_removidos > 0:
                        st.info(f"ℹ️ {duplicados_removidos:,} autos duplicados foram removidos (mantida a data de vencimento mais antiga).")
                    st.rerun()

# Exibir resultados
if 'df_com_ano' in st.session_state:
    df_com_ano = st.session_state['df_com_ano']
    anos_encontrados = st.session_state['anos_encontrados']
    stats_por_ano = st.session_state['stats_por_ano']
    coluna_auto = st.session_state['coluna_auto']
    coluna_vencimento = st.session_state['coluna_vencimento']
    coluna_cpf_cnpj = st.session_state.get('coluna_cpf_cnpj', '')
    coluna_modal = st.session_state.get('coluna_modal', '')
    coluna_protocolo = st.session_state.get('coluna_protocolo', '')
    
    st.markdown("---")
    st.markdown("## 📊 Dashboard - Análise por Ano de Vencimento")
    
    # Informação sobre duplicados removidos
    duplicados_removidos = st.session_state.get('duplicados_removidos', 0)
    total_original = st.session_state.get('total_original', 0)
    if duplicados_removidos > 0:
        st.info(f"ℹ️ **{duplicados_removidos:,} autos duplicados foram removidos** (de {total_original:,} registros originais). Mantida a data de vencimento mais antiga para cada auto.")
    
    # Métricas principais
    total_autos = len(df_com_ano)
    total_anos = len(anos_encontrados)
    ano_mais_autos = max(anos_encontrados, key=lambda x: stats_por_ano[x]['quantidade']) if anos_encontrados else None
    qtd_mais_autos = stats_por_ano[ano_mais_autos]['quantidade'] if ano_mais_autos else 0
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Total de Autos",
            f"{total_autos:,}",
            delta="Autos com vencimento válido"
        )
    
    with col2:
        st.metric(
            "Anos Encontrados",
            f"{total_anos}",
            delta="Anos diferentes"
        )
    
    with col3:
        st.metric(
            "Ano com Mais Autos",
            f"{ano_mais_autos}" if ano_mais_autos else "N/A",
            delta=f"{qtd_mais_autos:,} autos" if ano_mais_autos else ""
        )
    
    with col4:
        ano_menos_autos = min(anos_encontrados, key=lambda x: stats_por_ano[x]['quantidade']) if anos_encontrados else None
        qtd_menos_autos = stats_por_ano[ano_menos_autos]['quantidade'] if ano_menos_autos else 0
        st.metric(
            "Ano com Menos Autos",
            f"{ano_menos_autos}" if ano_menos_autos else "N/A",
            delta=f"{qtd_menos_autos:,} autos" if ano_menos_autos else ""
        )
    
    # Gráfico de barras
    st.markdown("---")
    st.markdown("### 📈 Distribuição de Autos por Ano")
    
    anos_ordenados = sorted(anos_encontrados)
    quantidades = [stats_por_ano[ano]['quantidade'] for ano in anos_ordenados]
    
    fig_barras = go.Figure(data=[
        go.Bar(
            x=anos_ordenados,
            y=quantidades,
            marker_color='#1f4e79',
            text=quantidades,
            textposition='auto'
        )
    ])
    fig_barras.update_layout(
        title="Quantidade de Autos por Ano de Vencimento",
        xaxis_title="Ano",
        yaxis_title="Quantidade de Autos",
        height=400,
        showlegend=False
    )
    st.plotly_chart(fig_barras, use_container_width=True)
    
    # Gráfico de pizza
    col1, col2 = st.columns(2)
    
    with col1:
        fig_pizza = go.Figure(data=[go.Pie(
            labels=[str(ano) for ano in anos_ordenados],
            values=quantidades,
            hole=0.4,
            marker_colors=px.colors.qualitative.Set3
        )])
        fig_pizza.update_layout(
            title="Distribuição Percentual por Ano",
            height=400
        )
        st.plotly_chart(fig_pizza, use_container_width=True)
    
    with col2:
        # Tabela resumo
        st.markdown("### 📋 Resumo por Ano")
        df_resumo = pd.DataFrame({
            'Ano': anos_ordenados,
            'Quantidade de Autos': quantidades,
            'Percentual': [f"{(qtd/total_autos)*100:.2f}%" for qtd in quantidades]
        })
        st.dataframe(df_resumo, use_container_width=True, hide_index=True)
    
    # Área de Exportação
    st.markdown("---")
    st.markdown("### 📥 Exportar por Ano de Vencimento")
    st.info("💡 Baixe planilhas separadas por ano de vencimento. Cada arquivo contém apenas os autos daquele ano específico.")
    
    # Data de extração
    data_arquivo = datetime.now().strftime('%d %m %Y %H:%M')
    
    # Criar colunas dinamicamente para os botões de download
    num_colunas_export = min(3, len(anos_encontrados))  # Máximo 3 colunas
    cols_export = st.columns(num_colunas_export)
    
    for idx, ano in enumerate(sorted(anos_encontrados, reverse=False)):  # Do mais antigo para o mais recente
        col_idx = idx % num_colunas_export
        with cols_export[col_idx]:
            st.markdown(f"##### 📅 Ano {ano}")
            
            df_ano = stats_por_ano[ano]['dataframe'].copy()
            qtd_autos = stats_por_ano[ano]['quantidade']
            
            # Preparar dados para exportação com nomes corretos das colunas
            # Ordem: IDENTIFICADOR DE DÉBITO, Nº DE PROCESSO, MODAL, CNPJ, DATA DE VENCIMENTO
            colunas_export = {}
            
            # IDENTIFICADOR DE DÉBITO (sempre incluir)
            colunas_export['IDENTIFICADOR DE DÉBITO'] = df_ano[coluna_auto].fillna('').astype(str).str.strip()
            
            # Nº DE PROCESSO (se existir)
            if coluna_protocolo and coluna_protocolo in df_ano.columns:
                colunas_export['Nº DE PROCESSO'] = df_ano[coluna_protocolo].fillna('').astype(str).str.strip()
            
            # MODAL (se existir)
            if coluna_modal and coluna_modal in df_ano.columns:
                colunas_export['MODAL'] = df_ano[coluna_modal].fillna('').astype(str).str.strip()
            
            # CNPJ (se existir)
            if coluna_cpf_cnpj and coluna_cpf_cnpj in df_ano.columns:
                colunas_export['CNPJ'] = df_ano[coluna_cpf_cnpj].fillna('').astype(str).str.strip().apply(formatar_cpf_cnpj_brasileiro)
            
            # DATA DE VENCIMENTO (sempre incluir)
            try:
                vencimento_dt = pd.to_datetime(
                    df_ano[coluna_vencimento],
                    errors='coerce',
                    dayfirst=True
                )
                colunas_export['DATA DE VENCIMENTO'] = vencimento_dt.dt.strftime('%d/%m/%Y').fillna('')
            except:
                colunas_export['DATA DE VENCIMENTO'] = df_ano[coluna_vencimento].fillna('').astype(str).str.strip()
            
            # Criar DataFrame de exportação com ordem correta
            ordem_colunas = ['IDENTIFICADOR DE DÉBITO']
            if 'Nº DE PROCESSO' in colunas_export:
                ordem_colunas.append('Nº DE PROCESSO')
            if 'MODAL' in colunas_export:
                ordem_colunas.append('MODAL')
            if 'CNPJ' in colunas_export:
                ordem_colunas.append('CNPJ')
            ordem_colunas.append('DATA DE VENCIMENTO')
            
            dados_exportacao = pd.DataFrame({col: colunas_export[col] for col in ordem_colunas}, index=df_ano.index)
            
            # Remover linhas completamente vazias
            dados_exportacao = dados_exportacao.dropna(how='all')
            
            if not dados_exportacao.empty:
                try:
                    excel_ano = gerar_excel_formatado(
                        dados_exportacao,
                        f'Autos_{ano}',
                        f'Autos Vencimento {ano} {data_arquivo}.xlsx'
                    )
                    
                    st.download_button(
                        label=f"📥 Download Autos {ano}",
                        data=excel_ano,
                        file_name=f"Autos Vencimento {ano} {data_arquivo}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"download_ano_{ano}",
                        help=f"Arquivo Excel com todos os autos de vencimento {ano}"
                    )
                    
                    st.success(f"✅ {qtd_autos:,} autos")
                    st.caption(f"📊 Ano {ano}")
                except Exception as e:
                    st.error(f"⚠️ Erro ao gerar arquivo: {str(e)}")
            else:
                st.warning(f"⚠️ Nenhum auto encontrado para {ano}")
    
    # Visualização detalhada por ano
    st.markdown("---")
    st.markdown("### 📋 Visualização Detalhada por Ano")
    
    # Seletor de ano para visualização
    ano_selecionado = st.selectbox(
        "Selecione um ano para visualizar os autos:",
        options=sorted(anos_encontrados, reverse=True),
        key="ano_visualizacao"
    )
    
    if ano_selecionado:
        df_ano_selecionado = stats_por_ano[ano_selecionado]['dataframe'].copy()
        
        st.markdown(f"#### 📅 Autos de Vencimento {ano_selecionado}")
        st.info(f"💡 Mostrando {len(df_ano_selecionado):,} autos com vencimento em {ano_selecionado}")
        
        # Preparar colunas para exibição
        colunas_exibicao = [coluna_auto]
        if coluna_protocolo and coluna_protocolo in df_ano_selecionado.columns:
            colunas_exibicao.append(coluna_protocolo)
        if coluna_modal and coluna_modal in df_ano_selecionado.columns:
            colunas_exibicao.append(coluna_modal)
        if coluna_cpf_cnpj and coluna_cpf_cnpj in df_ano_selecionado.columns:
            colunas_exibicao.append(coluna_cpf_cnpj)
        if coluna_vencimento in df_ano_selecionado.columns:
            colunas_exibicao.append(coluna_vencimento)
        
        # Mostrar apenas colunas que existem
        colunas_exibicao = [col for col in colunas_exibicao if col in df_ano_selecionado.columns]
        
        if colunas_exibicao:
            st.dataframe(
                df_ano_selecionado[colunas_exibicao],
                use_container_width=True,
                height=400
            )
        else:
            st.warning("⚠️ Nenhuma coluna disponível para exibição")
    
    # Resumo final
    st.markdown("---")
    st.caption(f"📅 Data de extração: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    st.caption(f"✅ **Total de autos analisados:** {total_autos:,}")
    st.caption(f"📊 **Anos encontrados:** {', '.join([str(ano) for ano in sorted(anos_encontrados)])}")

else:
    st.info("👆 Por favor, faça o upload da planilha na barra lateral para iniciar a análise.")
    
    # Mostrar instruções
    with st.expander("ℹ️ Como usar o sistema"):
        st.markdown("""
        ### Instruções de Uso:
        
        1. **Upload de Arquivo**: Faça o upload da planilha na barra lateral
        2. **Configuração de Colunas**: Informe os nomes exatos das colunas:
           - ⚠️ **Coluna Auto de Infração** (OBRIGATÓRIA)
           - ⚠️ **Coluna Vencimento** (OBRIGATÓRIA)
           - Coluna CPF/CNPJ (Opcional)
           - Coluna Valor (Opcional)
           - Coluna Nº do Processo (Opcional)
        3. **Executar Análise**: Clique no botão "Analisar por Ano de Vencimento"
        4. **Visualizar Resultados**: Explore o dashboard com métricas e gráficos
        5. **Exportar**: Baixe planilhas separadas por ano de vencimento
        
        ### Funcionalidades:
        - ✅ Análise automática por ano de vencimento
        - ✅ Dashboard com métricas e gráficos
        - ✅ Exportação separada por ano
        - ✅ Visualização detalhada por ano
        - ✅ Formatação profissional das planilhas Excel
        """)

