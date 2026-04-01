import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date, timedelta
import io
import json
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED2

# Configuração da página
st.set_page_config(
    page_title="Sistema de Análise SERASA x Dívida Ativa",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f4e79;
        text-align: center;
        padding: 1rem 0;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f4e79;
    }
    .stButton>button {
        width: 100%;
        background-color: #1f4e79;
        color: white;
        font-weight: bold;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Título principal
st.markdown('<div class="main-header">📊 Sistema de Análise de Autos de Infração ANTT</div>', unsafe_allow_html=True)
st.markdown('<p style="text-align: center; color: #666; font-size: 1.1rem;">SERASA × Dívida Ativa - Análise Inteligente de Dados</p>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image("https://via.placeholder.com/200x80/1f4e79/ffffff?text=ANTT", use_container_width=True)
    st.markdown("### 📁 Upload de Arquivos")
    
    st.markdown("#### Base SERASA")
    arquivo_serasa = st.file_uploader(
        "Selecione a planilha SERASA",
        type=['xlsx', 'xls', 'csv'],
        key='serasa'
    )
    
    st.markdown("#### Base Dívida Ativa")
    arquivo_divida = st.file_uploader(
        "Selecione a planilha Dívida Ativa",
        type=['xlsx', 'xls', 'csv'],
        key='divida'
    )
    
    st.markdown("---")
    st.markdown("### ⚙️ Configurações")
    
    st.markdown("#### 🔑 Coluna Principal (Obrigatória)")
    # Campo para identificar Auto de Infração (PRINCIPAL)
    coluna_auto = st.text_input(
        "Nome da coluna Auto de Infração",
        value="Identificador do Débito",
        help="⚠️ OBRIGATÓRIO: Digite o nome exato da coluna que contém os Autos de Infração (ex: CRGPF00074552019)"
    )
    
    st.markdown("#### 📋 Colunas Adicionais")
    # Campo para identificar CPF/CNPJ
    coluna_cpf_cnpj = st.text_input(
        "Nome da coluna CPF/CNPJ",
        value="CPF/CNPJ",
        help="Digite o nome exato da coluna que contém CPF/CNPJ"
    )
    
    # Campo para identificar valor na SERASA
    coluna_valor = st.text_input(
        "Nome da coluna Valor (SERASA)",
        value="Valor Multa Atualizado",
        help="Digite o nome exato da coluna de valor na base SERASA"
    )
    
    # Campo para identificar valor na Dívida Ativa
    coluna_valor_divida = st.text_input(
        "Nome da coluna Valor (Dívida Ativa)",
        value="Valor Atualizado do Débito",
        help="Digite o nome exato da coluna de valor na base Dívida Ativa"
    )
    
    # Campo para identificar data de vencimento
    coluna_vencimento = st.text_input(
        "Nome da coluna Vencimento",
        value="Data do Vencimento",
        help="Digite o nome exato da coluna que contém as datas de vencimento"
    )
    
    # Campo para identificar número de protocolos
    coluna_protocolo = st.text_input(
        "Nome da coluna Nº do Processo",
        value="Nº do Processo",
        help="Digite o nome exato da coluna que contém os números de protocolos"
    )

    st.markdown("#### 🚛 Colunas de Modais")
    # Campo para identificar modal na SERASA
    coluna_modal_serasa = st.text_input(
        "Nome da coluna Modal (SERASA)",
        value="Tipo Modal",
        help="Digite o nome exato da coluna que contém os modais na base SERASA"
    )
    
    # Campo para identificar modal na Dívida Ativa
    coluna_modal_divida = st.text_input(
        "Nome da coluna Modal (Dívida Ativa)",
        value="Subtipo de Débito",
        help="Digite o nome exato da coluna que contém os modais na base Dívida Ativa"
    )

    st.markdown("#### 👤 Classificação de Autuados (exportação base SERASA)")
    coluna_nome_autuado = st.text_input(
        "Nome da coluna do autuado (SERASA)",
        value="Nome Autuado",
        help="Coluna usada para classificar se o autuado pode ou não ser cobrado (órgãos, bancos, leasing)"
    )

CONFIG_CLASSIFICACAO_PATH = Path(__file__).with_name("config_classificacao_autuados.json")

DEFAULT_CLASSIFICACAO_CONFIG = {
    "extras_orgao": [],
    "extras_banco": [],
    "extras_leasing": [],
    "excecoes_pode_cobrar": [],
}

EXCECOES_FIXAS_PODE_COBRAR = ["SAFRA"]

def carregar_config_classificacao():
    """Carrega configurações persistidas da classificação de autuados."""
    try:
        if CONFIG_CLASSIFICACAO_PATH.exists():
            data = json.loads(CONFIG_CLASSIFICACAO_PATH.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                config = DEFAULT_CLASSIFICACAO_CONFIG.copy()
                for chave in config.keys():
                    valor = data.get(chave, [])
                    config[chave] = valor if isinstance(valor, list) else []
                return config
    except Exception:
        pass
    return DEFAULT_CLASSIFICACAO_CONFIG.copy()

def salvar_config_classificacao(config):
    """Salva configurações persistidas da classificação de autuados."""
    config_limpo = {}
    for chave, valor in DEFAULT_CLASSIFICACAO_CONFIG.items():
        itens = config.get(chave, valor)
        if not isinstance(itens, list):
            itens = []
        itens = [str(v).strip() for v in itens if str(v).strip()]
        config_limpo[chave] = itens
    CONFIG_CLASSIFICACAO_PATH.write_text(
        json.dumps(config_limpo, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

def parse_lista_multilinha(texto):
    """Converte textarea em lista de termos."""
    if not texto:
        return []
    return [linha.strip() for linha in str(texto).splitlines() if linha.strip()]

if "classificacao_config" not in st.session_state:
    st.session_state["classificacao_config"] = carregar_config_classificacao()

with st.sidebar.expander("🛠️ Regras da classificação de autuados", expanded=False):
    st.caption("Adicione exceções e termos extras sem precisar alterar o código.")
    config_atual = st.session_state.get("classificacao_config", DEFAULT_CLASSIFICACAO_CONFIG.copy())
    extras_orgao_text = st.text_area(
        "Termos extras para Órgão",
        value="\n".join(config_atual.get("extras_orgao", [])),
        help="Um termo por linha. Ex.: POLICIA CIENTIFICA"
    )
    extras_banco_text = st.text_area(
        "Termos extras para Banco",
        value="\n".join(config_atual.get("extras_banco", [])),
        help="Um termo por linha. Ex.: BANCO MERCEDES"
    )
    extras_leasing_text = st.text_area(
        "Termos extras para Leasing",
        value="\n".join(config_atual.get("extras_leasing", [])),
        help="Um termo por linha. Ex.: LEASING OPERACIONAL"
    )
    excecoes_text = st.text_area(
        "Exceções - sempre Pode cobrar",
        value="\n".join(config_atual.get("excecoes_pode_cobrar", [])),
        help="Use aqui nomes/termos que estavam gerando falso positivo. Um por linha."
    )
    if st.button("💾 Salvar regras da classificação", use_container_width=True):
        novo_config = {
            "extras_orgao": parse_lista_multilinha(extras_orgao_text),
            "extras_banco": parse_lista_multilinha(extras_banco_text),
            "extras_leasing": parse_lista_multilinha(extras_leasing_text),
            "excecoes_pode_cobrar": parse_lista_multilinha(excecoes_text),
        }
        salvar_config_classificacao(novo_config)
        st.session_state["classificacao_config"] = novo_config
        st.success("✅ Regras salvas com sucesso.")

# Função para carregar dados
@st.cache_data
def carregar_dados(arquivo, nome_base):
    try:
        if arquivo.name.endswith('.csv'):
            # Carregar CSV garantindo que primeira linha seja cabeçalho
            # IMPORTANTE: Não usar skiprows ou outras opções que removam linhas
            df = pd.read_csv(arquivo, encoding='utf-8', sep=';', decimal=',', header=0)
        else:
            # Carregar Excel garantindo que primeira linha seja cabeçalho
            # IMPORTANTE: Não usar skiprows ou outras opções que removam linhas
            df = pd.read_excel(arquivo, header=0)
        
        # IMPORTANTE: Contar total ANTES de qualquer remoção
        # Isso garante que mostramos o total exato da planilha
        total_original = len(df)
        
        # IMPORTANTE: NÃO remover linhas automaticamente para preservar total exato
        # Apenas remover linhas completamente vazias (todas as colunas vazias)
        # Isso preserva linhas que tenham pelo menos uma coluna com valor
        df = df.dropna(how='all')
        
        # NÃO remover cabeçalhos duplicados automaticamente
        # Isso pode remover linhas válidas que por acaso tenham valores similares aos nomes das colunas
        
        # Armazenar total original no DataFrame para referência
        df.attrs['total_original'] = total_original
        
        return df
    except Exception as e:
        st.error(f"Erro ao carregar {nome_base}: {str(e)}")
        return None

# Função para normalizar CPF/CNPJ
def normalizar_cpf_cnpj(valor):
    if pd.isna(valor):
        return None
    # Remove caracteres não numéricos
    valor_str = str(valor).replace('.', '').replace('-', '').replace('/', '').strip()
    return valor_str if valor_str else None

# Função para normalizar Auto de Infração
def normalizar_auto(valor):
    """
    Normaliza o valor do Auto de Infração para comparação exata.
    Remove espaços, converte para maiúsculas e garante que seja string.
    """
    if pd.isna(valor):
        return None
    # Converter para string, remover espaços no início e fim, e converter para maiúsculas
    valor_str = str(valor).strip().upper()
    # Remover espaços múltiplos internos (caso existam)
    valor_str = ' '.join(valor_str.split())
    return valor_str if valor_str else None

# Função para converter valores como no SQL (limpeza completa)
def converter_valor_sql(valor):
    """
    Converte valor de texto para decimal seguindo a mesma lógica do SQL:
    TRY_CONVERT(decimal(18,2), REPLACE(REPLACE(REPLACE(REPLACE([Valor], 'R$', ''), ' ', ''), '.', ''), ',', '.'))

    ATENÇÃO: Se o valor já for numérico (int/float), retorna diretamente SEM
    manipulação de string, evitando o bug onde "241.11" → remove '.' → "24111"
    (100x maior). Esse bug ocorre quando a coluna da SERASA é lida pelo pandas
    como float, enquanto a da Dívida era texto formatado em BR ("R$ 241,11").
    """
    if pd.isna(valor):
        return None
    # Se já é número (Excel leu como float/int), retorna direto — não há vírgula/ponto a tratar
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        valor_str = str(valor).strip()
        # Remover R$
        valor_str = valor_str.replace('R$', '').replace('r$', '').replace('R$', '')
        # Remover espaços
        valor_str = valor_str.replace(' ', '')
        # Checar se é formato BR (tem vírgula como decimal): "1.234,56" ou "241,11"
        # Nesse caso: remover pontos (milhar) e trocar vírgula por ponto decimal
        if ',' in valor_str:
            valor_str = valor_str.replace('.', '')
            valor_str = valor_str.replace(',', '.')
        # Se não tem vírgula, pode ser já no padrão EN ("241.11") ou inteiro — não mexer nos pontos
        return float(valor_str)
    except:
        return None

# Função para resolver coluna de vencimento com variações de nome e sufixo
def resolver_coluna_vencimento(df, coluna_vencimento):
    if df is None or df.empty:
        return None
    colunas = list(df.columns)
    colunas_map = {c.lower(): c for c in colunas}

    candidatos = []
    if coluna_vencimento:
        candidatos.append(coluna_vencimento)
        if " do " in coluna_vencimento:
            candidatos.append(coluna_vencimento.replace(" do ", " de "))
        if " de " in coluna_vencimento:
            candidatos.append(coluna_vencimento.replace(" de ", " do "))
    candidatos += ["Data de Vencimento", "Data do Vencimento"]

    for cand in candidatos:
        for sufixo in ["_serasa", "_divida", ""]:
            chave = f"{cand}{sufixo}".lower()
            if chave in colunas_map:
                return colunas_map[chave]

    # Fallback: qualquer coluna que contenha "vencimento"
    cols_venc = [c for c in colunas if "vencimento" in c.lower()]
    if cols_venc:
        for pref in ["serasa", "divida"]:
            for c in cols_venc:
                if pref in c.lower():
                    return c
        return cols_venc[0]

    return None

# --- Classificação de autuados (não pode cobrar: órgão, banco, leasing) ---
import re
import unicodedata

def _normalizar_texto_para_busca(texto):
    """Remove acentos, padroniza separadores e adiciona espaços nas bordas."""
    if not texto or (isinstance(texto, float) and pd.isna(texto)):
        return ""
    s = unicodedata.normalize("NFD", str(texto).strip().upper())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return f" {s} " if s else ""

def _contem_expressao(texto_norm, expressao):
    """Confere expressao completa, evitando falso positivo por pedaço de palavra."""
    expr_norm = _normalizar_texto_para_busca(expressao)
    return bool(expr_norm) and expr_norm in texto_norm

def _contem_alguma_expressao(texto_norm, expressoes):
    return any(_contem_expressao(texto_norm, expr) for expr in expressoes)

_EXPRESSOES_LEASING = [
    "LEASING",
    "ARRENDAMENTO MERCANTIL",
    "ARREND MERCANTIL",
    "ARREND. MERCANTIL",
    "LOCACAO FINANCEIRA",
    "LOCACAO MERCANTIL",
]

_EXPRESSOES_ORGAO = [
    "CORPO DE BOMBEIROS",
    "BOMBEIRO MILITAR",
    "POLICIA MILITAR",
    "POLICIA CIVIL",
    "POLICIA FEDERAL",
    "POLICIA RODOVIARIA FEDERAL",
    "RODOVIARIA FEDERAL",
    "PRF",
    "EXERCITO BRASILEIRO",
    "EXERCITO",
    "MARINHA DO BRASIL",
    "MARINHA",
    "FORCA AEREA BRASILEIRA",
    "AERONAUTICA",
    "PREFEITURA MUNICIPAL",
    "MUNICIPIO DE",
    "GOVERNO DO ESTADO",
    "GOVERNO FEDERAL",
    "RECEITA FEDERAL",
    "MINISTERIO PUBLICO",
    "DEFENSORIA PUBLICA",
    "TRIBUNAL DE",
    "TRIBUNAL REGIONAL",
    "CAMARA MUNICIPAL",
    "SECRETARIA DE ESTADO",
    "SECRETARIA MUNICIPAL",
    "MINISTERIO DA",
    "MINISTERIO DO",
    "AUTARQUIA",
    "FUNDACAO PUBLICA",
    "INSS",
    "DATAPREV",
    "IBAMA",
    "INCRA",
    "DETRAN",
    "DEPARTAMENTO ESTADUAL DE TRANSITO",
    "CORREIOS",
    "EMPRESA BRASILEIRA DE CORREIOS E TELEGRAFOS",
    "AGENCIA NACIONAL DE AGUAS",
    "AGENCIA NACIONAL DE TRANSPORTES TERRESTRES",
]

_MARCADORES_BANCO = [
    "BANCO",
    "BANK",
    "FINANCEIRA",
    "INSTITUICAO FINANCEIRA",
    "SERVICOS FINANCEIROS",
    "FINANCIAL SERVICES",
    "CREDITO FINANCIAMENTO E INVESTIMENTO",
    "COOPERATIVA DE CREDITO",
]

_NOMES_BANCOS_ESPECIFICOS = [
    "BANCO DO BRASIL",
    "CAIXA ECONOMICA FEDERAL",
    "BANCO DA AMAZONIA",
    "BANCO DO NORDESTE",
    "BANCO DE BRASILIA",
    "BNDES",
    "BANRISUL",
    "BANESTES",
    "BANPARA",
    "BANESE",
    "BRADESCO",
    "SANTANDER",
    "ITAU UNIBANCO",
    "ITAU",
    "BTG PACTUAL",
    "SAFRA",
    "CITIBANK",
    "C6 BANK",
    "NUBANK",
    "BANCO INTER",
    "PAGBANK",
    "BANCO ORIGINAL",
    "BANCO PAN",
    "DAYCOVAL",
    "SICREDI",
    "SICOOB",
    "CRESOL",
    "AGIBANK",
    "OMNI BANCO",
    "PARANA BANCO",
    "MERCANTIL DO BRASIL",
    "BANCO SOFISA",
    "BANCO FIBRA",
    "BANCO GENIAL",
    "BANCO MODAL",
    "BANCO BS2",
    "DIGIO",
    "OURIBANK",
    "TRIBANCO",
    "ABC BRASIL",
    "ABN AMRO",
    "BNP PARIBAS",
    "BANK OF AMERICA",
    "SCOTIABANK",
    "DEUTSCHE BANK",
    "GOLDMAN SACHS",
    "JP MORGAN",
    "MORGAN STANLEY",
    "HSBC",
    "RABOBANK",
    "BMG",
    "BV",
]

def obter_config_classificacao_ativa():
    """Retorna a configuração ativa da classificação de autuados."""
    config = st.session_state.get("classificacao_config")
    if not isinstance(config, dict):
        config = carregar_config_classificacao()
        st.session_state["classificacao_config"] = config
    return config

def _lista_regras_classificacao(config):
    """Monta as listas efetivas de regras combinando defaults + extras."""
    config = config or DEFAULT_CLASSIFICACAO_CONFIG
    return {
        "orgao": _EXPRESSOES_ORGAO + list(config.get("extras_orgao", [])),
        "banco": _MARCADORES_BANCO + _NOMES_BANCOS_ESPECIFICOS + list(config.get("extras_banco", [])),
        "leasing": _EXPRESSOES_LEASING + list(config.get("extras_leasing", [])),
        "excecoes": EXCECOES_FIXAS_PODE_COBRAR + list(config.get("excecoes_pode_cobrar", [])),
    }

def classificar_autuado_detalhado(nome, config=None):
    """
    Retorna (classificacao, motivo, termo_encontrado) para o nome do autuado.
    """
    texto_norm = _normalizar_texto_para_busca(nome)
    if not texto_norm:
        return "Pode cobrar", "Nome vazio ou não informado", ""

    regras = _lista_regras_classificacao(config or obter_config_classificacao_ativa())

    for termo in regras["excecoes"]:
        if _contem_expressao(texto_norm, termo):
            return "Pode cobrar", "Exceção cadastrada", termo

    for termo in regras["leasing"]:
        if _contem_expressao(texto_norm, termo):
            return "Não pode cobrar - Leasing", "Correspondência com regra de leasing", termo

    for termo in regras["orgao"]:
        if _contem_expressao(texto_norm, termo):
            return "Não pode cobrar - Órgão", "Correspondência com regra de órgão público", termo

    for termo in regras["banco"]:
        if _contem_expressao(texto_norm, termo):
            return "Não pode cobrar - Banco", "Correspondência com regra de instituição financeira", termo

    return "Pode cobrar", "Nenhuma regra impeditiva encontrada", ""

def classificar_autuado(nome):
    """Wrapper simples para manter compatibilidade com chamadas antigas."""
    classificacao, _, _ = classificar_autuado_detalhado(nome)
    return classificacao

def filtrar_autuados_cobraveis(df_base, coluna_nome_autuado):
    """
    Remove da base os autuados classificados como:
    - Não pode cobrar - Órgão
    - Não pode cobrar - Banco
    - Não pode cobrar - Leasing

    Exceções configuradas (como SAFRA) permanecem na base.
    """
    if df_base is None or df_base.empty or not coluna_nome_autuado or coluna_nome_autuado not in df_base.columns:
        return df_base

    detalhes = df_base[coluna_nome_autuado].apply(classificar_autuado_detalhado)
    detalhes_df = pd.DataFrame(
        detalhes.tolist(),
        index=df_base.index,
        columns=['_CLASSIFICACAO_AUTUADO', '_MOTIVO_CLASSIFICACAO', '_TERMO_IDENTIFICADO']
    )
    df_filtrado = df_base.copy()
    df_filtrado['_CLASSIFICACAO_AUTUADO'] = detalhes_df['_CLASSIFICACAO_AUTUADO']
    df_filtrado['_MOTIVO_CLASSIFICACAO'] = detalhes_df['_MOTIVO_CLASSIFICACAO']
    df_filtrado['_TERMO_IDENTIFICADO'] = detalhes_df['_TERMO_IDENTIFICADO']
    return df_filtrado[df_filtrado['_CLASSIFICACAO_AUTUADO'] == 'Pode cobrar'].copy()


# Função para normalizar e mesclar modais
def normalizar_e_mesclar_modais(modal_serasa, modal_divida):
    """
    Normaliza e mescla modais de ambas as bases, padronizando para maiúsculas.
    Se ambos existirem e forem diferentes, mescla com separador.
    """
    modal_serasa_str = str(modal_serasa).strip().upper() if pd.notna(modal_serasa) and str(modal_serasa).strip() != '' else None
    modal_divida_str = str(modal_divida).strip().upper() if pd.notna(modal_divida) and str(modal_divida).strip() != '' else None
    
    # Se ambos existirem
    if modal_serasa_str and modal_divida_str:
        # Se forem iguais, retorna apenas um
        if modal_serasa_str == modal_divida_str:
            return modal_serasa_str
        # Se forem diferentes, mescla com separador
        else:
            return f"{modal_serasa_str} / {modal_divida_str}"
    # Se apenas SERASA existir
    elif modal_serasa_str:
        return modal_serasa_str
    # Se apenas Dívida existir
    elif modal_divida_str:
        return modal_divida_str
    # Se nenhum existir
    else:
        return ''

# Função para formatar CPF/CNPJ no formato brasileiro
def formatar_cpf_cnpj_brasileiro(valor):
    """
    Formata CPF/CNPJ no formato brasileiro:
    - CPF: XXX.XXX.XXX-XX (11 dígitos)
    - CNPJ: XX.XXX.XXX/XXXX-XX (14 dígitos)
    """
    if pd.isna(valor) or valor == '' or valor is None:
        return ''
    
    # Remove caracteres não numéricos
    valor_str = str(valor).replace('.', '').replace('-', '').replace('/', '').strip()
    
    # Se estiver vazio, retorna vazio
    if not valor_str or not valor_str.isdigit():
        return str(valor)  # Retorna o valor original se não for numérico
    
    # Formatar CPF (11 dígitos)
    if len(valor_str) == 11:
        return f"{valor_str[0:3]}.{valor_str[3:6]}.{valor_str[6:9]}-{valor_str[9:11]}"
    
    # Formatar CNPJ (14 dígitos)
    elif len(valor_str) == 14:
        return f"{valor_str[0:2]}.{valor_str[2:5]}.{valor_str[5:8]}/{valor_str[8:12]}-{valor_str[12:14]}"
    
    # Se não tiver 11 ou 14 dígitos, retorna o valor original
    return str(valor)


def formatar_valor_br(valor):
    """Formata valor numérico no padrão brasileiro: R$ 1.234,56 ou R$ 0,00 para zero/nulo."""
    if valor is None or pd.isna(valor):
        return "R$ 0,00"
    try:
        v = float(valor)
        if v == 0:
            return "R$ 0,00"
        s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return "R$ " + s
    except (TypeError, ValueError):
        return "R$ 0,00"

# Regras de decadência:
# - Início do prazo: primeiro dia útil >= Data Infração (a própria data se for dia útil).
# - Autuação: 31 dias CORRIDOS a partir desse dia 1; data de notificação ajustada +4 dias.
# - Multa: 180 dias CORRIDOS a partir desse dia 1; data de notificação ajustada +4 dias.
PRAZO_DIAS_AUTUACAO = 31   # dias corridos
PRAZO_DIAS_MULTA = 181     # dias corridos

AJUSTE_DIAS_AUTUACAO = 4   # ajuste sobre a data de autuação (planilha → expedição)
AJUSTE_DIAS_MULTA = 4      # ajuste sobre a data de notificação de multa


def _easter_year(ano):
    """Retorna a data do domingo de Páscoa no ano dado (algoritmo de Butcher/Meeus)."""
    a = ano % 19
    b = ano // 100
    c = ano % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return date(ano, mes, dia)


def _feriados_nacionais_brasil(ano):
    """
    Retorna um set de datas (date) dos feriados nacionais do Brasil no ano.
    Usado para definir o primeiro dia útil a partir da data da infração (início do prazo de decadência).
    Referência: Lei 662/1949 e legislação federal vigente.
    """
    pascoa = _easter_year(ano)
    feriados = {
        date(ano, 1, 1),     # Confraternização Universal (Ano Novo)
        date(ano, 4, 21),   # Tiradentes
        date(ano, 5, 1),    # Dia do Trabalho
        date(ano, 9, 7),    # Independência do Brasil
        date(ano, 10, 12),  # Nossa Senhora Aparecida
        date(ano, 11, 2),   # Finados
        date(ano, 11, 15),  # Proclamação da República
        date(ano, 12, 25),  # Natal
    }
    # Feriados móveis (calculados a partir da Páscoa)
    feriados.add(pascoa - timedelta(days=2))   # Sexta-feira Santa (Paixão de Cristo)
    feriados.add(pascoa - timedelta(days=48))  # Segunda-feira de Carnaval
    feriados.add(pascoa - timedelta(days=47))  # Terça-feira de Carnaval
    feriados.add(pascoa + timedelta(days=60))  # Corpus Christi
    return feriados


def _resolver_coluna_data(df, nome_base):
    """Retorna o nome da coluna de data no DataFrame (com ou sem sufixo _serasa/_divida)."""
    if nome_base in df.columns:
        return nome_base
    if f"{nome_base}_serasa" in df.columns:
        return f"{nome_base}_serasa"
    if f"{nome_base}_divida" in df.columns:
        return f"{nome_base}_divida"
    return None

# Modais para os quais a decadência deve ser calculada
_MODAIS_COM_DECADENCIA = ['EXCESSO DE PESO', 'EVASAO DE PEDAGIO']

def _modal_tem_decadencia(modal_str):
    """Retorna True se o modal se enquadra nos tipos com cálculo de decadência
    (Excesso de Peso ou Evasão de Pedágio). Comparação sem sensibilidade a acentos."""
    if not modal_str or (isinstance(modal_str, float)):
        return False
    import unicodedata
    m = unicodedata.normalize('NFD', str(modal_str).upper().strip())
    m_sem_acento = ''.join(c for c in m if unicodedata.category(c) != 'Mn')
    return any(kw in m_sem_acento for kw in _MODAIS_COM_DECADENCIA)


def calcular_situacao_decadente(df, coluna_modal=None):
    """
    Calcula a coluna [Situação decadente] com base nas datas da SERASA:
    - Data Infração: referência. Início do prazo = primeiro dia útil >= Data Infração
      (a própria data se for dia útil; caso contrário, o próximo dia útil).
    - A partir desse dia 1, a contagem é em DIAS CORRIDOS (não dias úteis).
    - Para compensar o atraso entre emissão e expedição:
        * Autuação: soma-se AJUSTE_DIAS_AUTUACAO (4 dias) à Data Primeira Notificação Autuação.
        * Multa: soma-se AJUSTE_DIAS_MULTA (4 dias) à Data Primeira Notificação Multa.
    - Data Primeira Notificação Autuação (ajustada): não pode ultrapassar 31 dias corridos
      a partir do primeiro dia útil.
    - Data Primeira Notificação Multa (ajustada): não pode ultrapassar 181 dias corridos
      a partir do primeiro dia útil.
    Retorna Series: '' | 'Decadente autuação' | 'Decadente multa' | 'Decadente autuação e multa'.
    """
    col_infracao = _resolver_coluna_data(df, "Data Infração")
    col_notif_autuacao = _resolver_coluna_data(df, "Data Primeira Notificação Autuação")
    col_notif_multa = _resolver_coluna_data(df, "Data Primeira Notificação Multa")
    if not col_infracao or (not col_notif_autuacao and not col_notif_multa):
        return pd.Series([''] * len(df), index=df.index)
    data_infracao = pd.to_datetime(df[col_infracao], errors='coerce', dayfirst=True)
    data_notif_autuacao = pd.to_datetime(df[col_notif_autuacao], errors='coerce', dayfirst=True) if col_notif_autuacao else pd.Series([pd.NaT] * len(df), index=df.index)
    data_notif_multa = pd.to_datetime(df[col_notif_multa], errors='coerce', dayfirst=True) if col_notif_multa else pd.Series([pd.NaT] * len(df), index=df.index)

    # Construir conjunto de feriados para achar o primeiro dia útil a partir da infração
    series_todas_datas = pd.concat([
        data_infracao.dropna(),
        data_notif_autuacao.dropna(),
        data_notif_multa.dropna(),
    ]) if (data_notif_autuacao is not None and data_notif_multa is not None) else data_infracao.dropna()

    anos = set()
    for v in series_todas_datas:
        try:
            t = pd.Timestamp(v)
            if pd.notna(t):
                anos.add(t.year)
        except Exception:
            pass
    feriados = set()
    for a in anos:
        feriados |= _feriados_nacionais_brasil(a)
        feriados |= _feriados_nacionais_brasil(a + 1)

    def primeiro_dia_util_a_partir(dt):
        """Primeiro dia útil >= dt (considerando sábado/domingo/feriado)."""
        if pd.isna(dt):
            return pd.NaT
        d = dt.date()
        while True:
            if d.weekday() < 5 and d not in feriados:
                return pd.Timestamp(d)
            d += timedelta(days=1)

    # Início do prazo = primeiro dia útil a partir da Data Infração
    inicio_prazo = data_infracao.apply(primeiro_dia_util_a_partir)

    # Ajustar datas de notificação (compensar atraso entre emissão e expedição)
    data_notif_autuacao_ajustada = data_notif_autuacao + pd.to_timedelta(AJUSTE_DIAS_AUTUACAO, unit="D")
    data_notif_multa_ajustada = data_notif_multa + pd.to_timedelta(AJUSTE_DIAS_MULTA, unit="D")

    # Contagem em DIAS CORRIDOS a partir do primeiro dia útil
    dias_corridos_autuacao = (data_notif_autuacao_ajustada - inicio_prazo).dt.days
    dias_corridos_multa = (data_notif_multa_ajustada - inicio_prazo).dt.days

    decadente_autuacao = dias_corridos_autuacao > PRAZO_DIAS_AUTUACAO
    decadente_multa = dias_corridos_multa > PRAZO_DIAS_MULTA

    # A decadência de multa só é considerada para infrações a partir de 11/04/2021.
    # Antes dessa data, apenas a decadência de autuação pode ser marcada.
    data_corte_multa = pd.Timestamp('2021-04-11')
    mask_permite_multa = data_infracao >= data_corte_multa
    decadente_multa = decadente_multa & mask_permite_multa.fillna(False)
    situacao = pd.Series([''] * len(df), index=df.index, dtype=object)
    both_ = decadente_autuacao & decadente_multa
    only_aut = decadente_autuacao & ~decadente_multa
    only_multa = ~decadente_autuacao & decadente_multa
    situacao = situacao.mask(both_, 'Decadente autuação e multa').mask(only_aut, 'Decadente autuação').mask(only_multa, 'Decadente multa')

    # Aplicar filtro por modal: somente Excesso de Peso e Evasão de Pedágio têm decadência calculada
    col_modal_efetivo = coluna_modal if (coluna_modal and coluna_modal in df.columns) else ('Modais' if 'Modais' in df.columns else None)
    if col_modal_efetivo:
        mask_modal = df[col_modal_efetivo].apply(_modal_tem_decadencia)
        situacao = situacao.where(mask_modal, '')

    return situacao

# Função principal de análise - SEGUINDO A ORDEM DO SQL
def analisar_bases(df_serasa, df_divida, col_auto, col_cpf, col_valor, col_vencimento, coluna_protocolo, coluna_modal_serasa=None, coluna_modal_divida=None):
    resultados = {}
    
    # Verificar se a coluna de Auto de Infração existe (OBRIGATÓRIA)
    if col_auto not in df_serasa.columns:
        st.error(f"⚠️ Coluna '{col_auto}' não encontrada na base SERASA. Esta coluna é OBRIGATÓRIA!")
        return None
    
    if col_auto not in df_divida.columns:
        st.error(f"⚠️ Coluna '{col_auto}' não encontrada na base Dívida Ativa. Esta coluna é OBRIGATÓRIA!")
        return None
    
    # IMPORTANTE: Contar total de registros ANTES de remover nulos (para mostrar total exato da planilha)
    # Usar total_original se disponível (do carregamento), senão usar len atual
    total_registros_serasa_original = df_serasa.attrs.get('total_original', len(df_serasa))
    total_registros_divida_original = df_divida.attrs.get('total_original', len(df_divida))
    
    # Normalizar Auto de Infração (CHAVE PRINCIPAL)
    df_serasa['AUTO_NORM'] = df_serasa[col_auto].apply(normalizar_auto)
    df_divida['AUTO_NORM'] = df_divida[col_auto].apply(normalizar_auto)
    
    # Remover nulos de Auto de Infração (apenas para análises de correspondência)
    df_serasa_clean = df_serasa[df_serasa['AUTO_NORM'].notna()].copy()
    df_divida_clean = df_divida[df_divida['AUTO_NORM'].notna()].copy()
    
    # ==========================================
    # PASSO 1: JOIN - CRUZAR AS DUAS BASES (como no SQL: FROM a, b WHERE a.[Identificador do Débito] = b.[Identificador do Débito])
    # ==========================================
    # Fazer INNER JOIN usando Identificador do Débito (col_auto)
    df_joined = pd.merge(
        df_serasa_clean,
        df_divida_clean,
        on='AUTO_NORM',
        how='inner',
        suffixes=('_serasa', '_divida')
    )
    
    if df_joined.empty:
        st.warning("⚠️ Nenhum auto encontrado em ambas as bases!")
        return None
    
    # ==========================================
    # PASSO 2: CONVERTER VALORES (como no SQL: TRY_CONVERT com REPLACE)
    # ==========================================
    # A PARTIR DE AGORA:
    # - col_valor  -> coluna de valor da base SERASA (configurada na interface)
    # - coluna_valor_divida (session) -> coluna de valor da base Dívida Ativa
    from streamlit import session_state as _ss  # acesso leve apenas para pegar config
    col_valor_divida_cfg = _ss.get('coluna_valor_divida', col_valor)
    
    # Descobrir nomes das colunas após o merge (sufixos _serasa / _divida)
    col_valor_serasa = f"{col_valor}_serasa" if f"{col_valor}_serasa" in df_joined.columns else col_valor
    col_valor_divida = f"{col_valor_divida_cfg}_divida" if f"{col_valor_divida_cfg}_divida" in df_joined.columns else col_valor_divida_cfg
    
    # Usar SEMPRE a coluna de valor da SERASA como principal; se não existir, cai para Dívida
    col_valor_usar = col_valor_serasa if col_valor_serasa in df_joined.columns else col_valor_divida
    
    if col_valor_usar in df_joined.columns:
        # Converter valores usando a mesma lógica do SQL
        df_joined['Valor'] = df_joined[col_valor_usar].apply(converter_valor_sql)
    else:
        # Se não encontrar, tentar usar a coluna original
        if col_valor in df_joined.columns:
            df_joined['Valor'] = df_joined[col_valor].apply(converter_valor_sql)
        else:
            df_joined['Valor'] = None
    
    # ==========================================
    # PASSO 3: REMOVER VALORES ZERO (como no SQL: where Valor not like '0.00')
    # ==========================================
    # Remover valores zero (equivalente a: where Valor not like '0.00')
    df_sem_zero = df_joined[
        (df_joined['Valor'].notna()) & 
        (df_joined['Valor'] != 0.00) & 
        (df_joined['Valor'] > 0)
    ].copy()
    
    # ==========================================
    # PASSO 4: REMOVER DUPLICADOS (como no SQL: PARTITION BY [Nº do Processo])
    # ==========================================
    # Remover duplicados baseado em Nº do Processo, mantendo a DATA DE VENCIMENTO MAIS RECENTE
    # Identificar coluna de vencimento (pode ter sufixos após merge)
    col_vencimento_para_ordenacao = None
    if f"{col_vencimento}_serasa" in df_sem_zero.columns:
        col_vencimento_para_ordenacao = f"{col_vencimento}_serasa"
    elif f"{col_vencimento}_divida" in df_sem_zero.columns:
        col_vencimento_para_ordenacao = f"{col_vencimento}_divida"
    elif col_vencimento in df_sem_zero.columns:
        col_vencimento_para_ordenacao = col_vencimento
    
    if coluna_protocolo in df_sem_zero.columns:
        # Ordenar por data de vencimento (mais recente primeiro) antes de remover duplicados
        if col_vencimento_para_ordenacao and col_vencimento_para_ordenacao in df_sem_zero.columns:
            # Converter para datetime para ordenação correta
            df_sem_zero_temp = df_sem_zero.copy()
            df_sem_zero_temp['_VENCIMENTO_ORD'] = pd.to_datetime(
                df_sem_zero_temp[col_vencimento_para_ordenacao], 
                errors='coerce'
            )
            # Ordenar: mais recente primeiro (descendente), depois manter primeira ocorrência por protocolo
            df_sem_zero_temp = df_sem_zero_temp.sort_values(
                by='_VENCIMENTO_ORD', 
                ascending=False,  # Mais recente primeiro
                na_position='last'  # Datas inválidas vão para o final
            )
            # Remover duplicados mantendo a primeira (que será a mais recente)
            df_sem_duplicados = df_sem_zero_temp.drop_duplicates(
                subset=[coluna_protocolo], 
                keep='first'
            ).copy()
            # Remover coluna auxiliar
            df_sem_duplicados = df_sem_duplicados.drop(columns=['_VENCIMENTO_ORD'])
        else:
            # Se não tiver coluna de vencimento, usar lógica padrão
            df_sem_duplicados = df_sem_zero.drop_duplicates(subset=[coluna_protocolo], keep='first').copy()
    else:
        # Se não tiver Nº do Processo, usar Auto de Infração (com mesma lógica de data mais recente)
        if col_vencimento_para_ordenacao and col_vencimento_para_ordenacao in df_sem_zero.columns:
            df_sem_zero_temp = df_sem_zero.copy()
            df_sem_zero_temp['_VENCIMENTO_ORD'] = pd.to_datetime(
                df_sem_zero_temp[col_vencimento_para_ordenacao], 
                errors='coerce'
            )
            df_sem_zero_temp = df_sem_zero_temp.sort_values(
                by='_VENCIMENTO_ORD', 
                ascending=False,  # Mais recente primeiro
                na_position='last'
            )
            df_sem_duplicados = df_sem_zero_temp.drop_duplicates(
                subset=['AUTO_NORM'], 
                keep='first'
            ).copy()
            df_sem_duplicados = df_sem_duplicados.drop(columns=['_VENCIMENTO_ORD'])
        else:
            df_sem_duplicados = df_sem_zero.drop_duplicates(subset=['AUTO_NORM'], keep='first').copy()
    
    # ==========================================
    # PASSO 5: FILTRAR POR DATA (vencimento em 2025: 01/01/2025 a 31/12/2025)
    # ==========================================
    # Filtrar por data usando comparação de data (mais confiável que texto)
    col_vencimento_usar = f"{col_vencimento}_serasa" if f"{col_vencimento}_serasa" in df_sem_duplicados.columns else col_vencimento
    if col_vencimento_usar not in df_sem_duplicados.columns:
        col_vencimento_usar = f"{col_vencimento}_divida" if f"{col_vencimento}_divida" in df_sem_duplicados.columns else col_vencimento
    
    data_limite = pd.Timestamp('2025-01-01')
    data_limite_fim = pd.Timestamp('2025-12-31')
    
    if col_vencimento_usar in df_sem_duplicados.columns:
        try:
            # Criar cópia para não modificar o original
            df_temp = df_sem_duplicados.copy()
            
            # Converter para datetime com múltiplos formatos suportados
            if df_temp[col_vencimento_usar].dtype != 'datetime64[ns]':
                # Tentar converter para datetime com diferentes formatos
                df_temp['_VENCIMENTO_DT'] = pd.to_datetime(
                    df_temp[col_vencimento_usar], 
                    errors='coerce',
                    dayfirst=True,  # Suporta formato DD/MM/YYYY
                    infer_datetime_format=True
                )
            else:
                df_temp['_VENCIMENTO_DT'] = df_temp[col_vencimento_usar]
            
            # Filtrar apenas datas em 2025 (01/01/2025 a 31/12/2025, excluindo NaT)
            df_final = df_temp[
                (df_temp['_VENCIMENTO_DT'].notna()) & 
                (df_temp['_VENCIMENTO_DT'] >= data_limite) &
                (df_temp['_VENCIMENTO_DT'] <= data_limite_fim)
        ].copy()
            
            # Remover coluna auxiliar
            if '_VENCIMENTO_DT' in df_final.columns:
                df_final = df_final.drop(columns=['_VENCIMENTO_DT'])
        except Exception as e:
            # Se falhar, tentar método alternativo com comparação de texto (apenas ano 2025)
            try:
                df_temp = df_sem_duplicados.copy()
                # Converter para string para fazer comparação LIKE (fallback)
                df_temp[col_vencimento_usar] = df_temp[col_vencimento_usar].astype(str)
                # Filtrar apenas datas que contenham 2025 (excluindo 2026)
                df_final = df_temp[
                    (df_temp[col_vencimento_usar].str.contains('/2025', na=False)) |
                    (df_temp[col_vencimento_usar].str.contains('-2025', na=False)) |
                    (df_temp[col_vencimento_usar].str.contains('2025', na=False))
                ].copy()
                # Remover valores inválidos (NaT, nan, etc)
                df_final = df_final[
                    (~df_final[col_vencimento_usar].isin(['NaT', 'nan', 'None', '']))
                ].copy()
            except:
                # Se tudo falhar, usar todos os dados
                df_final = df_sem_duplicados.copy()
    else:
        # Se não encontrar coluna de vencimento, usar todos os dados
        df_final = df_sem_duplicados.copy()
    
    # Preparar dados para retorno (usar estrutura similar à anterior para compatibilidade)
    # Separar novamente em SERASA e Dívida para manter compatibilidade com o resto do código
    # Identificar autos únicos que estão no resultado final
    autos_em_ambas = set(df_final['AUTO_NORM'].unique())
    
    # Filtrar dados originais para manter estrutura
    df_serasa_filtrado = df_serasa_clean[df_serasa_clean['AUTO_NORM'].isin(autos_em_ambas)].copy()
    df_divida_filtrado = df_divida_clean[df_divida_clean['AUTO_NORM'].isin(autos_em_ambas)].copy()
    
    # Identificar autos únicos para estatísticas
    autos_serasa = set(df_serasa_clean['AUTO_NORM'].unique())
    autos_divida = set(df_divida_clean['AUTO_NORM'].unique())
    autos_apenas_serasa = autos_serasa - autos_divida
    autos_apenas_divida = autos_divida - autos_serasa
    
    # VALIDAÇÃO: Verificar se os autos normalizados correspondem exatamente
    # Garantir que estamos comparando os mesmos autos
    autos_serasa_filtrado = set(df_serasa_filtrado['AUTO_NORM'].unique())
    autos_divida_filtrado = set(df_divida_filtrado['AUTO_NORM'].unique())
    
    # Verificar se há diferença (não deveria haver)
    if autos_serasa_filtrado != autos_divida_filtrado:
        # Se houver diferença, usar apenas os que estão realmente em ambas
        autos_em_ambas_validados = autos_serasa_filtrado.intersection(autos_divida_filtrado)
        df_serasa_filtrado = df_serasa_filtrado[df_serasa_filtrado['AUTO_NORM'].isin(autos_em_ambas_validados)].copy()
        df_divida_filtrado = df_divida_filtrado[df_divida_filtrado['AUTO_NORM'].isin(autos_em_ambas_validados)].copy()
        autos_em_ambas = autos_em_ambas_validados
    
    # Normalizar CPF/CNPJ (para análises adicionais)
    if col_cpf in df_serasa_filtrado.columns:
        df_serasa_filtrado['CPF_CNPJ_NORM'] = df_serasa_filtrado[col_cpf].apply(normalizar_cpf_cnpj)
    else:
        df_serasa_filtrado['CPF_CNPJ_NORM'] = None
    
    if col_cpf in df_divida_filtrado.columns:
        df_divida_filtrado['CPF_CNPJ_NORM'] = df_divida_filtrado[col_cpf].apply(normalizar_cpf_cnpj)
    else:
        df_divida_filtrado['CPF_CNPJ_NORM'] = None
    
    # Extrair e mesclar modais de ambas as bases
    if coluna_modal_serasa and coluna_modal_divida:
        # Criar mapeamento de modais por AUTO_NORM (usar bases originais para ter todos os dados)
        modal_serasa_map = {}
        if coluna_modal_serasa in df_serasa_clean.columns:
            for idx, row in df_serasa_clean.iterrows():
                auto_norm = row.get('AUTO_NORM')
                if auto_norm:
                    modal_val = row.get(coluna_modal_serasa)
                    if auto_norm not in modal_serasa_map or pd.notna(modal_val):
                        modal_serasa_map[auto_norm] = modal_val
        elif coluna_modal_serasa in df_serasa_filtrado.columns:
            for idx, row in df_serasa_filtrado.iterrows():
                auto_norm = row.get('AUTO_NORM')
                if auto_norm:
                    modal_val = row.get(coluna_modal_serasa)
                    if auto_norm not in modal_serasa_map or pd.notna(modal_val):
                        modal_serasa_map[auto_norm] = modal_val
        
        modal_divida_map = {}
        if coluna_modal_divida in df_divida_clean.columns:
            for idx, row in df_divida_clean.iterrows():
                auto_norm = row.get('AUTO_NORM')
                if auto_norm:
                    modal_val = row.get(coluna_modal_divida)
                    if auto_norm not in modal_divida_map or pd.notna(modal_val):
                        modal_divida_map[auto_norm] = modal_val
        elif coluna_modal_divida in df_divida_filtrado.columns:
            for idx, row in df_divida_filtrado.iterrows():
                auto_norm = row.get('AUTO_NORM')
                if auto_norm:
                    modal_val = row.get(coluna_modal_divida)
                    if auto_norm not in modal_divida_map or pd.notna(modal_val):
                        modal_divida_map[auto_norm] = modal_val
        
        # Mesclar modais e adicionar coluna 'Modais' nos dataframes filtrados
        if 'AUTO_NORM' in df_serasa_filtrado.columns:
            df_serasa_filtrado['Modais'] = df_serasa_filtrado['AUTO_NORM'].apply(
                lambda auto: normalizar_e_mesclar_modais(
                    modal_serasa_map.get(auto),
                    modal_divida_map.get(auto)
                )
            )
        
        if 'AUTO_NORM' in df_divida_filtrado.columns:
            df_divida_filtrado['Modais'] = df_divida_filtrado['AUTO_NORM'].apply(
                lambda auto: normalizar_e_mesclar_modais(
                    modal_serasa_map.get(auto),
                    modal_divida_map.get(auto)
                )
            )
        
        # Adicionar coluna 'Modais' no df_final também
        if 'AUTO_NORM' in df_final.columns:
            df_final['Modais'] = df_final['AUTO_NORM'].apply(
                lambda auto: normalizar_e_mesclar_modais(
                    modal_serasa_map.get(auto),
                    modal_divida_map.get(auto)
                )
            )
    
    # Identificar CPF/CNPJ para análises adicionais
    cpf_serasa = set(df_serasa_filtrado[df_serasa_filtrado['CPF_CNPJ_NORM'].notna()]['CPF_CNPJ_NORM'].unique())
    cpf_divida = set(df_divida_filtrado[df_divida_filtrado['CPF_CNPJ_NORM'].notna()]['CPF_CNPJ_NORM'].unique())
    
    # Análise de correspondência por CPF/CNPJ (adicional)
    cpf_em_ambas = cpf_serasa.intersection(cpf_divida)
    cpf_apenas_serasa = cpf_serasa - cpf_divida
    cpf_apenas_divida = cpf_divida - cpf_serasa
    
    # IMPORTANTE: Calcular totais gerais (TODOS os autos com vencimento em 2025: 01/01/2025 a 31/12/2025)
    # Totais gerais da SERASA: pegar coluna "Data do Vencimento" e filtrar apenas ano 2025
    data_limite = pd.Timestamp('2025-01-01')
    data_limite_fim = pd.Timestamp('2025-12-31')
    if col_vencimento in df_serasa_clean.columns:
        try:
            # Criar cópia para não modificar o original
            df_serasa_temp = df_serasa_clean.copy()
            
            # Converter para datetime se ainda não estiver
            if df_serasa_temp[col_vencimento].dtype != 'datetime64[ns]':
                df_serasa_temp['_VENCIMENTO_DT'] = pd.to_datetime(
                    df_serasa_temp[col_vencimento], 
                    errors='coerce',
                    dayfirst=True,  # Suporta formato DD/MM/YYYY
                    infer_datetime_format=True
                )
            else:
                df_serasa_temp['_VENCIMENTO_DT'] = df_serasa_temp[col_vencimento]
            
            # Filtrar apenas datas em 2025 (01/01/2025 a 31/12/2025, excluindo NaT)
            df_serasa_total_2025 = df_serasa_temp[
                (df_serasa_temp['_VENCIMENTO_DT'].notna()) & 
                (df_serasa_temp['_VENCIMENTO_DT'] >= data_limite) &
                (df_serasa_temp['_VENCIMENTO_DT'] <= data_limite_fim)
            ].copy()
            
            # Remover coluna auxiliar
            if '_VENCIMENTO_DT' in df_serasa_total_2025.columns:
                df_serasa_total_2025 = df_serasa_total_2025.drop(columns=['_VENCIMENTO_DT'])
        except Exception as e:
            # Se falhar, tentar método alternativo (apenas ano 2025)
            try:
                df_serasa_temp = df_serasa_clean.copy()
                df_serasa_temp[col_vencimento] = df_serasa_temp[col_vencimento].astype(str)
                df_serasa_total_2025 = df_serasa_temp[
                    (df_serasa_temp[col_vencimento].str.contains('/2025', na=False)) |
                    (df_serasa_temp[col_vencimento].str.contains('-2025', na=False))
                ].copy()
                df_serasa_total_2025 = df_serasa_total_2025[
                    (~df_serasa_total_2025[col_vencimento].isin(['NaT', 'nan', 'None', '']))
                ].copy()
            except Exception:
                # Se ainda assim falhar, usar a base limpa sem filtro de data
                df_serasa_total_2025 = df_serasa_clean.copy()
    else:
        df_serasa_total_2025 = df_serasa_clean.copy()
    
    # Totais gerais da Dívida Ativa: pegar coluna "Data do Vencimento" e filtrar apenas ano 2025
    if col_vencimento in df_divida_clean.columns:
        try:
            # Criar cópia para não modificar o original
            df_divida_temp = df_divida_clean.copy()
            
            # Converter para datetime se ainda não estiver
            if df_divida_temp[col_vencimento].dtype != 'datetime64[ns]':
                df_divida_temp['_VENCIMENTO_DT'] = pd.to_datetime(
                    df_divida_temp[col_vencimento], 
                    errors='coerce',
                    dayfirst=True,  # Suporta formato DD/MM/YYYY
                    infer_datetime_format=True
                )
            else:
                df_divida_temp['_VENCIMENTO_DT'] = df_divida_temp[col_vencimento]
            
            # Filtrar apenas datas em 2025 (01/01/2025 a 31/12/2025, excluindo NaT)
            df_divida_total_2025 = df_divida_temp[
                (df_divida_temp['_VENCIMENTO_DT'].notna()) & 
                (df_divida_temp['_VENCIMENTO_DT'] >= data_limite) &
                (df_divida_temp['_VENCIMENTO_DT'] <= data_limite_fim)
            ].copy()
            
            # Remover coluna auxiliar
            if '_VENCIMENTO_DT' in df_divida_total_2025.columns:
                df_divida_total_2025 = df_divida_total_2025.drop(columns=['_VENCIMENTO_DT'])
        except Exception as e:
            # Se falhar, tentar método alternativo (apenas ano 2025)
            try:
                df_divida_temp = df_divida_clean.copy()
                df_divida_temp[col_vencimento] = df_divida_temp[col_vencimento].astype(str)
                df_divida_total_2025 = df_divida_temp[
                    (df_divida_temp[col_vencimento].str.contains('/2025', na=False)) |
                    (df_divida_temp[col_vencimento].str.contains('-2025', na=False))
                ].copy()
                df_divida_total_2025 = df_divida_total_2025[
                    (~df_divida_total_2025[col_vencimento].isin(['NaT', 'nan', 'None', '']))
                ].copy()
            except Exception:
                # Se ainda assim falhar, usar a base limpa sem filtro de data
                df_divida_total_2025 = df_divida_clean.copy()
    else:
        df_divida_total_2025 = df_divida_clean.copy()
    
    # IMPORTANTE: df_serasa_filtrado e df_divida_filtrado JÁ estão filtrados por vencimento em 2025
    # Não precisamos filtrar novamente, apenas garantir que as datas estão no formato correto
    df_serasa_2025 = df_serasa_filtrado.copy()
    df_divida_2025 = df_divida_filtrado.copy()
    
    # Garantir que as colunas de vencimento estão no formato datetime (já foram convertidas antes)
    if col_vencimento in df_serasa_2025.columns:
        if df_serasa_2025[col_vencimento].dtype != 'datetime64[ns]':
            df_serasa_2025[col_vencimento] = pd.to_datetime(df_serasa_2025[col_vencimento], errors='coerce')
    
    if col_vencimento in df_divida_2025.columns:
        if df_divida_2025[col_vencimento].dtype != 'datetime64[ns]':
            df_divida_2025[col_vencimento] = pd.to_datetime(df_divida_2025[col_vencimento], errors='coerce')
    
    # Agrupamento por CPF/CNPJ (ordenado do maior para menor)
    if col_valor in df_serasa_2025.columns:
        try:
            # Converter valores para numérico, ignorando valores inválidos (NaN)
            df_serasa_2025[col_valor] = pd.to_numeric(df_serasa_2025[col_valor], errors='coerce')
            # Remover linhas com valores NaN antes de agrupar (garantir exatidão)
            df_serasa_2025_valido = df_serasa_2025[df_serasa_2025[col_valor].notna()].copy()
            
            if not df_serasa_2025_valido.empty:
                # Agrupar e somar valores (NaN já foram removidos)
                agrupado_serasa = df_serasa_2025_valido.groupby('CPF_CNPJ_NORM').agg({
                    col_valor: ['sum', 'count']
                }).reset_index()
                agrupado_serasa.columns = ['CPF_CNPJ_NORM', 'VALOR_TOTAL', 'QTD_AUTOS']
                # Garantir que VALOR_TOTAL seja numérico e não tenha NaN
                agrupado_serasa['VALOR_TOTAL'] = pd.to_numeric(agrupado_serasa['VALOR_TOTAL'], errors='coerce').fillna(0)
                # Adicionar coluna CPF_CNPJ para exibição (formatação)
                agrupado_serasa['CPF_CNPJ'] = agrupado_serasa['CPF_CNPJ_NORM']
                agrupado_serasa = agrupado_serasa.sort_values('QTD_AUTOS', ascending=False)
            else:
                agrupado_serasa = pd.DataFrame()
        except Exception as e:
            agrupado_serasa = pd.DataFrame()
    else:
        agrupado_serasa = pd.DataFrame()
    
    # Separação por valores SERASA (apenas acima e abaixo de 1000)
    if col_valor in df_serasa_2025.columns:
        try:
            # Garantir que valores já foram convertidos (se não foram, converter agora)
            if df_serasa_2025[col_valor].dtype not in ['int64', 'float64']:
                df_serasa_2025[col_valor] = pd.to_numeric(df_serasa_2025[col_valor], errors='coerce')
            
            # PASSO 1: PRIMEIRO - Filtrar apenas valores válidos (não NaN) e maiores que zero (remover R$ 0,00)
            df_serasa_2025_sem_zero = df_serasa_2025[
                (df_serasa_2025[col_valor].notna()) & 
                (df_serasa_2025[col_valor] > 0)
            ].copy()
            
            # PASSO 2: DEPOIS - Remover duplicados baseado em Auto de Infração, mantendo DATA MAIS RECENTE
            # Identificar coluna de vencimento para ordenação
            col_vencimento_serasa = None
            if col_vencimento in df_serasa_2025_sem_zero.columns:
                col_vencimento_serasa = col_vencimento
            
            if col_auto in df_serasa_2025_sem_zero.columns:
                # Ordenar por data de vencimento (mais recente primeiro) antes de remover duplicados
                if col_vencimento_serasa and col_vencimento_serasa in df_serasa_2025_sem_zero.columns:
                    df_serasa_temp = df_serasa_2025_sem_zero.copy()
                    df_serasa_temp['_VENCIMENTO_ORD'] = pd.to_datetime(
                        df_serasa_temp[col_vencimento_serasa], 
                        errors='coerce'
                    )
                    # Ordenar: mais recente primeiro (descendente)
                    df_serasa_temp = df_serasa_temp.sort_values(
                        by='_VENCIMENTO_ORD', 
                        ascending=False,  # Mais recente primeiro
                        na_position='last'
                    )
                    # Remover duplicados mantendo a primeira (mais recente)
                    df_serasa_2025_valido = df_serasa_temp.drop_duplicates(
                        subset=[col_auto], 
                        keep='first'
                    ).copy()
                    df_serasa_2025_valido = df_serasa_2025_valido.drop(columns=['_VENCIMENTO_ORD'])
                else:
                    # Se não tiver coluna de vencimento, usar lógica padrão
                    df_serasa_2025_valido = df_serasa_2025_sem_zero.drop_duplicates(subset=[col_auto], keep='first').copy()
            else:
                df_serasa_2025_valido = df_serasa_2025_sem_zero.copy()
            
            # Individual - Autos com valor > 0 e <= R$ 999,99 (até 999,99, incluindo exatamente 999,99)
            serasa_abaixo_1000_ind = df_serasa_2025_valido[df_serasa_2025_valido[col_valor] <= 999.99].copy()
            # Individual - Autos com valor >= R$ 1.000,00 (maior ou igual a 1000, inclui exatamente 1000)
            serasa_acima_1000_ind = df_serasa_2025_valido[df_serasa_2025_valido[col_valor] >= 1000].copy()
            
            # VALIDAÇÃO: Verificar se há valores entre 999,99 e 1000 (excluindo 1000 - não devem ser contados em nenhum grupo)
            # Nota: Valores >= 1000 agora são incluídos no grupo "acima_1000"
            valores_entre_999_1000 = df_serasa_2025_valido[
                (df_serasa_2025_valido[col_valor] > 999.99) & 
                (df_serasa_2025_valido[col_valor] < 1000)
            ].copy()
            
            # VALIDAÇÃO: Verificar se a soma dos grupos + valores entre 999-1000 = total válido
            total_ate_999 = len(serasa_abaixo_1000_ind)
            total_acima_1000 = len(serasa_acima_1000_ind)
            total_entre_999_1000 = len(valores_entre_999_1000)
            total_valido_esperado = len(df_serasa_2025_valido)
            total_calculado = total_ate_999 + total_acima_1000 + total_entre_999_1000
            
            # VALIDAÇÃO: Verificar se ainda há duplicatas após remoção (apenas para informação)
            # Duplicados já foram removidos no PASSO 2, mas verificamos se há outros tipos de duplicatas
            if col_auto in df_serasa_2025_valido.columns:
                tem_duplicatas_restantes = len(df_serasa_2025_valido) != len(df_serasa_2025_valido.drop_duplicates(subset=[col_auto]))
                if tem_duplicatas_restantes:
                    # Se ainda houver duplicatas, remover novamente mantendo data mais recente (garantir exatidão)
                    if col_vencimento_serasa and col_vencimento_serasa in df_serasa_2025_valido.columns:
                        df_serasa_temp = df_serasa_2025_valido.copy()
                        df_serasa_temp['_VENCIMENTO_ORD'] = pd.to_datetime(
                            df_serasa_temp[col_vencimento_serasa], 
                            errors='coerce'
                        )
                        df_serasa_temp = df_serasa_temp.sort_values(
                            by='_VENCIMENTO_ORD', 
                            ascending=False,  # Mais recente primeiro
                            na_position='last'
                        )
                        df_serasa_2025_valido = df_serasa_temp.drop_duplicates(
                            subset=[col_auto], 
                            keep='first'
                        ).copy()
                        df_serasa_2025_valido = df_serasa_2025_valido.drop(columns=['_VENCIMENTO_ORD'])
                    else:
                        df_serasa_2025_valido = df_serasa_2025_valido.drop_duplicates(subset=[col_auto], keep='first').copy()
                    # Recalcular as separações por valor após remoção de duplicados
                    serasa_abaixo_1000_ind = df_serasa_2025_valido[df_serasa_2025_valido[col_valor] <= 999.99].copy()
                    serasa_acima_1000_ind = df_serasa_2025_valido[df_serasa_2025_valido[col_valor] >= 1000].copy()
            
            # Recalcular agrupado_serasa a partir da base limpa (sem zeros, sem duplicados) para números exatos
            if not df_serasa_2025_valido.empty and col_valor in df_serasa_2025_valido.columns and 'CPF_CNPJ_NORM' in df_serasa_2025_valido.columns:
                agrupado_serasa = df_serasa_2025_valido.groupby('CPF_CNPJ_NORM').agg({
                    col_valor: ['sum', 'count']
                }).reset_index()
                agrupado_serasa.columns = ['CPF_CNPJ_NORM', 'VALOR_TOTAL', 'QTD_AUTOS']
                agrupado_serasa['VALOR_TOTAL'] = pd.to_numeric(agrupado_serasa['VALOR_TOTAL'], errors='coerce').fillna(0)
                agrupado_serasa['CPF_CNPJ'] = agrupado_serasa['CPF_CNPJ_NORM']
                agrupado_serasa = agrupado_serasa.sort_values('QTD_AUTOS', ascending=False)
            
            # Acumulativo - CPF/CNPJ com soma < R$ 1.000 (até 999,99 - nem um centavo a mais)
            # Valor exatamente 1000 entra apenas no grupo "acima de 1000"
            if not agrupado_serasa.empty:
                serasa_abaixo_1000_acum = agrupado_serasa[agrupado_serasa['VALOR_TOTAL'] < 1000].copy()
                # Buscar os autos correspondentes aos CPF/CNPJ acumulados ≤ 1000
                # Usar df_serasa_2025_valido para garantir consistência (já sem valores zero e duplicados)
                cpf_abaixo_1000_acum = set(serasa_abaixo_1000_acum['CPF_CNPJ_NORM'].unique())
                serasa_abaixo_1000_acum_autos = df_serasa_2025_valido[df_serasa_2025_valido['CPF_CNPJ_NORM'].isin(cpf_abaixo_1000_acum)].copy()
            else:
                serasa_abaixo_1000_acum = pd.DataFrame()
                serasa_abaixo_1000_acum_autos = pd.DataFrame()
            
            # Acumulativo - CPF/CNPJ com soma >= R$ 1.000,00
            if not agrupado_serasa.empty:
                serasa_acima_1000_acum = agrupado_serasa[agrupado_serasa['VALOR_TOTAL'] >= 1000].copy()
                # Buscar os autos correspondentes aos CPF/CNPJ acumulados >= 1000
                # Usar df_serasa_2025_valido para garantir consistência (já sem valores zero e duplicados)
                cpf_acima_1000_acum = set(serasa_acima_1000_acum['CPF_CNPJ_NORM'].unique())
                serasa_acima_1000_acum_autos = df_serasa_2025_valido[df_serasa_2025_valido['CPF_CNPJ_NORM'].isin(cpf_acima_1000_acum)].copy()
            else:
                serasa_acima_1000_acum = pd.DataFrame()
                serasa_acima_1000_acum_autos = pd.DataFrame()
        except Exception as e:
            serasa_abaixo_1000_ind = pd.DataFrame()
            serasa_acima_1000_ind = pd.DataFrame()
            serasa_abaixo_1000_acum = pd.DataFrame()
            serasa_acima_1000_acum = pd.DataFrame()
            serasa_abaixo_1000_acum_autos = pd.DataFrame()
            serasa_acima_1000_acum_autos = pd.DataFrame()
    else:
        serasa_abaixo_1000_ind = pd.DataFrame()
        serasa_acima_1000_ind = pd.DataFrame()
        serasa_abaixo_1000_acum = pd.DataFrame()
        serasa_acima_1000_acum = pd.DataFrame()
        serasa_abaixo_1000_acum_autos = pd.DataFrame()
        serasa_acima_1000_acum_autos = pd.DataFrame()
    
    # Dataframes de divergências por AUTO DE INFRAÇÃO (PRINCIPAL)
    df_autos_apenas_serasa = df_serasa_clean[df_serasa_clean['AUTO_NORM'].isin(autos_apenas_serasa)].copy()
    df_autos_apenas_divida = df_divida_clean[df_divida_clean['AUTO_NORM'].isin(autos_apenas_divida)].copy()
    
    # Dataframes de divergências por CPF/CNPJ (adicional)
    df_cpf_apenas_serasa = df_serasa_filtrado[df_serasa_filtrado['CPF_CNPJ_NORM'].notna() & df_serasa_filtrado['CPF_CNPJ_NORM'].isin(cpf_apenas_serasa)].copy() if col_cpf in df_serasa_filtrado.columns else pd.DataFrame()
    df_cpf_apenas_divida = df_divida_filtrado[df_divida_filtrado['CPF_CNPJ_NORM'].notna() & df_divida_filtrado['CPF_CNPJ_NORM'].isin(cpf_apenas_divida)].copy() if col_cpf in df_divida_filtrado.columns else pd.DataFrame()
    
    # Calcular contagens de autos por faixa de valor (para dashboard)
    # NOVA LÓGICA: Usar análise ACUMULATIVA (soma por CPF/CNPJ) - igual à exportação
    # Os dataframes acumulativos já usam df_serasa_2025_valido (sem valores zero e duplicados)
    try:
        # Usar lógica acumulativa: contar autos cujo CNPJ tem soma total < 1000 (até 999,99; exatamente 1000 vai para acima_1000)
        if not serasa_abaixo_1000_acum_autos.empty:
            qtd_autos_ate_999 = len(serasa_abaixo_1000_acum_autos)
        else:
            qtd_autos_ate_999 = 0
    except Exception as e:
        qtd_autos_ate_999 = 0
    
    try:
        # Usar lógica acumulativa: contar autos cujo CNPJ tem soma total >= 1000
        if not serasa_acima_1000_acum_autos.empty:
            qtd_autos_acima_1000 = len(serasa_acima_1000_acum_autos)
        else:
            qtd_autos_acima_1000 = 0
    except Exception as e:
        qtd_autos_acima_1000 = 0
    
    # Calcular autos_em_ambas como LINHAS válidas (consistente com SQL)
    # IMPORTANTE: df_final já passou por todos os passos do SQL:
    # 1. JOIN, 2. Converter valores, 3. Remover zeros, 4. Remover duplicados, 5. Filtrar por data
    # Então df_final já contém o resultado final equivalente ao SQL
    autos_em_ambas_linhas = len(df_final)
    
    # ==========================================
    # ANÁLISE GERAL (SEM FILTRO DE VENCIMENTO) - Para comparação geral
    # ==========================================
    # Usar df_sem_duplicados (já passou por: JOIN → Valores → Zeros → Duplicados, mas SEM filtro de data)
    # Isso permite comparar todos os autos, independente da data de vencimento
    df_final_geral = df_sem_duplicados.copy()  # Sem filtro de data
    autos_em_ambas_geral = set(df_final_geral['AUTO_NORM'].unique())
    autos_em_ambas_geral_linhas = len(df_final_geral)
    
    # Adicionar modais e CPF_CNPJ_NORM ao df_final_geral (se modais foram processados)
    if coluna_modal_serasa and coluna_modal_divida:
        # Criar mapeamentos de modais para df_final_geral
        modal_serasa_map_geral = {}
        if coluna_modal_serasa in df_serasa_clean.columns:
            for idx, row in df_serasa_clean.iterrows():
                auto_norm = row.get('AUTO_NORM')
                if auto_norm:
                    modal_val = row.get(coluna_modal_serasa)
                    if auto_norm not in modal_serasa_map_geral or pd.notna(modal_val):
                        modal_serasa_map_geral[auto_norm] = modal_val
        
        modal_divida_map_geral = {}
        if coluna_modal_divida in df_divida_clean.columns:
            for idx, row in df_divida_clean.iterrows():
                auto_norm = row.get('AUTO_NORM')
                if auto_norm:
                    modal_val = row.get(coluna_modal_divida)
                    if auto_norm not in modal_divida_map_geral or pd.notna(modal_val):
                        modal_divida_map_geral[auto_norm] = modal_val
        
        # Adicionar coluna 'Modais' no df_final_geral
        if 'AUTO_NORM' in df_final_geral.columns:
            df_final_geral['Modais'] = df_final_geral['AUTO_NORM'].apply(
                lambda auto: normalizar_e_mesclar_modais(
                    modal_serasa_map_geral.get(auto),
                    modal_divida_map_geral.get(auto)
                )
            )
    
    # Adicionar CPF_CNPJ_NORM no df_final_geral para compatibilidade
    col_cpf_serasa_geral = f"{col_cpf}_serasa" if f"{col_cpf}_serasa" in df_final_geral.columns else col_cpf
    if col_cpf_serasa_geral in df_final_geral.columns:
        df_final_geral['CPF_CNPJ_NORM'] = df_final_geral[col_cpf_serasa_geral].apply(normalizar_cpf_cnpj)
    elif col_cpf in df_final_geral.columns:
        df_final_geral['CPF_CNPJ_NORM'] = df_final_geral[col_cpf].apply(normalizar_cpf_cnpj)
    else:
        df_final_geral['CPF_CNPJ_NORM'] = None
    
    # Regra de decadência (após comparação): Data Infração + 37 dias (notificação autuação) e 187 dias (notificação multa)
    df_final['Situação decadente'] = calcular_situacao_decadente(df_final)
    df_final_geral['Situação decadente'] = calcular_situacao_decadente(df_final_geral)
    
    resultados = {
        'df_serasa_original': df_serasa,
        'df_divida_original': df_divida,
        'df_serasa_filtrado': df_serasa_2025,  # Autos em ambas as bases com vencimento em 2025
        'df_divida_filtrado': df_divida_2025,  # Autos em ambas as bases com vencimento em 2025
        'df_final_sql': df_final,  # DataFrame final seguindo a ordem do SQL (para exportação)
        'df_final_geral': df_final_geral,  # DataFrame geral SEM filtro de vencimento (para comparação geral)
        'df_serasa_total_2025': df_serasa_total_2025,  # TODOS os autos SERASA com vencimento em 2025
        'df_divida_total_2025': df_divida_total_2025,  # TODOS os autos Dívida Ativa com vencimento em 2025
        # Estatísticas por AUTO DE INFRAÇÃO (PRINCIPAL - COM filtro de vencimento em 2025)
        'autos_em_ambas': autos_em_ambas_linhas,  # Número de LINHAS válidas (vencimento em 2025 e valor > 0)
        'autos_em_ambas_unicos': len(autos_em_ambas),  # Número de AUTOS ÚNICOS em ambas as bases (com filtro em 2025)
        # Estatísticas GERAIS (SEM filtro de vencimento)
        'autos_em_ambas_geral': autos_em_ambas_geral_linhas,  # Número de LINHAS válidas (TODOS os autos, sem filtro de data)
        'autos_em_ambas_geral_unicos': len(autos_em_ambas_geral),  # Número de AUTOS ÚNICOS em ambas as bases (sem filtro de data)
        'autos_apenas_serasa': len(autos_apenas_serasa),
        'autos_apenas_divida': len(autos_apenas_divida),
        # Totais de registros (TODOS os registros da planilha, incluindo os sem Auto de Infração)
        'total_registros_serasa': total_registros_serasa_original,  # Total exato de registros da planilha SERASA
        'total_registros_divida': total_registros_divida_original,  # Total exato de registros da planilha Dívida Ativa
        # Totais de autos únicos (apenas os que têm Auto de Infração válido)
        'total_autos_serasa': len(autos_serasa),  # Autos únicos com Auto de Infração válido
        'total_autos_divida': len(autos_divida),  # Autos únicos com Auto de Infração válido
        'df_autos_apenas_serasa': df_autos_apenas_serasa,
        'df_autos_apenas_divida': df_autos_apenas_divida,
        # Estatísticas por CPF/CNPJ (adicional)
        'cpf_em_ambas': len(cpf_em_ambas),
        'cpf_apenas_serasa': len(cpf_apenas_serasa),
        'cpf_apenas_divida': len(cpf_apenas_divida),
        'total_cpf_serasa': len(cpf_serasa),
        'total_cpf_divida': len(cpf_divida),
        'df_cpf_apenas_serasa': df_cpf_apenas_serasa,
        'df_cpf_apenas_divida': df_cpf_apenas_divida,
        # Dados processados
        'agrupado_serasa': agrupado_serasa,
        'serasa_abaixo_1000_ind': serasa_abaixo_1000_ind,
        'serasa_acima_1000_ind': serasa_acima_1000_ind,
        'serasa_abaixo_1000_acum': serasa_abaixo_1000_acum,
        'serasa_acima_1000_acum': serasa_acima_1000_acum,
        'serasa_abaixo_1000_acum_autos': serasa_abaixo_1000_acum_autos,
        'serasa_acima_1000_acum_autos': serasa_acima_1000_acum_autos,
        # Contagens por faixa de valor (para dashboard)
        'qtd_autos_ate_999': qtd_autos_ate_999,
        'qtd_autos_acima_1000': qtd_autos_acima_1000,
        # Listas para referência
        'autos_em_ambas_lista': autos_em_ambas,
        'autos_apenas_serasa_lista': autos_apenas_serasa,
        'autos_apenas_divida_lista': autos_apenas_divida
    }
    
    return resultados

# Main
if arquivo_serasa and arquivo_divida:
    st.markdown("---")
    
    # Carregar dados
    with st.spinner("Carregando bases de dados..."):
        df_serasa = carregar_dados(arquivo_serasa, "SERASA")
        df_divida = carregar_dados(arquivo_divida, "Dívida Ativa")
    
    if df_serasa is not None and df_divida is not None:
        # Mostrar preview
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📋 Preview - Base SERASA")
            st.dataframe(df_serasa.head(), use_container_width=True)
            st.caption(f"Total de registros: {len(df_serasa)}")
            st.caption(f"Colunas: {', '.join(df_serasa.columns.tolist()[:5])}...")
        
        with col2:
            st.markdown("### 📋 Preview - Base Dívida Ativa")
            st.dataframe(df_divida.head(), use_container_width=True)
            st.caption(f"Total de registros: {len(df_divida)}")
            st.caption(f"Colunas: {', '.join(df_divida.columns.tolist()[:5])}...")
        
        st.markdown("---")
        
        # Botão de análise
        if st.button("🚀 Executar Análise Completa", type="primary", use_container_width=True):
            if not coluna_auto:
                st.error("⚠️ Por favor, informe o nome da coluna de Auto de Infração!")
            else:
                with st.spinner("Analisando bases de dados por Auto de Infração... Isso pode levar alguns instantes."):
                    resultados = analisar_bases(
                        df_serasa, 
                        df_divida, 
                        coluna_auto,
                        coluna_cpf_cnpj, 
                        coluna_valor, 
                        coluna_vencimento,
                        coluna_protocolo,
                        coluna_modal_serasa,
                        coluna_modal_divida
                    )
            
            if resultados:
                st.session_state['resultados'] = resultados
                # Armazenar configurações de colunas
                st.session_state['coluna_auto'] = coluna_auto
                st.session_state['coluna_cpf_cnpj'] = coluna_cpf_cnpj
                # Valor principal (SERASA)
                st.session_state['coluna_valor'] = coluna_valor
                # Valor da base Dívida Ativa
                st.session_state['coluna_valor_divida'] = coluna_valor_divida
                st.session_state['coluna_vencimento'] = coluna_vencimento
                st.session_state['coluna_protocolo'] = coluna_protocolo
                st.session_state['coluna_modal_serasa'] = coluna_modal_serasa
                st.session_state['coluna_modal_divida'] = coluna_modal_divida
                st.session_state['coluna_nome_autuado'] = coluna_nome_autuado
                st.session_state['export_run_id'] = datetime.now().strftime('%Y%m%d%H%M%S%f')
                st.session_state['export_run_label'] = datetime.now().strftime('%d %m %Y %H:%M')
                st.success("✅ Análise concluída com sucesso!")
                st.rerun()

# Exibir resultados
if 'resultados' in st.session_state:
    resultados = st.session_state['resultados']
    # Recuperar configurações de colunas
    coluna_auto = st.session_state.get('coluna_auto', 'Identificador do Débito')
    coluna_cpf_cnpj = st.session_state.get('coluna_cpf_cnpj', 'CPF/CNPJ')
    # Valor principal (SERASA)
    coluna_valor = st.session_state.get('coluna_valor', 'Valor Multa Atualizado')
    # Valor da base Dívida Ativa
    coluna_valor_divida = st.session_state.get('coluna_valor_divida', 'Valor Atualizado do Débito')
    coluna_vencimento = st.session_state.get('coluna_vencimento', 'Data do Vencimento')
    coluna_protocolo = st.session_state.get('coluna_protocolo', 'Nº do Processo')
    coluna_modal_serasa = st.session_state.get('coluna_modal_serasa', 'Tipo Modal')
    coluna_modal_divida = st.session_state.get('coluna_modal_divida', 'Subtipo de Débito')
    coluna_nome_autuado = st.session_state.get('coluna_nome_autuado', 'Nome Autuado')
    
    st.markdown("---")
    st.markdown("## 📊 Dashboard de Resultados - Análise por Auto de Infração")
    
    st.markdown("### 🔑 Análise Principal: Autos de Infração")
    # Métricas principais - AUTOS DE INFRAÇÃO
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Total Registros SERASA",
            f"{resultados['total_registros_serasa']:,}",
            delta=f"{resultados['total_autos_serasa']:,} autos únicos válidos"
        )
        st.caption(f"📋 {len(resultados['df_serasa_filtrado']):,} registros após filtros (vencimento em 2025: 01/01 a 31/12)")
    
    with col2:
        st.metric(
            "Total Registros Dívida Ativa",
            f"{resultados['total_registros_divida']:,}",
            delta=f"{resultados['total_autos_divida']:,} autos únicos válidos"
        )
        st.caption(f"📋 {len(resultados['df_divida_filtrado']):,} registros após filtros (vencimento em 2025: 01/01 a 31/12)")
    
    with col3:
        # Mostrar número de linhas (registros) que estão em ambas as bases
        # Isso corresponde exatamente aos dados que aparecem na exportação
        st.metric(
            "Autos em Ambas as Bases",
            f"{resultados['autos_em_ambas']:,}",
            delta=f"{resultados.get('autos_em_ambas_unicos', 0):,} autos únicos"
        )
    
    with col4:
        taxa_match_autos = (resultados['autos_em_ambas'] / max(resultados['total_autos_serasa'], 1)) * 100
        st.metric(
            "Taxa de Correspondência",
            f"{taxa_match_autos:.1f}%",
            delta="Autos SERASA vs Dívida Ativa"
        )
    
    # Métricas adicionais - Autos por faixa de valor
    st.markdown("#### 💰 Autos por Faixa de Valor (SERASA - Vencimento em 2025)")
    col5, col6 = st.columns(2)
    
    with col5:
        st.metric(
            "Autos até R$ 999,99",
            f"{resultados.get('qtd_autos_ate_999', 0):,}",
            delta="Soma acumulativa por CNPJ"
        )
    
    with col6:
        st.metric(
            "Autos acima de R$ 1.000,00",
            f"{resultados.get('qtd_autos_acima_1000', 0):,}",
            delta="Soma acumulativa por CNPJ"
        )
    
    # NOVA SEÇÃO: Comparação Geral (SEM filtro de vencimento)
    st.markdown("---")
    st.markdown("#### 📊 Comparação Geral - Todos os Autos (Sem Filtro de Vencimento)")
    st.info("💡 Esta seção mostra a comparação de TODOS os autos em ambas as bases, **independente da data de vencimento**. Mantém todos os outros filtros (valores > 0, remoção de duplicados, etc.).")
    
    # Calcular métricas gerais
    autos_geral = resultados.get('autos_em_ambas_geral', 0)
    autos_geral_unicos = resultados.get('autos_em_ambas_geral_unicos', 0)
    
    # Calcular valores totais gerais
    df_final_geral = resultados.get('df_final_geral', pd.DataFrame())
    valor_total_geral = 0.0
    qtd_autos_geral_valor = 0
    if not df_final_geral.empty and 'Valor' in df_final_geral.columns:
        valores_validos_geral = df_final_geral['Valor'][df_final_geral['Valor'].notna() & (df_final_geral['Valor'] > 0)]
        valor_total_geral = float(valores_validos_geral.sum()) if len(valores_validos_geral) > 0 else 0.0
        qtd_autos_geral_valor = len(valores_validos_geral)
    
    col_geral1, col_geral2, col_geral3, col_geral4 = st.columns(4)
    
    with col_geral1:
        st.metric(
            "Autos em Ambas (Geral)",
            f"{autos_geral:,}",
            delta=f"{autos_geral_unicos:,} autos únicos"
        )
        st.caption("📋 Todos os autos (sem filtro de data)")
    
    with col_geral2:
        st.metric(
            "Valor Total (Geral)",
            f"R$ {valor_total_geral:,.2f}",
            delta=f"{qtd_autos_geral_valor:,} autos"
        )
        st.caption("💰 Soma de todos os valores")
    
    with col_geral3:
        # Comparar com a análise filtrada
        autos_filtrado = resultados.get('autos_em_ambas', 0)
        diferenca_autos = autos_geral - autos_filtrado
        st.metric(
            "Diferença (Geral vs Filtrado)",
            f"{diferenca_autos:,} autos",
            delta="Geral - Filtrado (vencimento em 2025)"
        )
        st.caption("📊 Diferença entre geral e filtrado")
    
    with col_geral4:
        # Taxa de cobertura
        if autos_geral > 0:
            taxa_cobertura = (autos_filtrado / autos_geral) * 100
            st.metric(
                "Cobertura 2025+",
                f"{taxa_cobertura:.1f}%",
                delta=f"{autos_filtrado:,} de {autos_geral:,}"
            )
            st.caption("📈 % dos autos com vencimento em 2025")
        else:
            st.metric("Cobertura 2025+", "N/A")
    
    # Gráfico comparativo Geral vs Filtrado
    st.markdown("---")
    st.markdown("##### 📊 Comparação Visual: Geral vs Filtrado (vencimento em 2025)")
    fig_comparacao = go.Figure(data=[
        go.Bar(name='Todos os Autos (Geral)', x=['Comparação'], y=[autos_geral], marker_color='#3498db', text=f"{autos_geral:,}", textposition='auto'),
        go.Bar(name='Autos em 2025 (Filtrado)', x=['Comparação'], y=[autos_filtrado], marker_color='#2ecc71', text=f"{autos_filtrado:,}", textposition='auto')
    ])
    fig_comparacao.update_layout(
        title="Comparação: Todos os Autos vs Autos com Vencimento em 2025",
        xaxis_title="",
        yaxis_title="Quantidade de Autos",
        barmode='group',
        height=400,
        showlegend=True
    )
    st.plotly_chart(fig_comparacao, use_container_width=True)
    st.caption(f"💡 Comparação entre todos os autos ({autos_geral:,}) e apenas os com vencimento em 2025 ({autos_filtrado:,})")
    
    # VALIDAÇÃO: Seção de verificação de exatidão
    st.markdown("---")
    st.markdown("#### ✅ Validação de Exatidão dos Dados")
    st.info("💡 Esta seção valida se os cálculos estão 100% corretos conforme os dados da planilha.")
    
    # Calcular totais para validação
    total_autos_em_ambas = resultados.get('autos_em_ambas', 0)
    total_ate_999 = resultados.get('qtd_autos_ate_999', 0)
    total_acima_1000 = resultados.get('qtd_autos_acima_1000', 0)
    soma_grupos = total_ate_999 + total_acima_1000
    diferenca = total_autos_em_ambas - soma_grupos
    
    col_val1, col_val2, col_val3 = st.columns(3)
    
    with col_val1:
        st.metric(
            "Total Autos em Ambas",
            f"{total_autos_em_ambas:,}",
            delta="Linhas válidas (vencimento em 2025 e valor > 0)"
        )
    
    with col_val2:
        st.metric(
            "Soma dos Grupos",
            f"{soma_grupos:,}",
            delta=f"Até 999,99: {total_ate_999:,} + Acima 1000: {total_acima_1000:,}"
        )
    
    with col_val3:
        if diferenca == 0:
            st.metric(
                "✅ Validação",
                "CORRETO",
                delta="Soma dos grupos = Total"
            )
        else:
            st.metric(
                "⚠️ Validação",
                f"Diferença: {diferenca:,}",
                delta="Valores entre 999 e 1000 (não contados nos grupos)"
            )
            st.caption(f"💡 {diferenca:,} autos têm valores entre R$ 999,99 e R$ 1.000,00 (não incluídos nos grupos filtrados)")
    
    # Gráficos - AUTOS DE INFRAÇÃO
    col1, col2 = st.columns(2)
    
    with col1:
        # Gráfico de correspondência de autos
        fig_corresp = go.Figure(data=[
            go.Bar(name='Em Ambas', x=['Autos de Infração'], y=[resultados['autos_em_ambas']], marker_color='#2ecc71'),
            go.Bar(name='Apenas SERASA', x=['Autos de Infração'], y=[resultados['autos_apenas_serasa']], marker_color='#3498db'),
            go.Bar(name='Apenas Dívida Ativa', x=['Autos de Infração'], y=[resultados['autos_apenas_divida']], marker_color='#e74c3c')
        ])
        fig_corresp.update_layout(
            title='Análise de Correspondência de Autos de Infração',
            barmode='group',
            height=400,
            showlegend=True
        )
        st.plotly_chart(fig_corresp, use_container_width=True)
    
    with col2:
        # Gráfico de pizza de autos
        labels = ['Em Ambas', 'Apenas SERASA', 'Apenas Dívida Ativa']
        values = [resultados['autos_em_ambas'], resultados['autos_apenas_serasa'], resultados['autos_apenas_divida']]
        colors = ['#2ecc71', '#3498db', '#e74c3c']
        
        fig_pizza = go.Figure(data=[go.Pie(
            labels=labels,
            values=values,
            hole=0.4,
            marker_colors=colors
        )])
        fig_pizza.update_layout(
            title='Distribuição de Autos de Infração',
            height=400
        )
        st.plotly_chart(fig_pizza, use_container_width=True)
    
    # Métricas adicionais - CPF/CNPJ
    st.markdown("### 📋 Análise Adicional: CPF/CNPJ")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "CPF/CNPJ SERASA",
            f"{resultados['total_cpf_serasa']:,}",
            delta="CPF/CNPJ únicos"
        )
    
    with col2:
        st.metric(
            "CPF/CNPJ Dívida Ativa",
            f"{resultados['total_cpf_divida']:,}",
            delta="CPF/CNPJ únicos"
        )
    
    with col3:
        st.metric(
            "CPF/CNPJ em Ambas",
            f"{resultados['cpf_em_ambas']:,}",
            delta="CPF/CNPJ correspondentes"
        )
    
    with col4:
        if resultados['total_cpf_serasa'] > 0:
            taxa_match_cpf = (resultados['cpf_em_ambas'] / resultados['total_cpf_serasa']) * 100
            st.metric(
                "Taxa CPF/CNPJ",
                f"{taxa_match_cpf:.1f}%",
                delta="Correspondência"
            )
        else:
            st.metric("Taxa CPF/CNPJ", "N/A", delta="Sem dados")
    
    # Análise de Decadência (regra: primeiro dia útil >= Data Infração; 31 dias corridos autuação, 180 dias corridos multa; +4 dias nas datas de notificação)
    df_final_sql = resultados.get('df_final_sql', pd.DataFrame())
    if not df_final_sql.empty and 'Situação decadente' in df_final_sql.columns:
        st.markdown("---")
        st.markdown("### ⏱️ Análise de Decadência")
        st.caption("Regra: o prazo começa no **primeiro dia útil** a partir da Data Infração (a própria data se for dia útil). A partir desse dia 1, a contagem é em **dias corridos**: a notificação de autuação (ajustada em +4 dias) não pode ultrapassar **31 dias corridos**, e a notificação de multa (ajustada em +4 dias) não pode ultrapassar **181 dias corridos**. Autos que ultrapassam são decadentes.")
        sit = df_final_sql['Situação decadente'].fillna('').astype(str).str.strip()
        qtd_decad_autuacao = int(((sit == 'Decadente autuação') | (sit == 'Decadente autuação e multa')).sum())
        qtd_decad_multa = int(((sit == 'Decadente multa') | (sit == 'Decadente autuação e multa')).sum())
        qtd_decad_ambos = int((sit == 'Decadente autuação e multa').sum())
        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            st.metric("Decadentes (autuação)", f"{qtd_decad_autuacao:,}", delta="> 31 dias corridos")
        with col_d2:
            st.metric("Decadentes (multa)", f"{qtd_decad_multa:,}", delta="> 181 dias corridos")
        with col_d3:
            st.metric("Decadentes (autuação e multa)", f"{qtd_decad_ambos:,}", delta="ambos prazos ultrapassados")
        # Gráfico por ano (ano da Data Infração)
        col_infracao = _resolver_coluna_data(df_final_sql, "Data Infração")
        if col_infracao:
            data_infracao_dt = pd.to_datetime(df_final_sql[col_infracao], errors='coerce', dayfirst=True)
            df_final_sql = df_final_sql.copy()
            df_final_sql['_ano_infracao'] = data_infracao_dt.dt.year
            df_decad = df_final_sql[sit != ''].copy()
            if not df_decad.empty and df_decad['_ano_infracao'].notna().any():
                df_decad['Decadente autuação'] = (df_decad['Situação decadente'].fillna('').str.contains('autuação', na=False)).astype(int)
                df_decad['Decadente multa'] = (df_decad['Situação decadente'].fillna('').str.contains('multa', na=False)).astype(int)
                por_ano = df_decad.groupby('_ano_infracao').agg({'Decadente autuação': 'sum', 'Decadente multa': 'sum'}).reset_index()
                por_ano = por_ano.rename(columns={'_ano_infracao': 'Ano'})
                fig_decad = go.Figure()
                fig_decad.add_trace(go.Bar(name='Decadente autuação', x=por_ano['Ano'].astype(int), y=por_ano['Decadente autuação'], marker_color='#e74c3c'))
                fig_decad.add_trace(go.Bar(name='Decadente multa', x=por_ano['Ano'].astype(int), y=por_ano['Decadente multa'], marker_color='#f39c12'))
                fig_decad.update_layout(
                    title='Decadentes por ano (Data Infração)',
                    barmode='group',
                    xaxis_title='Ano',
                    yaxis_title='Quantidade',
                    height=400,
                    showlegend=True
                )
                st.plotly_chart(fig_decad, use_container_width=True)
    
    # Abas de análise detalhada
    tab1, tab2, tab3, tab4 = st.tabs([
        "🔑 Autos de Infração",
        "📈 Agrupamento por CPF/CNPJ",
        "💰 Separação por Valores",
        "⚠️ Divergências"
    ])
    
    with tab1:
        st.markdown("### 🔑 Análise de Autos de Infração")
        st.info("💡 Esta é a análise PRINCIPAL. Os autos de infração são a chave de comparação entre as bases.")
        
        st.markdown("#### ✅ Autos Presentes em Ambas as Bases (Vencimento em 2025: 01/01 a 31/12)")
        st.success(f"Total de {resultados['autos_em_ambas']:,} autos de infração encontrados em ambas as bases com vencimento em 2025")
        
        # Criar visualização comparativa com valores e vencimentos
        if not resultados['df_serasa_filtrado'].empty and not resultados['df_divida_filtrado'].empty:
            # Preparar dados comparativos
            df_serasa_comp = resultados['df_serasa_filtrado'].copy()
            df_divida_comp = resultados['df_divida_filtrado'].copy()
            
            # Garantir que as colunas de valor e vencimento estejam visíveis
            colunas_importantes = []
            if coluna_auto in df_serasa_comp.columns:
                colunas_importantes.append(coluna_auto)
            if coluna_valor in df_serasa_comp.columns:
                colunas_importantes.append(coluna_valor)
            if coluna_vencimento in df_serasa_comp.columns:
                colunas_importantes.append(coluna_vencimento)
            
            # Adicionar todas as outras colunas
            outras_colunas = [c for c in df_serasa_comp.columns if c not in colunas_importantes and c not in ['AUTO_NORM', 'CPF_CNPJ_NORM']]
            colunas_ordenadas = colunas_importantes + outras_colunas
            
            # Reordenar colunas para destacar Auto, Valor e Vencimento
            df_serasa_display = df_serasa_comp[[c for c in colunas_ordenadas if c in df_serasa_comp.columns]]
            df_divida_display = df_divida_comp[[c for c in colunas_ordenadas if c in df_divida_comp.columns]]
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("##### 📊 Base SERASA - Autos Correspondentes")
                st.markdown("**Valores e Vencimentos 2025 em diante**")
                st.info("💡 Mostrando apenas autos que estão em **AMBAS as bases** (SERASA e Dívida Ativa)")
                st.dataframe(df_serasa_display, use_container_width=True, height=400)
                
                # Estatísticas de valores - AUTOS EM AMBAS AS BASES
                if coluna_valor in df_serasa_comp.columns:
                    try:
                        # Garantir que a coluna de valor está convertida para numérico
                        if df_serasa_comp[coluna_valor].dtype not in ['int64', 'float64']:
                            valores_serasa = pd.to_numeric(df_serasa_comp[coluna_valor], errors='coerce')
                        else:
                            valores_serasa = df_serasa_comp[coluna_valor]
                        
                        # Remover NaN antes de calcular soma e média (garantir exatidão)
                        valores_serasa_validos = valores_serasa[valores_serasa.notna()]
                        if len(valores_serasa_validos) > 0:
                            # Soma exata (como no Excel - somando todos os valores válidos)
                            soma_serasa = float(valores_serasa_validos.sum())
                            media_serasa = float(valores_serasa_validos.mean())
                            st.metric("💰 Valor Total SERASA (Em Ambas)", f"R$ {soma_serasa:,.2f}", 
                                     delta=f"Média: R$ {media_serasa:,.2f}")
                    except Exception as e:
                        st.error(f"Erro ao calcular valores SERASA: {str(e)}")
                        pass
                
                st.caption(f"📋 Total: {len(resultados['df_serasa_filtrado'])} autos de infração em ambas as bases (vencimento em 2025)")
            
            with col2:
                st.markdown("##### 📊 Base Dívida Ativa - Autos Correspondentes")
                st.markdown("**Valores e Vencimentos 2025 em diante**")
                st.info("💡 Mostrando apenas autos que estão em **AMBAS as bases** (SERASA e Dívida Ativa)")
                st.dataframe(df_divida_display, use_container_width=True, height=400)
                
                # Estatísticas de valores - AUTOS EM AMBAS AS BASES
                if coluna_valor in df_divida_comp.columns:
                    try:
                        # Garantir que a coluna de valor está convertida para numérico
                        if df_divida_comp[coluna_valor].dtype not in ['int64', 'float64']:
                            valores_divida = pd.to_numeric(df_divida_comp[coluna_valor], errors='coerce')
                        else:
                            valores_divida = df_divida_comp[coluna_valor]
                        
                        # Remover NaN antes de calcular soma e média (garantir exatidão)
                        valores_divida_validos = valores_divida[valores_divida.notna()]
                        if len(valores_divida_validos) > 0:
                            # Soma exata (como no Excel - somando todos os valores válidos)
                            soma_divida = float(valores_divida_validos.sum())
                            media_divida = float(valores_divida_validos.mean())
                            st.metric("💰 Valor Total Dívida Ativa (Em Ambas)", f"R$ {soma_divida:,.2f}", 
                                     delta=f"Média: R$ {media_divida:,.2f}")
                    except Exception as e:
                        st.error(f"Erro ao calcular valores Dívida Ativa: {str(e)}")
                        pass
                
                st.caption(f"📋 Total: {len(resultados['df_divida_filtrado'])} autos de infração em ambas as bases (vencimento em 2025)")
            
            # Comparação de valores - TOTAIS GERAIS (TODOS os autos com vencimento em 2025)
            st.markdown("---")
            st.markdown("#### 💰 Totais Gerais (TODOS os autos com vencimento em 2025: 01/01 a 31/12)")
            st.warning("⚠️ **IMPORTANTE:** Estes são os totais de TODOS os autos de cada base com vencimento em 2025, não apenas os que estão em ambas as bases.")
            
            if coluna_valor in resultados['df_serasa_total_2025'].columns and coluna_valor in resultados['df_divida_total_2025'].columns:
                col1, col2, col3 = st.columns(3)
                
                try:
                    # Calcular totais da SERASA (TODOS os autos)
                    df_serasa_total = resultados['df_serasa_total_2025'].copy()
                    if df_serasa_total[coluna_valor].dtype not in ['int64', 'float64']:
                        valores_serasa_total = pd.to_numeric(df_serasa_total[coluna_valor], errors='coerce')
                    else:
                        valores_serasa_total = df_serasa_total[coluna_valor]
                    
                    valores_serasa_total_validos = valores_serasa_total[valores_serasa_total.notna()]
                    soma_serasa_total = float(valores_serasa_total_validos.sum()) if len(valores_serasa_total_validos) > 0 else 0.0
                    
                    # Calcular totais da Dívida Ativa (TODOS os autos)
                    df_divida_total = resultados['df_divida_total_2025'].copy()
                    if df_divida_total[coluna_valor].dtype not in ['int64', 'float64']:
                        valores_divida_total = pd.to_numeric(df_divida_total[coluna_valor], errors='coerce')
                    else:
                        valores_divida_total = df_divida_total[coluna_valor]
                    
                    valores_divida_total_validos = valores_divida_total[valores_divida_total.notna()]
                    soma_divida_total = float(valores_divida_total_validos.sum()) if len(valores_divida_total_validos) > 0 else 0.0
                    
                    with col1:
                        st.metric("💰 Total SERASA (Todos)", f"R$ {soma_serasa_total:,.2f}", 
                                 delta=f"{len(valores_serasa_total_validos):,} autos")
                        st.caption("Soma de TODOS os valores da coluna de valor da SERASA com vencimento em 2025")
                    
                    with col2:
                        st.metric("💰 Total Dívida Ativa (Todos)", f"R$ {soma_divida_total:,.2f}", 
                                 delta=f"{len(valores_divida_total_validos):,} autos")
                        st.caption("Soma de TODOS os valores da coluna de valor da Dívida Ativa com vencimento em 2025")
                    
                    with col3:
                        diferenca_total = soma_serasa_total - soma_divida_total
                        st.metric("Diferença", f"R$ {diferenca_total:,.2f}", 
                                 delta="SERASA - Dívida Ativa")
                except Exception as e:
                    st.error(f"Erro ao calcular totais: {str(e)}")
            
            # Comparação de valores - APENAS AUTOS EM AMBAS AS BASES
            st.markdown("---")
            st.markdown("#### 💰 Comparação de Valores (Apenas autos em AMBAS as bases)")
            st.info("💡 Estes são os totais apenas dos autos que estão presentes em AMBAS as bases (SERASA e Dívida Ativa) com vencimento em 2025.")
            
            if coluna_valor in df_serasa_comp.columns and coluna_valor in df_divida_comp.columns:
                col1, col2, col3 = st.columns(3)
                
                try:
                    # Garantir que as colunas de valor estão convertidas para numérico
                    if df_serasa_comp[coluna_valor].dtype not in ['int64', 'float64']:
                        valores_serasa = pd.to_numeric(df_serasa_comp[coluna_valor], errors='coerce')
                    else:
                        valores_serasa = df_serasa_comp[coluna_valor]
                    
                    if df_divida_comp[coluna_valor].dtype not in ['int64', 'float64']:
                        valores_divida = pd.to_numeric(df_divida_comp[coluna_valor], errors='coerce')
                    else:
                        valores_divida = df_divida_comp[coluna_valor]
                    
                    # Remover NaN antes de calcular (garantir exatidão)
                    valores_serasa_validos = valores_serasa[valores_serasa.notna()]
                    valores_divida_validos = valores_divida[valores_divida.notna()]
                    
                    # Soma exata (como no Excel - somando todos os valores válidos)
                    soma_serasa = float(valores_serasa_validos.sum()) if len(valores_serasa_validos) > 0 else 0.0
                    soma_divida = float(valores_divida_validos.sum()) if len(valores_divida_validos) > 0 else 0.0
                    
                    with col1:
                        st.metric("Total SERASA (Em Ambas)", f"R$ {soma_serasa:,.2f}", 
                                 delta=f"{len(valores_serasa_validos):,} autos")
                        st.caption("Soma dos valores dos autos que estão em ambas as bases")
                    
                    with col2:
                        st.metric("Total Dívida Ativa (Em Ambas)", f"R$ {soma_divida:,.2f}", 
                                 delta=f"{len(valores_divida_validos):,} autos")
                        st.caption("Soma dos valores dos autos que estão em ambas as bases")
                    
                    with col3:
                        diferenca = soma_serasa - soma_divida
                        st.metric("Diferença", f"R$ {diferenca:,.2f}", 
                                 delta="SERASA - Dívida Ativa")
                except Exception as e:
                    st.error(f"Erro ao calcular comparação: {str(e)}")
            
            # Resumo de vencimentos - TOTAIS GERAIS
            st.markdown("---")
            st.markdown("#### 📅 Resumo de Vencimentos - Totais Gerais (vencimento em 2025)")
            st.warning("⚠️ **IMPORTANTE:** Estes são os totais de TODOS os autos de cada base com vencimento em 2025 (01/01 a 31/12).")
            col1, col2 = st.columns(2)
            
            with col1:
                try:
                    df_serasa_total = resultados['df_serasa_total_2025'].copy()
                    if coluna_vencimento in df_serasa_total.columns:
                        venc_serasa = pd.to_datetime(df_serasa_total[coluna_vencimento], errors='coerce')
                        data_limite = pd.Timestamp('2025-01-01')
                        data_limite_fim = pd.Timestamp('2025-12-31')
                        venc_serasa_filtrado = venc_serasa[(venc_serasa >= data_limite) & (venc_serasa <= data_limite_fim)]
                        if len(venc_serasa_filtrado) > 0:
                            st.markdown("**SERASA (Todos os autos):**")
                            st.write(f"- Primeiro vencimento: {venc_serasa_filtrado.min().strftime('%d/%m/%Y')}")
                            st.write(f"- Último vencimento: {venc_serasa_filtrado.max().strftime('%d/%m/%Y')}")
                            st.write(f"- **Total com vencimento em 2025: {len(df_serasa_total):,} autos**")
                except Exception as e:
                    st.error(f"Erro ao processar vencimentos SERASA: {str(e)}")
            
            with col2:
                try:
                    df_divida_total = resultados['df_divida_total_2025'].copy()
                    if coluna_vencimento in df_divida_total.columns:
                        venc_divida = pd.to_datetime(df_divida_total[coluna_vencimento], errors='coerce')
                        data_limite = pd.Timestamp('2025-01-01')
                        data_limite_fim = pd.Timestamp('2025-12-31')
                        venc_divida_filtrado = venc_divida[(venc_divida >= data_limite) & (venc_divida <= data_limite_fim)]
                        if len(venc_divida_filtrado) > 0:
                            st.markdown("**Dívida Ativa (Todos os autos):**")
                            st.write(f"- Primeiro vencimento: {venc_divida_filtrado.min().strftime('%d/%m/%Y')}")
                            st.write(f"- Último vencimento: {venc_divida_filtrado.max().strftime('%d/%m/%Y')}")
                            st.write(f"- **Total com vencimento em 2025: {len(df_divida_total):,} autos**")
                except Exception as e:
                    st.error(f"Erro ao processar vencimentos Dívida Ativa: {str(e)}")
            
            st.markdown("---")
            
            # Área de Exportação - Download Principal (SEPARADO POR FAIXA DE VALOR + TODOS)
            st.markdown("#### 📥 Exportar Base Comparada")
            # Verificar se a coluna de protocolo existe nos dados
            tem_protocolo = coluna_protocolo in resultados['df_serasa_filtrado'].columns if not resultados['df_serasa_filtrado'].empty else False
            if tem_protocolo:
                st.info("💡 Exporte os autos que estão em **AMBAS as bases** (SERASA e Dívida Ativa). **IMPORTANTE:** Antes da exportação, o sistema **exclui autuados classificados como Órgão, Banco ou Leasing** pela coluna de nome do autuado da SERASA, preservando exceções permitidas como **SAFRA**. A classificação por faixa de valor (até R$ 999,99 ou acima de R$ 1.000,00) é feita pela **SOMA dos valores por CPF/CNPJ**. Se um CPF/CNPJ tiver soma <= R$ 999,99, TODOS os seus autos vão para 'até R$ 999,99'. Se tiver soma >= R$ 1.000,00, TODOS os seus autos vão para 'acima de R$ 1.000,00'. Cada arquivo contém: Auto de Infração, **Número de Protocolo**, CPF/CNPJ e **Valor Individual de cada auto**, ordenados por CPF/CNPJ (do maior para menor número de autos).")
            else:
                st.info("💡 Exporte os autos que estão em **AMBAS as bases** (SERASA e Dívida Ativa). **IMPORTANTE:** Antes da exportação, o sistema **exclui autuados classificados como Órgão, Banco ou Leasing** pela coluna de nome do autuado da SERASA, preservando exceções permitidas como **SAFRA**. A classificação por faixa de valor (até R$ 999,99 ou acima de R$ 1.000,00) é feita pela **SOMA dos valores por CPF/CNPJ**. Se um CPF/CNPJ tiver soma <= R$ 999,99, TODOS os seus autos vão para 'até R$ 999,99'. Se tiver soma >= R$ 1.000,00, TODOS os seus autos vão para 'acima de R$ 1.000,00'. Cada arquivo contém: Auto de Infração, CPF/CNPJ e **Valor Individual de cada auto**, ordenados por CPF/CNPJ (do maior para menor número de autos).")
            
            # Preparar dados base para exportação - ANÁLISE ACUMULATIVA POR CPF/CNPJ
            # Usar df_final que segue a ordem do SQL (já processado: JOIN → Valores → Zeros → Duplicados → Data)
            # O df_final já tem os valores convertidos e processados corretamente
            if 'df_final_sql' in resultados and not resultados['df_final_sql'].empty:
                df_final_work = resultados['df_final_sql'].copy()
                
                # Criar df_export mapeando colunas do df_final para estrutura esperada
                # Usar colunas _serasa como padrão (ou _divida se não existir _serasa)
                df_export = pd.DataFrame()
                
                # Mapear Auto de Infração
                if f"{coluna_auto}_serasa" in df_final_work.columns:
                    df_export[coluna_auto] = df_final_work[f"{coluna_auto}_serasa"]
                elif f"{coluna_auto}_divida" in df_final_work.columns:
                    df_export[coluna_auto] = df_final_work[f"{coluna_auto}_divida"]
                elif 'AUTO_NORM' in df_final_work.columns:
                    # Se tiver AUTO_NORM, buscar do df_serasa_filtrado
                    autos_finais = set(df_final_work['AUTO_NORM'].unique())
                    df_export = resultados['df_serasa_filtrado'][resultados['df_serasa_filtrado']['AUTO_NORM'].isin(autos_finais)].copy()
                else:
                    df_export = resultados['df_serasa_filtrado'].copy()
                
                # Se conseguiu mapear do df_final, adicionar outras colunas
                if coluna_auto in df_export.columns and len(df_export) == len(df_final_work):
                    # Adicionar coluna de valor convertida (do df_final)
                    if 'Valor' in df_final_work.columns:
                        df_export[coluna_valor] = df_final_work['Valor']
                    
                    # Adicionar outras colunas necessárias
                    if coluna_cpf_cnpj in resultados['df_serasa_filtrado'].columns:
                        # Mapear CPF/CNPJ
                        if f"{coluna_cpf_cnpj}_serasa" in df_final_work.columns:
                            df_export[coluna_cpf_cnpj] = df_final_work[f"{coluna_cpf_cnpj}_serasa"]
                        elif f"{coluna_cpf_cnpj}_divida" in df_final_work.columns:
                            df_export[coluna_cpf_cnpj] = df_final_work[f"{coluna_cpf_cnpj}_divida"]
                    
                    # Adicionar coluna de vencimento (sempre incluir)
                    vencimento_adicionado = False
                    col_venc_df_final = resolver_coluna_vencimento(df_final_work, coluna_vencimento)
                    if col_venc_df_final:
                        df_export[coluna_vencimento] = df_final_work[col_venc_df_final]
                        vencimento_adicionado = True
                    
                    # Se não encontrou no df_final, buscar do df_serasa_filtrado usando AUTO_NORM
                    if not vencimento_adicionado:
                        col_venc_serasa = resolver_coluna_vencimento(resultados['df_serasa_filtrado'], coluna_vencimento)
                        if col_venc_serasa and 'AUTO_NORM' in df_final_work.columns and 'AUTO_NORM' in resultados['df_serasa_filtrado'].columns:
                            # Criar mapeamento por AUTO_NORM
                            vencimento_map = resultados['df_serasa_filtrado'].set_index('AUTO_NORM')[col_venc_serasa].to_dict()
                            df_export[coluna_vencimento] = df_final_work['AUTO_NORM'].map(vencimento_map)
                            vencimento_adicionado = True
                    
                    # Se ainda não encontrou, buscar diretamente do df_serasa_filtrado (garantir que sempre tenha a coluna)
                    if not vencimento_adicionado:
                        if coluna_vencimento in resultados['df_serasa_filtrado'].columns and 'AUTO_NORM' in df_export.columns:
                            # Mapear usando AUTO_NORM
                            vencimento_map = resultados['df_serasa_filtrado'].set_index('AUTO_NORM')[coluna_vencimento].to_dict()
                            df_export[coluna_vencimento] = df_export['AUTO_NORM'].map(vencimento_map).fillna('')
                        else:
                            # Se não conseguir mapear, criar coluna vazia (mas sempre incluir)
                            df_export[coluna_vencimento] = pd.Series([''] * len(df_export), index=df_export.index)
                    
                    # Adicionar coluna de protocolo
                    if coluna_protocolo in resultados['df_serasa_filtrado'].columns:
                        if f"{coluna_protocolo}_serasa" in df_final_work.columns:
                            df_export[coluna_protocolo] = df_final_work[f"{coluna_protocolo}_serasa"]
                        elif f"{coluna_protocolo}_divida" in df_final_work.columns:
                            df_export[coluna_protocolo] = df_final_work[f"{coluna_protocolo}_divida"]
                    
                    # Adicionar coluna de Modais
                    if 'Modais' in df_final_work.columns:
                        df_export['Modais'] = df_final_work['Modais']
                    elif 'Modais' in resultados['df_serasa_filtrado'].columns and 'AUTO_NORM' in df_export.columns:
                        # Mapear usando AUTO_NORM
                        modais_map = resultados['df_serasa_filtrado'].set_index('AUTO_NORM')['Modais'].to_dict()
                        df_export['Modais'] = df_export['AUTO_NORM'].map(modais_map).fillna('')

                    # Adicionar Nome Autuado da base SERASA para exclusão de órgãos/bancos/leasing na exportação comparada
                    if coluna_nome_autuado in resultados['df_serasa_filtrado'].columns:
                        if f"{coluna_nome_autuado}_serasa" in df_final_work.columns:
                            df_export[coluna_nome_autuado] = df_final_work[f"{coluna_nome_autuado}_serasa"]
                        elif 'AUTO_NORM' in df_export.columns and 'AUTO_NORM' in resultados['df_serasa_filtrado'].columns:
                            nome_autuado_map = resultados['df_serasa_filtrado'].set_index('AUTO_NORM')[coluna_nome_autuado].to_dict()
                            df_export[coluna_nome_autuado] = df_export['AUTO_NORM'].map(nome_autuado_map).fillna('')
                    
                    # Adicionar AUTO_NORM para compatibilidade
                    if 'AUTO_NORM' in df_final_work.columns:
                        df_export['AUTO_NORM'] = df_final_work['AUTO_NORM']
                    
                    # IMPORTANTE: Adicionar CPF_CNPJ_NORM para a lógica acumulativa funcionar
                    if coluna_cpf_cnpj in df_export.columns:
                        if 'CPF_CNPJ_NORM' not in df_export.columns:
                            df_export['CPF_CNPJ_NORM'] = df_export[coluna_cpf_cnpj].apply(normalizar_cpf_cnpj)
                    elif 'CPF_CNPJ_NORM' in resultados['df_serasa_filtrado'].columns and 'AUTO_NORM' in df_export.columns:
                        # Mapear CPF_CNPJ_NORM do df_serasa_filtrado usando AUTO_NORM
                        cpf_norm_map = resultados['df_serasa_filtrado'].set_index('AUTO_NORM')['CPF_CNPJ_NORM'].to_dict()
                        df_export['CPF_CNPJ_NORM'] = df_export['AUTO_NORM'].map(cpf_norm_map)
                    # Coluna Situação decadente (regra de datas: 37 dias autuação, 187 dias multa)
                    if 'Situação decadente' in df_final_work.columns:
                        df_export['Situação decadente'] = df_final_work['Situação decadente']
            else:
                # Fallback: usar df_serasa_filtrado
                df_export = resultados['df_serasa_filtrado'].copy()
                # Garantir que CPF_CNPJ_NORM existe
                if 'CPF_CNPJ_NORM' not in df_export.columns and coluna_cpf_cnpj in df_export.columns:
                    df_export['CPF_CNPJ_NORM'] = df_export[coluna_cpf_cnpj].apply(normalizar_cpf_cnpj)
            
            if df_export.empty:
                st.warning("⚠️ Nenhum dado encontrado para exportar!")
            else:
                # Função auxiliar para gerar Excel formatado
                def gerar_excel_formatado(dados_df, nome_aba, nome_arquivo):
                    """Gera arquivo Excel formatado a partir de um DataFrame.
                    Remove linhas totalmente vazias e reindexa para evitar linhas em branco entre os dados."""
                    if dados_df is None or dados_df.empty:
                        return None
                    # Remover linhas onde todos os valores são vazios ('' ou NaN) e reindexar
                    dados_df = dados_df.replace('', np.nan).dropna(how='all').reset_index(drop=True)
                    if dados_df.empty:
                        return None
                    buffer = io.BytesIO()
                    try:
                        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                            dados_df.to_excel(
                                writer, 
                                sheet_name=nome_aba, 
                                index=False,
                                header=True
                            )
                            
                            worksheet = writer.sheets[nome_aba]
                            
                            # Aplicar formatação completa
                            # Ajustar larguras das colunas baseado no número de colunas
                            num_colunas = len(dados_df.columns)
                            if num_colunas == 5:  # Auto, Protocolo, Data Vencimento, CPF/CNPJ, Valor
                                worksheet.column_dimensions['A'].width = 25  # Auto de Infração
                                worksheet.column_dimensions['B'].width = 20  # Número de Protocolo
                                worksheet.column_dimensions['C'].width = 18  # Data de Vencimento
                                worksheet.column_dimensions['D'].width = 18  # CPF/CNPJ
                                worksheet.column_dimensions['E'].width = 15  # Valor
                            elif num_colunas == 4:  # Auto, Protocolo, CPF/CNPJ, Valor (sem data) OU Auto, Data Vencimento, CPF/CNPJ, Valor (sem protocolo)
                                # Verificar se tem protocolo ou data
                                if 'Número de Protocolo' in dados_df.columns:
                                    worksheet.column_dimensions['A'].width = 25  # Auto de Infração
                                    worksheet.column_dimensions['B'].width = 20  # Número de Protocolo
                                    worksheet.column_dimensions['C'].width = 18  # CPF/CNPJ
                                    worksheet.column_dimensions['D'].width = 15  # Valor
                                else:  # Tem Data de Vencimento mas não tem Protocolo
                                    worksheet.column_dimensions['A'].width = 25  # Auto de Infração
                                    worksheet.column_dimensions['B'].width = 18  # Data de Vencimento
                                    worksheet.column_dimensions['C'].width = 18  # CPF/CNPJ
                                    worksheet.column_dimensions['D'].width = 15  # Valor
                            else:  # Auto, CPF/CNPJ, Valor (sem protocolo e sem data)
                                worksheet.column_dimensions['A'].width = 25
                                worksheet.column_dimensions['B'].width = 18
                                worksheet.column_dimensions['C'].width = 15
                            
                            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
                            header_font = Font(bold=True, color="FFFFFF", size=11)
                            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            align_left_center = Alignment(horizontal="left", vertical="center")
                            align_center_center = Alignment(horizontal="center", vertical="center")
                            align_right_center = Alignment(horizontal="right", vertical="center")
                            
                            for cell in worksheet[1]:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = header_alignment
                            
                            # Formatar colunas
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            
                            # Determinar número de colunas e índices
                            num_colunas = len(dados_df.columns)
                            # Determinar índices das colunas baseado na estrutura
                            tem_protocolo = 'Número de Protocolo' in dados_df.columns
                            tem_data_venc = 'Data de Vencimento' in dados_df.columns
                            tem_data_infracao = 'Data Infração' in dados_df.columns
                            tem_modais = 'Modais' in dados_df.columns
                            
                            # Calcular índices das colunas (ordem: Auto, Protocolo, Data Venc, Data Infração, Modais, CPF/CNPJ, Valor, Valor (R$))
                            col_names = list(dados_df.columns)
                            idx_auto = 1
                            idx_protocolo = 2 if tem_protocolo else None
                            idx_data_venc = None
                            idx_data_infracao = None
                            idx_modais = None
                            idx_cpf = None
                            idx_valor = col_names.index('Valor') + 1 if 'Valor' in col_names else None
                            tem_valor_r = 'Valor (R$)' in col_names
                            idx_valor_r = col_names.index('Valor (R$)') + 1 if tem_valor_r else None
                            tem_situacao_divida = 'Situação Dívida' in col_names
                            idx_situacao_divida = col_names.index('Situação Dívida') + 1 if tem_situacao_divida else None
                            tem_situacao_congelamento = 'Situação Congelamento' in col_names
                            idx_situacao_congelamento = col_names.index('Situação Congelamento') + 1 if tem_situacao_congelamento else None
                            tem_data_pagamento = 'Data Pagamento' in col_names
                            idx_data_pagamento = col_names.index('Data Pagamento') + 1 if tem_data_pagamento else None
                            tem_nome_autuado = 'Nome Autuado' in col_names
                            idx_nome_autuado = col_names.index('Nome Autuado') + 1 if tem_nome_autuado else None
                            tem_classificacao_autuado = 'Classificação Autuado' in col_names
                            idx_classificacao_autuado = col_names.index('Classificação Autuado') + 1 if tem_classificacao_autuado else None
                            tem_motivo_classificacao = 'Motivo Classificação' in col_names
                            idx_motivo_classificacao = col_names.index('Motivo Classificação') + 1 if tem_motivo_classificacao else None
                            tem_termo_identificado = 'Termo Identificado' in col_names
                            idx_termo_identificado = col_names.index('Termo Identificado') + 1 if tem_termo_identificado else None
                            
                            # Calcular índices dinamicamente baseado nas colunas presentes
                            col_idx = 1
                            if tem_protocolo:
                                col_idx += 1
                            if tem_data_venc:
                                idx_data_venc = col_idx
                                col_idx += 1
                            if tem_data_infracao:
                                idx_data_infracao = col_idx
                                col_idx += 1
                            if tem_modais:
                                idx_modais = col_idx
                                col_idx += 1
                            idx_cpf = col_idx
                            
                            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                                for cell in row:
                                    cell.border = thin_border
                                    if cell.column == idx_cpf and cell.row > 1:  # CPF/CNPJ
                                        cell.number_format = '@'
                                        cell.alignment = align_center_center
                                    elif idx_valor and cell.column == idx_valor and cell.row > 1:  # Valor
                                        if cell.value is not None:
                                            cell.number_format = '#,##0.00'
                                            cell.alignment = align_right_center
                                    elif cell.column == idx_auto and cell.row > 1:  # Auto de Infração
                                        cell.alignment = align_left_center
                                    elif idx_protocolo and cell.column == idx_protocolo and cell.row > 1:  # Protocolo (se existir)
                                        cell.alignment = align_left_center
                                    elif idx_situacao_divida and cell.column == idx_situacao_divida and cell.row > 1:  # Situação Dívida (se existir)
                                        cell.number_format = '@'
                                        cell.alignment = align_left_center
                                    elif idx_situacao_congelamento and cell.column == idx_situacao_congelamento and cell.row > 1:  # Situação Congelamento (se existir)
                                        cell.number_format = '@'
                                        cell.alignment = align_left_center
                                    elif idx_data_pagamento and cell.column == idx_data_pagamento and cell.row > 1:  # Data Pagamento (se existir)
                                        cell.number_format = '@'
                                        cell.alignment = align_center_center
                                    elif idx_nome_autuado and cell.column == idx_nome_autuado and cell.row > 1:  # Nome Autuado (se existir)
                                        cell.number_format = '@'
                                        cell.alignment = align_left_center
                                    elif idx_classificacao_autuado and cell.column == idx_classificacao_autuado and cell.row > 1:  # Classificação Autuado (se existir)
                                        cell.number_format = '@'
                                        cell.alignment = align_left_center
                                    elif idx_motivo_classificacao and cell.column == idx_motivo_classificacao and cell.row > 1:  # Motivo Classificação
                                        cell.number_format = '@'
                                        cell.alignment = align_left_center
                                    elif idx_termo_identificado and cell.column == idx_termo_identificado and cell.row > 1:  # Termo Identificado
                                        cell.number_format = '@'
                                        cell.alignment = align_left_center
                                    elif idx_data_venc and cell.column == idx_data_venc and cell.row > 1:  # Data de Vencimento (se existir)
                                        cell.alignment = align_center_center
                                        cell.number_format = '@'  # Formato texto para manter formato DD/MM/YYYY
                                    elif idx_data_infracao and cell.column == idx_data_infracao and cell.row > 1:  # Data Infração (se existir)
                                        cell.alignment = align_center_center
                                        cell.number_format = '@'  # Formato texto para manter formato DD/MM/YYYY
                                    elif idx_modais and cell.column == idx_modais and cell.row > 1:  # Modais (se existir)
                                        cell.alignment = align_left_center
                                        cell.number_format = '@'  # Formato texto
                                    elif idx_valor_r and cell.column == idx_valor_r and cell.row > 1:  # Valor (R$) - texto
                                        cell.number_format = '@'
                                        cell.alignment = align_right_center
                            
                            # Largura da coluna Data Infração (quando presente)
                            if tem_data_infracao and idx_data_infracao is not None and idx_data_infracao <= 26:
                                col_letter_infracao = chr(64 + idx_data_infracao)
                                worksheet.column_dimensions[col_letter_infracao].width = 18
                            if tem_situacao_divida and idx_situacao_divida is not None and idx_situacao_divida <= 26:
                                col_letter_sit = chr(64 + idx_situacao_divida)
                                worksheet.column_dimensions[col_letter_sit].width = 22
                            if tem_situacao_congelamento and idx_situacao_congelamento is not None and idx_situacao_congelamento <= 26:
                                col_letter_cong = chr(64 + idx_situacao_congelamento)
                                worksheet.column_dimensions[col_letter_cong].width = 22
                            if tem_data_pagamento and idx_data_pagamento is not None and idx_data_pagamento <= 26:
                                col_letter_pag = chr(64 + idx_data_pagamento)
                                worksheet.column_dimensions[col_letter_pag].width = 18
                            if tem_nome_autuado and idx_nome_autuado is not None and idx_nome_autuado <= 26:
                                col_letter_nome = chr(64 + idx_nome_autuado)
                                worksheet.column_dimensions[col_letter_nome].width = 40
                            if tem_classificacao_autuado and idx_classificacao_autuado is not None and idx_classificacao_autuado <= 26:
                                col_letter_cla = chr(64 + idx_classificacao_autuado)
                                worksheet.column_dimensions[col_letter_cla].width = 28
                            if tem_motivo_classificacao and idx_motivo_classificacao is not None and idx_motivo_classificacao <= 26:
                                col_letter_motivo = chr(64 + idx_motivo_classificacao)
                                worksheet.column_dimensions[col_letter_motivo].width = 38
                            if tem_termo_identificado and idx_termo_identificado is not None and idx_termo_identificado <= 26:
                                col_letter_termo = chr(64 + idx_termo_identificado)
                                worksheet.column_dimensions[col_letter_termo].width = 28
                            if tem_valor_r and idx_valor_r is not None and idx_valor_r <= 26:
                                col_letter_valor_r = chr(64 + idx_valor_r)
                                worksheet.column_dimensions[col_letter_valor_r].width = 15
                            
                            worksheet.freeze_panes = 'A2'
                        
                        buffer.seek(0)
                        excel_data = buffer.getvalue()
                        return excel_data
                    except Exception as e:
                        buffer.close()
                        raise e
                
                # Função auxiliar para preparar dados de exportação
                def preparar_dados_exportacao(df_base, filtro_valor=None):
                    """Prepara dados para exportação com filtro opcional por valor"""
                    # Excluir da exportação comparada os autuados que não podem ser cobrados
                    # (órgão, banco, leasing), preservando exceções permitidas como SAFRA.
                    df_base = filtrar_autuados_cobraveis(df_base, coluna_nome_autuado)
                    if df_base is None or df_base.empty:
                        return None

                    # PASSO 1: PRIMEIRO - Remover valores 0,00 e NaN
                    # Garantir que valores já estão convertidos para numérico
                    if coluna_valor in df_base.columns:
                        if df_base[coluna_valor].dtype not in ['int64', 'float64']:
                            df_base[coluna_valor] = pd.to_numeric(df_base[coluna_valor], errors='coerce')
                    
                    # PRIMEIRO: Remover NaN e valores zero (R$ 0,00)
                    if coluna_valor in df_base.columns:
                        df_base_sem_zero = df_base[
                            (df_base[coluna_valor].notna()) & 
                            (df_base[coluna_valor] > 0)
                        ].copy()
                    else:
                        df_base_sem_zero = df_base.copy()
                    
                    # Aplicar filtro de valor se especificado (após remover zeros)
                    # NOVA LÓGICA: Usar análise ACUMULATIVA (soma por CPF/CNPJ)
                    if filtro_valor is not None and coluna_valor in df_base_sem_zero.columns:
                        # Verificar se existe coluna CPF_CNPJ_NORM para agrupar
                        # Se não existir, tentar criar a partir da coluna CPF/CNPJ
                        if 'CPF_CNPJ_NORM' not in df_base_sem_zero.columns and coluna_cpf_cnpj in df_base_sem_zero.columns:
                            df_base_sem_zero['CPF_CNPJ_NORM'] = df_base_sem_zero[coluna_cpf_cnpj].apply(normalizar_cpf_cnpj)

                        if 'CPF_CNPJ_NORM' in df_base_sem_zero.columns:
                            # Remover CPF/CNPJ nulos antes de agrupar
                            df_base_para_agrupar = df_base_sem_zero[df_base_sem_zero['CPF_CNPJ_NORM'].notna()].copy()

                            if not df_base_para_agrupar.empty:
                                # Agrupar por CPF/CNPJ e somar os valores
                                agrupado_export = df_base_para_agrupar.groupby('CPF_CNPJ_NORM').agg({
                                    coluna_valor: 'sum'
                                }).reset_index()
                                agrupado_export.columns = ['CPF_CNPJ_NORM', 'VALOR_TOTAL']

                                if filtro_valor == 'ate_999':
                                    # CPF/CNPJ com soma total < 1000 (até 999,99 - valor exato 1000 vai para acima_1000)
                                    cpf_ate_999 = set(
                                        agrupado_export[agrupado_export['VALOR_TOTAL'] < 1000]['CPF_CNPJ_NORM'].unique()
                                    )
                                    # Filtrar autos que pertencem a esses CPF/CNPJ
                                    df_filtrado = df_base_sem_zero[df_base_sem_zero['CPF_CNPJ_NORM'].isin(cpf_ate_999)].copy()
                                elif filtro_valor == 'acima_1000':
                                    # CPF/CNPJ com soma total >= 1000
                                    cpf_acima_1000 = set(
                                        agrupado_export[agrupado_export['VALOR_TOTAL'] >= 1000]['CPF_CNPJ_NORM'].unique()
                                    )
                                    # Filtrar autos que pertencem a esses CPF/CNPJ
                                    df_filtrado = df_base_sem_zero[df_base_sem_zero['CPF_CNPJ_NORM'].isin(cpf_acima_1000)].copy()
                                else:
                                    df_filtrado = df_base_sem_zero.copy()
                            else:
                                # Se não houver CPF/CNPJ válidos, usar lógica individual
                                if filtro_valor == 'ate_999':
                                    df_filtrado = df_base_sem_zero[df_base_sem_zero[coluna_valor] < 1000].copy()
                                elif filtro_valor == 'acima_1000':
                                    df_filtrado = df_base_sem_zero[df_base_sem_zero[coluna_valor] >= 1000].copy()
                                else:
                                    df_filtrado = df_base_sem_zero.copy()
                        else:
                            # Fallback: se não tiver CPF_CNPJ_NORM, usar lógica individual (compatibilidade)
                            if filtro_valor == 'ate_999':
                                df_filtrado = df_base_sem_zero[df_base_sem_zero[coluna_valor] < 1000].copy()
                            elif filtro_valor == 'acima_1000':
                                df_filtrado = df_base_sem_zero[df_base_sem_zero[coluna_valor] >= 1000].copy()
                            else:
                                df_filtrado = df_base_sem_zero.copy()
                    else:
                        # Se não há filtro, usar todos os autos (já sem valores zero)
                        df_filtrado = df_base_sem_zero.copy()
                    
                    if df_filtrado.empty:
                        return None
                    
                    # PASSO 2: DEPOIS - Remover duplicados baseado em Nº do Processo, mantendo DATA MAIS RECENTE
                    # Remover duplicados mantendo a data de vencimento mais recente
                    # Verificar se coluna_protocolo existe (pode estar com sufixo após merge)
                    col_protocolo_para_dup = None
                    if coluna_protocolo in df_filtrado.columns:
                        col_protocolo_para_dup = coluna_protocolo
                    elif f"{coluna_protocolo}_serasa" in df_filtrado.columns:
                        col_protocolo_para_dup = f"{coluna_protocolo}_serasa"
                    elif f"{coluna_protocolo}_divida" in df_filtrado.columns:
                        col_protocolo_para_dup = f"{coluna_protocolo}_divida"
                    
                    # Identificar coluna de vencimento para ordenação
                    col_vencimento_para_ordenacao_export = resolver_coluna_vencimento(df_filtrado, coluna_vencimento)
                    
                    if col_protocolo_para_dup:
                        # Ordenar por data de vencimento (mais recente primeiro) antes de remover duplicados
                        if col_vencimento_para_ordenacao_export and col_vencimento_para_ordenacao_export in df_filtrado.columns:
                            df_filtrado_temp = df_filtrado.copy()
                            df_filtrado_temp['_VENCIMENTO_ORD'] = pd.to_datetime(
                                df_filtrado_temp[col_vencimento_para_ordenacao_export], 
                                errors='coerce'
                            )
                            # Ordenar: mais recente primeiro (descendente)
                            df_filtrado_temp = df_filtrado_temp.sort_values(
                                by='_VENCIMENTO_ORD', 
                                ascending=False,  # Mais recente primeiro
                                na_position='last'
                            )
                            # Remover duplicados mantendo a primeira (mais recente)
                            df_filtrado = df_filtrado_temp.drop_duplicates(
                                subset=[col_protocolo_para_dup], 
                                keep='first'
                            ).copy()
                            df_filtrado = df_filtrado.drop(columns=['_VENCIMENTO_ORD'])
                        else:
                            # Se não tiver coluna de vencimento, usar lógica padrão
                            df_filtrado = df_filtrado.drop_duplicates(subset=[col_protocolo_para_dup], keep='first').copy()
                    elif coluna_auto in df_filtrado.columns:
                        # Fallback: usar Auto de Infração com mesma lógica de data mais recente
                        if col_vencimento_para_ordenacao_export and col_vencimento_para_ordenacao_export in df_filtrado.columns:
                            df_filtrado_temp = df_filtrado.copy()
                            df_filtrado_temp['_VENCIMENTO_ORD'] = pd.to_datetime(
                                df_filtrado_temp[col_vencimento_para_ordenacao_export], 
                                errors='coerce'
                            )
                            df_filtrado_temp = df_filtrado_temp.sort_values(
                                by='_VENCIMENTO_ORD', 
                                ascending=False,  # Mais recente primeiro
                                na_position='last'
                            )
                            df_filtrado = df_filtrado_temp.drop_duplicates(
                                subset=[coluna_auto], 
                                keep='first'
                            ).copy()
                            df_filtrado = df_filtrado.drop(columns=['_VENCIMENTO_ORD'])
                        else:
                            df_filtrado = df_filtrado.drop_duplicates(subset=[coluna_auto], keep='first').copy()
                    
                    # Criar DataFrame com as colunas necessárias PRESERVANDO O ÍNDICE ORIGINAL
                    # Incluir protocolo se a coluna foi informada
                    colunas_export = {
                        'Auto de Infração': df_filtrado[coluna_auto].fillna('').astype(str).str.strip() if coluna_auto in df_filtrado.columns else ''
                    }
                    
                    # Adicionar coluna de Protocolo se existir no DataFrame (verificar também com sufixos)
                    col_protocolo_encontrada = None
                    if coluna_protocolo in df_filtrado.columns:
                        col_protocolo_encontrada = coluna_protocolo
                    elif f"{coluna_protocolo}_serasa" in df_filtrado.columns:
                        col_protocolo_encontrada = f"{coluna_protocolo}_serasa"
                    elif f"{coluna_protocolo}_divida" in df_filtrado.columns:
                        col_protocolo_encontrada = f"{coluna_protocolo}_divida"
                    
                    if col_protocolo_encontrada:
                        colunas_export['Número de Protocolo'] = df_filtrado[col_protocolo_encontrada].fillna('').astype(str).str.strip()
                    
                    # Adicionar coluna de Data de Vencimento (sempre incluir, mesmo que vazia)
                    # Resolver coluna de vencimento com tolerância a variações de nome
                    col_vencimento_encontrada = resolver_coluna_vencimento(df_filtrado, coluna_vencimento)
                    
                    if col_vencimento_encontrada:
                        # Converter para datetime e formatar no padrão brasileiro (DD/MM/YYYY)
                        try:
                            vencimento_dt = pd.to_datetime(
                                df_filtrado[col_vencimento_encontrada],
                                errors='coerce',
                                dayfirst=True
                            )
                            colunas_export['Data de Vencimento'] = vencimento_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except:
                            # Se falhar, usar como string
                            colunas_export['Data de Vencimento'] = df_filtrado[col_vencimento_encontrada].fillna('').astype(str).str.strip()
                    else:
                        # Se não encontrar, criar coluna vazia
                        colunas_export['Data de Vencimento'] = pd.Series([''] * len(df_filtrado), index=df_filtrado.index)
                    
                    # Adicionar coluna de Modais se existir
                    if 'Modais' in df_filtrado.columns:
                        colunas_export['Modais'] = df_filtrado['Modais'].fillna('').astype(str).str.strip()
                    
                    # Adicionar CPF/CNPJ e Valor
                    colunas_export['CPF_CNPJ'] = df_filtrado[coluna_cpf_cnpj].fillna('').astype(str).str.strip() if coluna_cpf_cnpj in df_filtrado.columns else ''
                    colunas_export['Valor'] = df_filtrado[coluna_valor] if coluna_valor in df_filtrado.columns else None
                    # Coluna Situação decadente (Decadente autuação / Decadente multa / Decadente autuação e multa)
                    if 'Situação decadente' in df_filtrado.columns:
                        colunas_export['Situação decadente'] = df_filtrado['Situação decadente'].fillna('').astype(str).str.strip()
                    
                    dados_exportacao = pd.DataFrame(colunas_export, index=df_filtrado.index)  # PRESERVAR ÍNDICE ORIGINAL
                    
                    # Garantir que Valor seja numérico (já foi convertido antes)
                    if 'Valor' in dados_exportacao.columns:
                        dados_exportacao['Valor'] = pd.to_numeric(dados_exportacao['Valor'], errors='coerce')
                    
                    # Adicionar coluna de contagem para ordenação (sem alterar os valores individuais)
                    if coluna_cpf_cnpj in df_filtrado.columns and 'CPF_CNPJ_NORM' in df_filtrado.columns:
                        contagem_autos = df_filtrado.groupby('CPF_CNPJ_NORM').size().to_dict()
                        # Usar o índice preservado para mapear corretamente
                        dados_exportacao['_QTD_AUTOS'] = df_filtrado['CPF_CNPJ_NORM'].map(contagem_autos).fillna(0)
                        # Ordenar: 1º por quantidade de autos (maior para menor), 2º por CPF/CNPJ para agrupar autos do mesmo CNPJ
                        dados_exportacao['_CPF_CNPJ_NORM'] = df_filtrado['CPF_CNPJ_NORM']
                        dados_exportacao = dados_exportacao.sort_values(['_QTD_AUTOS', '_CPF_CNPJ_NORM'], ascending=[False, True])
                        dados_exportacao = dados_exportacao.drop(columns=['_QTD_AUTOS', '_CPF_CNPJ_NORM'])
                    
                    # Formatar CPF/CNPJ no formato brasileiro
                    dados_exportacao['CPF_CNPJ'] = dados_exportacao['CPF_CNPJ'].apply(formatar_cpf_cnpj_brasileiro)
                    
                    # Remover linhas completamente vazias
                    dados_exportacao = dados_exportacao.dropna(how='all')
                    
                    # Garantir que Auto de Infração não está vazio (linha válida)
                    if 'Auto de Infração' in dados_exportacao.columns:
                        dados_exportacao = dados_exportacao[dados_exportacao['Auto de Infração'].str.strip() != '']
                    
                    # Validação final: garantir que não há valores zero ou NaN (já foram removidos antes, mas verificar novamente)
                    if 'Valor' in dados_exportacao.columns:
                        # Manter apenas linhas com valores válidos e maiores que zero
                        dados_exportacao = dados_exportacao[
                            (dados_exportacao['Valor'].notna()) & 
                            (dados_exportacao['Valor'] > 0)
                        ]
                    
                    return dados_exportacao
                
                # Data de extração
                data_extracao = datetime.now().strftime('%d/%m/%Y')
                # Fixar o nome-base do arquivo por análise para evitar regenerações/desalinhamentos a cada rerun.
                data_arquivo = st.session_state.get('export_run_label', datetime.now().strftime('%d %m %Y %H:%M'))
                export_run_id = st.session_state.get('export_run_id', 'default')
                
                # Função auxiliar para gerar mensagem de colunas
                def get_colunas_msg(dados_df):
                    """Retorna mensagem com as colunas do arquivo"""
                    colunas = ["📊 Auto de Infração"]
                    if 'Número de Protocolo' in dados_df.columns:
                        colunas.append("Número de Protocolo")
                    if 'Situação Dívida' in dados_df.columns:
                        colunas.append("Situação Dívida")
                    if 'Situação Congelamento' in dados_df.columns:
                        colunas.append("Situação Congelamento")
                    if 'Nome Autuado' in dados_df.columns:
                        colunas.append("Nome Autuado")
                    if 'Classificação Autuado' in dados_df.columns:
                        colunas.append("Classificação Autuado")
                    if 'Motivo Classificação' in dados_df.columns:
                        colunas.append("Motivo Classificação")
                    if 'Termo Identificado' in dados_df.columns:
                        colunas.append("Termo Identificado")
                    if 'Data de Vencimento' in dados_df.columns:
                        colunas.append("Data de Vencimento")
                    if 'Data Infração' in dados_df.columns:
                        colunas.append("Data Infração")
                    if 'Data Pagamento' in dados_df.columns:
                        colunas.append("Data Pagamento")
                    if 'Modais' in dados_df.columns:
                        colunas.append("Modais")
                    colunas.append("CPF/CNPJ")
                    if 'Valor' in dados_df.columns:
                        colunas.append("Valor Individual")
                    if 'Valor (R$)' in dados_df.columns:
                        colunas.append("Valor (R$)")
                    if 'Situação decadente' in dados_df.columns:
                        colunas.append("Situação decadente")
                    return " | ".join(colunas)

                def get_resumo_classificacao_autuado(dados_df):
                    """Gera resumo textual da classificação dos autuados."""
                    if dados_df is None or dados_df.empty or 'Classificação Autuado' not in dados_df.columns:
                        return None
                    partes = []
                    contagem = dados_df['Classificação Autuado'].value_counts()
                    if not contagem.empty:
                        partes.append(" | ".join([f"{k}: {v:,}" for k, v in contagem.items()]))
                    if 'Termo Identificado' in dados_df.columns:
                        termos = dados_df['Termo Identificado'].fillna('').astype(str).str.strip()
                        termos = termos[termos != '']
                        if not termos.empty:
                            top_termos = termos.value_counts().head(5)
                            partes.append("Top termos: " + " | ".join([f"{k}: {v:,}" for k, v in top_termos.items()]))
                    return " || ".join(partes) if partes else None

                def render_exportacao_excel(chave, titulo_gerar, label_download, nome_aba, nome_arquivo, help_download, producer, empty_warning, success_template, extra_caption_fn=None):
                    """
                    Gera a exportação sob demanda.
                    Isso evita preparar vários DataFrames pesados e vários arquivos Excel antes dos botões aparecerem.
                    """
                    state_key = f"export_payload::{export_run_id}::{chave}"
                    state_df_key = f"export_df::{export_run_id}::{chave}"
                    if st.button(f"⚡ {titulo_gerar}", key=f"btn_generate::{export_run_id}::{chave}", use_container_width=True):
                        with st.spinner("Preparando arquivo Excel..."):
                            try:
                                dados_df = st.session_state.get(state_df_key)
                                if dados_df is None:
                                    dados_df = producer()
                                    if dados_df is not None and not dados_df.empty:
                                        st.session_state[state_df_key] = dados_df
                                if dados_df is None or dados_df.empty:
                                    st.session_state.pop(state_key, None)
                                    st.session_state.pop(state_df_key, None)
                                    st.warning(empty_warning)
                                else:
                                    excel_bytes = gerar_excel_formatado(dados_df, nome_aba, nome_arquivo)
                                    if excel_bytes is None:
                                        st.session_state.pop(state_key, None)
                                        st.warning(empty_warning)
                                    else:
                                        payload = {
                                            'excel': excel_bytes,
                                            'qtd': len(dados_df),
                                            'colunas_msg': get_colunas_msg(dados_df),
                                        }
                                        if extra_caption_fn:
                                            payload['extra_caption'] = extra_caption_fn(dados_df)
                                        st.session_state[state_key] = payload
                            except Exception as e:
                                st.session_state.pop(state_key, None)
                                st.error(f"⚠️ Erro ao gerar arquivo: {str(e)}")

                    payload = st.session_state.get(state_key)
                    if payload:
                        st.download_button(
                            label=label_download,
                            data=payload['excel'],
                            file_name=nome_arquivo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"btn_download::{export_run_id}::{chave}",
                            help=help_download
                        )
                        st.success(success_template.format(qtd=payload['qtd']))
                        st.caption(payload['colunas_msg'])
                        if payload.get('extra_caption'):
                            st.caption(payload['extra_caption'])
                
                def preparar_dados_apenas_serasa(df_apenas_serasa):
                    """Prepara DataFrame para exportação 'Apenas na SERASA' (autos que não constam na Dívida Ativa).
                    Colunas: Auto de Infração, Número de Protocolo, Situação Dívida, Situação Congelamento, Data de Vencimento, Data Infração, Data Pagamento (se existirem na base SERASA), Modais, CPF/CNPJ, Valor (R$), Situação decadente.
                    Remove duplicados por Auto de Infração, mantendo a linha com vencimento mais recente."""
                    if df_apenas_serasa is None or df_apenas_serasa.empty:
                        return None
                    df = df_apenas_serasa.copy()
                    col_venc_ord = resolver_coluna_vencimento(df, coluna_vencimento)
                    if col_venc_ord and coluna_auto in df.columns:
                        df['_VENC_ORD'] = pd.to_datetime(df[col_venc_ord], errors='coerce', dayfirst=True)
                        df = df.sort_values(by='_VENC_ORD', ascending=False, na_position='last')
                        df = df.drop_duplicates(subset=[coluna_auto], keep='first').copy()
                        df = df.drop(columns=['_VENC_ORD'], errors='ignore')
                    elif coluna_auto in df.columns:
                        df = df.drop_duplicates(subset=[coluna_auto], keep='first').copy()
                    if 'CPF_CNPJ_NORM' not in df.columns and coluna_cpf_cnpj in df.columns:
                        df['CPF_CNPJ_NORM'] = df[coluna_cpf_cnpj].apply(normalizar_cpf_cnpj)

                    # Opção B: manter apenas autos com valor > 0 (sem exportar zeros/nulos)
                    if coluna_valor in df.columns:
                        df['_VALOR_NUM'] = df[coluna_valor].apply(converter_valor_sql)
                    else:
                        df['_VALOR_NUM'] = None
                    df = df[(df['_VALOR_NUM'].notna()) & (df['_VALOR_NUM'] > 0)].copy()
                    if df.empty:
                        return None

                    # Coluna de decadência (apenas para Excesso de Peso e Evasão de Pedágio)
                    df['Situação decadente'] = calcular_situacao_decadente(df, coluna_modal=coluna_modal_serasa)

                    colunas_export = {}
                    colunas_export['Auto de Infração'] = df[coluna_auto].fillna('').astype(str).str.strip() if coluna_auto in df.columns else pd.Series([''] * len(df), index=df.index)
                    if coluna_protocolo in df.columns:
                        colunas_export['Número de Protocolo'] = df[coluna_protocolo].fillna('').astype(str).str.strip()
                    if 'Situação Dívida' in df.columns:
                        colunas_export['Situação Dívida'] = df['Situação Dívida'].fillna('').astype(str).str.strip()
                    if 'Situação Congelamento' in df.columns:
                        colunas_export['Situação Congelamento'] = df['Situação Congelamento'].fillna('').astype(str).str.strip()
                    col_venc = resolver_coluna_vencimento(df, coluna_vencimento)
                    if col_venc:
                        try:
                            venc_dt = pd.to_datetime(df[col_venc], errors='coerce', dayfirst=True)
                            colunas_export['Data de Vencimento'] = venc_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except Exception:
                            colunas_export['Data de Vencimento'] = df[col_venc].fillna('').astype(str).str.strip()
                    else:
                        colunas_export['Data de Vencimento'] = pd.Series([''] * len(df), index=df.index)
                    col_infracao = _resolver_coluna_data(df, "Data Infração")
                    if col_infracao:
                        try:
                            infracao_dt = pd.to_datetime(df[col_infracao], errors='coerce', dayfirst=True)
                            colunas_export['Data Infração'] = infracao_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except Exception:
                            colunas_export['Data Infração'] = df[col_infracao].fillna('').astype(str).str.strip()
                    else:
                        colunas_export['Data Infração'] = pd.Series([''] * len(df), index=df.index)
                    if 'Data Pagamento' in df.columns:
                        try:
                            pagamento_dt = pd.to_datetime(df['Data Pagamento'], errors='coerce', dayfirst=True)
                            colunas_export['Data Pagamento'] = pagamento_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except Exception:
                            colunas_export['Data Pagamento'] = df['Data Pagamento'].fillna('').astype(str).str.strip()
                    if coluna_modal_serasa and coluna_modal_serasa in df.columns:
                        colunas_export['Modais'] = df[coluna_modal_serasa].fillna('').astype(str).str.strip()
                    colunas_export['CPF_CNPJ'] = df[coluna_cpf_cnpj].fillna('').astype(str).str.strip() if coluna_cpf_cnpj in df.columns else pd.Series([''] * len(df), index=df.index)
                    colunas_export['Valor (R$)'] = df['_VALOR_NUM'].apply(formatar_valor_br)
                    colunas_export['Situação decadente'] = df['Situação decadente'].fillna('').astype(str).str.strip()
                    ordem = ['Auto de Infração']
                    if 'Número de Protocolo' in colunas_export:
                        ordem.append('Número de Protocolo')
                    if 'Situação Dívida' in colunas_export:
                        ordem.append('Situação Dívida')
                    if 'Situação Congelamento' in colunas_export:
                        ordem.append('Situação Congelamento')
                    ordem.append('Data de Vencimento')
                    ordem.append('Data Infração')
                    if 'Data Pagamento' in colunas_export:
                        ordem.append('Data Pagamento')
                    if 'Modais' in colunas_export:
                        ordem.append('Modais')
                    ordem.append('CPF_CNPJ')
                    ordem.append('Valor (R$)')
                    ordem.append('Situação decadente')
                    dados = pd.DataFrame({k: colunas_export[k] for k in ordem if k in colunas_export}, index=df.index)
                    dados['CPF_CNPJ'] = dados['CPF_CNPJ'].apply(formatar_cpf_cnpj_brasileiro)
                    if 'CPF_CNPJ_NORM' in df.columns and not dados.empty:
                        dados['_CPF'] = dados.index.map(lambda i: df.loc[i, 'CPF_CNPJ_NORM'] if i in df.index else '')
                        contagem = dados.groupby('_CPF').size().to_dict()
                        dados['_QTD'] = dados['_CPF'].map(contagem).fillna(0)
                        dados = dados.sort_values(['_QTD', '_CPF'], ascending=[False, True]).drop(columns=['_QTD', '_CPF'], errors='ignore')
                    # Remover linhas totalmente vazias e reindexar para evitar linhas em branco no Excel
                    dados = dados.replace('', np.nan).dropna(how='all').reset_index(drop=True)
                    return dados

                def preparar_dados_serasa_classificacao(df_serasa):
                    """Prepara DataFrame para exportação 'Base SERASA com classificação de autuado'.
                    Base inteira, sem filtro; colunas padrão + Nome Autuado + classificação detalhada."""
                    if df_serasa is None or df_serasa.empty:
                        return None
                    df = df_serasa.copy()
                    config_classificacao = obter_config_classificacao_ativa()
                    if 'CPF_CNPJ_NORM' not in df.columns and coluna_cpf_cnpj in df.columns:
                        df['CPF_CNPJ_NORM'] = df[coluna_cpf_cnpj].apply(normalizar_cpf_cnpj)
                    coluna_valor_serasa = coluna_valor
                    if coluna_valor_serasa in df.columns:
                        df['_VALOR_NUM'] = df[coluna_valor_serasa].apply(converter_valor_sql)
                    else:
                        df['_VALOR_NUM'] = 0.0
                    df['_VALOR_NUM'] = df['_VALOR_NUM'].fillna(0)
                    colunas_export = {}
                    colunas_export['Auto de Infração'] = df[coluna_auto].fillna('').astype(str).str.strip() if coluna_auto in df.columns else pd.Series([''] * len(df), index=df.index)
                    if coluna_protocolo in df.columns:
                        colunas_export['Número de Protocolo'] = df[coluna_protocolo].fillna('').astype(str).str.strip()
                    if coluna_nome_autuado in df.columns:
                        colunas_export['Nome Autuado'] = df[coluna_nome_autuado].fillna('').astype(str).str.strip()
                        classificacoes = df[coluna_nome_autuado].apply(
                            lambda nome: classificar_autuado_detalhado(nome, config=config_classificacao)
                        )
                        detalhes_df = pd.DataFrame(
                            classificacoes.tolist(),
                            index=df.index,
                            columns=['Classificação Autuado', 'Motivo Classificação', 'Termo Identificado']
                        )
                        colunas_export['Classificação Autuado'] = detalhes_df['Classificação Autuado']
                        colunas_export['Motivo Classificação'] = detalhes_df['Motivo Classificação']
                        colunas_export['Termo Identificado'] = detalhes_df['Termo Identificado']
                    else:
                        colunas_export['Nome Autuado'] = pd.Series([''] * len(df), index=df.index)
                        colunas_export['Classificação Autuado'] = pd.Series(['Pode cobrar'] * len(df), index=df.index)
                        colunas_export['Motivo Classificação'] = pd.Series(['Coluna Nome Autuado não encontrada'] * len(df), index=df.index)
                        colunas_export['Termo Identificado'] = pd.Series([''] * len(df), index=df.index)
                    if 'Situação Dívida' in df.columns:
                        colunas_export['Situação Dívida'] = df['Situação Dívida'].fillna('').astype(str).str.strip()
                    if 'Situação Congelamento' in df.columns:
                        colunas_export['Situação Congelamento'] = df['Situação Congelamento'].fillna('').astype(str).str.strip()
                    col_venc = resolver_coluna_vencimento(df, coluna_vencimento)
                    if col_venc:
                        try:
                            venc_dt = pd.to_datetime(df[col_venc], errors='coerce', dayfirst=True)
                            colunas_export['Data de Vencimento'] = venc_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except Exception:
                            colunas_export['Data de Vencimento'] = df[col_venc].fillna('').astype(str).str.strip()
                    else:
                        colunas_export['Data de Vencimento'] = pd.Series([''] * len(df), index=df.index)
                    col_infracao = _resolver_coluna_data(df, "Data Infração")
                    if col_infracao:
                        try:
                            infracao_dt = pd.to_datetime(df[col_infracao], errors='coerce', dayfirst=True)
                            colunas_export['Data Infração'] = infracao_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except Exception:
                            colunas_export['Data Infração'] = df[col_infracao].fillna('').astype(str).str.strip()
                    else:
                        colunas_export['Data Infração'] = pd.Series([''] * len(df), index=df.index)
                    if 'Data Pagamento' in df.columns:
                        try:
                            pagamento_dt = pd.to_datetime(df['Data Pagamento'], errors='coerce', dayfirst=True)
                            colunas_export['Data Pagamento'] = pagamento_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except Exception:
                            colunas_export['Data Pagamento'] = df['Data Pagamento'].fillna('').astype(str).str.strip()
                    if coluna_modal_serasa and coluna_modal_serasa in df.columns:
                        colunas_export['Modais'] = df[coluna_modal_serasa].fillna('').astype(str).str.strip()
                    colunas_export['CPF_CNPJ'] = df[coluna_cpf_cnpj].fillna('').astype(str).str.strip() if coluna_cpf_cnpj in df.columns else pd.Series([''] * len(df), index=df.index)
                    colunas_export['Valor (R$)'] = df['_VALOR_NUM'].apply(formatar_valor_br)
                    ordem = ['Auto de Infração']
                    if 'Número de Protocolo' in colunas_export:
                        ordem.append('Número de Protocolo')
                    ordem.append('Nome Autuado')
                    ordem.append('Classificação Autuado')
                    ordem.append('Motivo Classificação')
                    ordem.append('Termo Identificado')
                    if 'Situação Dívida' in colunas_export:
                        ordem.append('Situação Dívida')
                    if 'Situação Congelamento' in colunas_export:
                        ordem.append('Situação Congelamento')
                    ordem.append('Data de Vencimento')
                    ordem.append('Data Infração')
                    if 'Data Pagamento' in colunas_export:
                        ordem.append('Data Pagamento')
                    if 'Modais' in colunas_export:
                        ordem.append('Modais')
                    ordem.append('CPF_CNPJ')
                    ordem.append('Valor (R$)')
                    dados = pd.DataFrame({k: colunas_export[k] for k in ordem if k in colunas_export}, index=df.index)
                    dados['CPF_CNPJ'] = dados['CPF_CNPJ'].apply(formatar_cpf_cnpj_brasileiro)
                    dados = dados.replace('', np.nan).dropna(how='all').reset_index(drop=True)
                    return dados

                # Layout PADRÃO: botão "Todos" grande, e os dois botões de faixa de valor lado a lado
                st.markdown("##### 📥 Todos os Autos")
                st.markdown("**Sem filtro de valor**")
                render_exportacao_excel(
                    chave="todos_autos",
                    titulo_gerar="Gerar arquivo Todos os Autos",
                    label_download="📥 Download Todos os Autos",
                    nome_aba="Todos_Autos",
                    nome_arquivo=f"Base Comparada Todos {data_arquivo}.xlsx",
                    help_download="Arquivo Excel com TODOS os autos que estão em ambas as bases (sem filtro de valor)",
                    producer=lambda: preparar_dados_exportacao(df_export, filtro_valor=None),
                    empty_warning="⚠️ Nenhum auto encontrado",
                    success_template="✅ {qtd:,} autos (todos)"
                )

                col_exp1, col_exp2 = st.columns(2)

                # BOTÃO: Autos até R$ 999,99
                with col_exp1:
                    st.markdown("##### 📥 Autos até R$ 999,99")
                    render_exportacao_excel(
                        chave="autos_ate_999",
                        titulo_gerar="Gerar arquivo Autos ≤ R$ 999,99",
                        label_download="📥 Download Autos ≤ R$ 999,99",
                        nome_aba="Autos_Ate_999",
                        nome_arquivo=f"Base Comparada Ate 999 {data_arquivo}.xlsx",
                        help_download="Arquivo Excel com autos de valor até R$ 999,99",
                        producer=lambda: preparar_dados_exportacao(df_export, filtro_valor='ate_999'),
                        empty_warning="⚠️ Nenhum auto encontrado até R$ 999,99",
                        success_template="✅ {qtd:,} autos até R$ 999,99"
                    )

                # BOTÃO: Autos acima de R$ 1.000,00
                with col_exp2:
                    st.markdown("##### 📥 Autos acima de R$ 1.000,00")
                    render_exportacao_excel(
                        chave="autos_acima_1000",
                        titulo_gerar="Gerar arquivo Autos > R$ 1.000,00",
                        label_download="📥 Download Autos > R$ 1.000,00",
                        nome_aba="Autos_Acima_1000",
                        nome_arquivo=f"Base Comparada Acima 1000 {data_arquivo}.xlsx",
                        help_download="Arquivo Excel com autos de valor acima de R$ 1.000,00",
                        producer=lambda: preparar_dados_exportacao(df_export, filtro_valor='acima_1000'),
                        empty_warning="⚠️ Nenhum auto encontrado acima de R$ 1.000,00",
                        success_template="✅ {qtd:,} autos acima de R$ 1.000,00"
                    )
                
                # BOTÃO 4: TODOS os autos SEM FILTRO DE DATA
                st.markdown("---")
                st.markdown("#### 📥 Exportar Base Sem Filtro de Data")
                st.info("💡 Exporte TODOS os autos que estão em ambas as bases, **independente da data de vencimento**. Mantém todos os outros filtros (valores > 0, remoção de duplicados, etc.) e também **exclui autuados classificados como Órgão, Banco ou Leasing**, preservando exceções permitidas como **SAFRA**.")
                
                if 'df_final_geral' in resultados and not resultados['df_final_geral'].empty:
                    df_final_geral_work = resultados['df_final_geral'].copy()
                    
                    # Criar df_export_geral mapeando colunas do df_final_geral para estrutura esperada
                    df_export_geral = pd.DataFrame()
                    
                    # Mapear Auto de Infração
                    if f"{coluna_auto}_serasa" in df_final_geral_work.columns:
                        df_export_geral[coluna_auto] = df_final_geral_work[f"{coluna_auto}_serasa"]
                    elif f"{coluna_auto}_divida" in df_final_geral_work.columns:
                        df_export_geral[coluna_auto] = df_final_geral_work[f"{coluna_auto}_divida"]
                    elif 'AUTO_NORM' in df_final_geral_work.columns:
                        # Se tiver AUTO_NORM, buscar do df_serasa_clean (sem filtro de data)
                        autos_gerais = set(df_final_geral_work['AUTO_NORM'].unique())
                        df_serasa_geral = resultados['df_serasa_original'][resultados['df_serasa_original']['AUTO_NORM'].isin(autos_gerais)].copy()
                        df_export_geral = df_serasa_geral.copy()
                    else:
                        # Fallback
                        autos_gerais = set(df_final_geral_work['AUTO_NORM'].unique())
                        df_serasa_geral = resultados['df_serasa_original'][resultados['df_serasa_original']['AUTO_NORM'].isin(autos_gerais)].copy()
                        df_export_geral = df_serasa_geral.copy()
                    
                    # Se conseguiu mapear, adicionar outras colunas
                    if coluna_auto in df_export_geral.columns and len(df_export_geral) == len(df_final_geral_work):
                        # Adicionar coluna de valor convertida (do df_final_geral)
                        if 'Valor' in df_final_geral_work.columns:
                            df_export_geral[coluna_valor] = df_final_geral_work['Valor']
                        
                        # Adicionar outras colunas necessárias
                        if coluna_cpf_cnpj in resultados['df_serasa_original'].columns:
                            # Mapear CPF/CNPJ
                            if f"{coluna_cpf_cnpj}_serasa" in df_final_geral_work.columns:
                                df_export_geral[coluna_cpf_cnpj] = df_final_geral_work[f"{coluna_cpf_cnpj}_serasa"]
                            elif f"{coluna_cpf_cnpj}_divida" in df_final_geral_work.columns:
                                df_export_geral[coluna_cpf_cnpj] = df_final_geral_work[f"{coluna_cpf_cnpj}_divida"]
                            
                            # Adicionar CPF_CNPJ_NORM para compatibilidade com preparar_dados_exportacao
                            if coluna_cpf_cnpj in df_export_geral.columns:
                                df_export_geral['CPF_CNPJ_NORM'] = df_export_geral[coluna_cpf_cnpj].apply(normalizar_cpf_cnpj)
                        
                        # Adicionar coluna de vencimento
                        vencimento_adicionado_geral = False
                        col_venc_geral = resolver_coluna_vencimento(df_final_geral_work, coluna_vencimento)
                        if col_venc_geral:
                            df_export_geral[coluna_vencimento] = df_final_geral_work[col_venc_geral]
                            vencimento_adicionado_geral = True
                        
                        if not vencimento_adicionado_geral:
                            if coluna_vencimento in resultados['df_serasa_original'].columns and 'AUTO_NORM' in df_export_geral.columns:
                                vencimento_map_geral = resultados['df_serasa_original'].set_index('AUTO_NORM')[coluna_vencimento].to_dict()
                                df_export_geral[coluna_vencimento] = df_export_geral['AUTO_NORM'].map(vencimento_map_geral).fillna('')
                            else:
                                df_export_geral[coluna_vencimento] = pd.Series([''] * len(df_export_geral), index=df_export_geral.index)
                        
                        # Adicionar coluna de protocolo
                        if coluna_protocolo in resultados['df_serasa_original'].columns:
                            if f"{coluna_protocolo}_serasa" in df_final_geral_work.columns:
                                df_export_geral[coluna_protocolo] = df_final_geral_work[f"{coluna_protocolo}_serasa"]
                            elif f"{coluna_protocolo}_divida" in df_final_geral_work.columns:
                                df_export_geral[coluna_protocolo] = df_final_geral_work[f"{coluna_protocolo}_divida"]
                        
                        # Adicionar coluna de Modais
                        if 'Modais' in df_final_geral_work.columns:
                            df_export_geral['Modais'] = df_final_geral_work['Modais']
                        elif 'Modais' in resultados['df_serasa_original'].columns and 'AUTO_NORM' in df_export_geral.columns:
                            modais_map_geral = resultados['df_serasa_original'].set_index('AUTO_NORM')['Modais'].to_dict()
                            df_export_geral['Modais'] = df_export_geral['AUTO_NORM'].map(modais_map_geral).fillna('')

                        # Adicionar Nome Autuado da base SERASA para exclusão na exportação comparada sem filtro de data
                        if coluna_nome_autuado in resultados['df_serasa_original'].columns:
                            if f"{coluna_nome_autuado}_serasa" in df_final_geral_work.columns:
                                df_export_geral[coluna_nome_autuado] = df_final_geral_work[f"{coluna_nome_autuado}_serasa"]
                            elif 'AUTO_NORM' in df_export_geral.columns and 'AUTO_NORM' in resultados['df_serasa_original'].columns:
                                nome_autuado_map_geral = resultados['df_serasa_original'].set_index('AUTO_NORM')[coluna_nome_autuado].to_dict()
                                df_export_geral[coluna_nome_autuado] = df_export_geral['AUTO_NORM'].map(nome_autuado_map_geral).fillna('')
                        
                        # Adicionar AUTO_NORM para compatibilidade
                        if 'AUTO_NORM' in df_final_geral_work.columns:
                            df_export_geral['AUTO_NORM'] = df_final_geral_work['AUTO_NORM']
                        # Coluna Situação decadente (regra de datas: 37 dias autuação, 187 dias multa)
                        if 'Situação decadente' in df_final_geral_work.columns:
                            df_export_geral['Situação decadente'] = df_final_geral_work['Situação decadente']
                    else:
                        # Fallback: usar df_final_geral diretamente
                        df_export_geral = df_final_geral_work.copy()
                    
                    render_exportacao_excel(
                        chave="todos_sem_filtro_data",
                        titulo_gerar="Gerar arquivo Todos os Autos (Sem Filtro de Data)",
                        label_download="📥 Download Todos os Autos (Sem Filtro de Data)",
                        nome_aba="Todos_Sem_Filtro_Data",
                        nome_arquivo=f"Base Comparada Sem Filtro Data {data_arquivo}.xlsx",
                        help_download="Arquivo Excel com TODOS os autos que estão em ambas as bases (sem filtro de data de vencimento)",
                        producer=lambda: preparar_dados_exportacao(df_export_geral, filtro_valor=None),
                        empty_warning="⚠️ Nenhum auto encontrado sem filtro de data",
                        success_template="✅ {qtd:,} autos únicos (sem filtro de data)",
                        extra_caption_fn=lambda dados_df: "📋 Todos os autos (sem filtro de data)"
                    )
                    
                    # Criar duas colunas para os dois novos botões de exportação por valor (sem filtro de data)
                    col_geral_exp1, col_geral_exp2 = st.columns(2)
                    
                    # BOTÃO 5: Autos até R$ 999,99 SEM FILTRO DE DATA
                    with col_geral_exp1:
                        st.markdown("##### 📥 Autos até R$ 999,99 (Sem Filtro de Data)")
                        render_exportacao_excel(
                            chave="ate_999_sem_filtro_data",
                            titulo_gerar="Gerar arquivo Autos ≤ R$ 999,99 (Sem Filtro de Data)",
                            label_download="📥 Download Autos ≤ R$ 999,99 (Sem Filtro de Data)",
                            nome_aba="Autos_Ate_999_Sem_Data",
                            nome_arquivo=f"Base Comparada Ate 999 Sem Filtro Data {data_arquivo}.xlsx",
                            help_download="Arquivo Excel com autos de valor até R$ 999,99 (sem filtro de data de vencimento)",
                            producer=lambda: preparar_dados_exportacao(df_export_geral, filtro_valor='ate_999'),
                            empty_warning="⚠️ Nenhum auto encontrado até R$ 999,99 (sem filtro de data)",
                            success_template="✅ {qtd:,} autos até R$ 999,99 (sem filtro de data)"
                        )
                    
                    # BOTÃO 6: Autos acima de R$ 1.000,00 SEM FILTRO DE DATA
                    with col_geral_exp2:
                        st.markdown("##### 📥 Autos acima de R$ 1.000,00 (Sem Filtro de Data)")
                        render_exportacao_excel(
                            chave="acima_1000_sem_filtro_data",
                            titulo_gerar="Gerar arquivo Autos > R$ 1.000,00 (Sem Filtro de Data)",
                            label_download="📥 Download Autos > R$ 1.000,00 (Sem Filtro de Data)",
                            nome_aba="Autos_Acima_1000_Sem_Data",
                            nome_arquivo=f"Base Comparada Acima 1000 Sem Filtro Data {data_arquivo}.xlsx",
                            help_download="Arquivo Excel com autos de valor acima de R$ 1.000,00 (sem filtro de data de vencimento)",
                            producer=lambda: preparar_dados_exportacao(df_export_geral, filtro_valor='acima_1000'),
                            empty_warning="⚠️ Nenhum auto encontrado acima de R$ 1.000,00 (sem filtro de data)",
                            success_template="✅ {qtd:,} autos acima de R$ 1.000,00 (sem filtro de data)"
                        )
                
                # Exportar Apenas na SERASA (autos que não constam na Dívida Ativa)
                st.markdown("---")
                st.markdown("#### 📥 Exportar Apenas na SERASA")
                st.info("💡 Esta exportação contém apenas os autos que estão na base SERASA e **não** constam na Dívida Ativa. O sistema compara as duas bases e exclui da relação todos os autos que têm em comum; o arquivo traz somente o que restou da SERASA após essa exclusão. **Sem filtro de vencimento nem de valor:** todos os anos e todos os valores. **Uma linha por auto:** duplicados por Auto de Infração são removidos, mantendo a linha com data de vencimento mais recente.")
                st.caption("📐 **Como é feita a conta:** A comparação usa **autos únicos** (cada Auto de Infração conta uma vez). O \"Total Registros SERASA\" é o total de **linhas** da planilha; o número de **autos únicos** pode ser um pouco menor quando o mesmo auto aparece em mais de uma linha. Apenas na SERASA = autos únicos SERASA − autos únicos em ambas.")
                if 'df_autos_apenas_serasa' in resultados and not resultados['df_autos_apenas_serasa'].empty:
                    render_exportacao_excel(
                        chave="apenas_serasa",
                        titulo_gerar="Gerar arquivo Apenas na SERASA",
                        label_download="📥 Download Apenas na SERASA (sem autos da Dívida)",
                        nome_aba="Apenas_SERASA",
                        nome_arquivo=f"Base Apenas SERASA {data_arquivo}.xlsx",
                        help_download="Arquivo Excel com autos que estão somente na SERASA (excluídos os que constam na Dívida Ativa), com coluna Data Infração.",
                        producer=lambda: preparar_dados_apenas_serasa(resultados['df_autos_apenas_serasa']),
                        empty_warning="⚠️ Nenhum dado disponível para exportar (apenas na SERASA).",
                        success_template="✅ {qtd:,} autos apenas na SERASA"
                    )
                else:
                    st.warning("⚠️ Nenhum auto encontrado apenas na SERASA (todos os autos da SERASA constam na Dívida Ativa).")

                # Exportar base SERASA inteira com classificação de autuado (não pode cobrar: órgão, banco, leasing)
                st.markdown("---")
                st.markdown("#### 📥 Exportar base SERASA com classificação de autuado")
                st.info("💡 Exporta **toda a base SERASA** sem filtro (sem comparação com Dívida). Adiciona as colunas **Classificação Autuado**, **Motivo Classificação** e **Termo Identificado** para explicar por que o nome foi marcado como **Não pode cobrar - Órgão**, **Não pode cobrar - Banco**, **Não pode cobrar - Leasing** ou **Pode cobrar**.")
                if 'df_serasa_original' in resultados and resultados['df_serasa_original'] is not None and not resultados['df_serasa_original'].empty:
                    render_exportacao_excel(
                        chave="serasa_classificacao_autuado",
                        titulo_gerar="Gerar arquivo Base SERASA com classificação",
                        label_download="📥 Download Base SERASA com classificação de autuado",
                        nome_aba="SERASA_Classificacao",
                        nome_arquivo=f"Base SERASA Classificacao Autuados {data_arquivo}.xlsx",
                        help_download="Planilha com toda a base SERASA e coluna Classificação Autuado (Órgão, Banco, Leasing, Pode cobrar)",
                        producer=lambda: preparar_dados_serasa_classificacao(resultados['df_serasa_original']),
                        empty_warning="⚠️ Nenhum dado disponível para exportar.",
                        success_template="✅ {qtd:,} registros com classificação",
                        extra_caption_fn=get_resumo_classificacao_autuado
                    )
                else:
                    st.warning("⚠️ Base SERASA não disponível para exportação.")
                
                # Resumo geral
                st.markdown("---")
                st.caption(f"📅 Data de extração: {data_extracao}")
                st.caption(f"✅ **Dados comparados:** Estes são os autos que estão presentes em AMBAS as bases (SERASA e Dívida Ativa)")
                st.caption(f"💡 **Classificação por faixa de valor:** Baseada na SOMA dos valores por CPF/CNPJ. Cada arquivo contém valores individuais de cada auto, mas a classificação (até R$ 999,99 ou acima de R$ 1.000,00) é feita pela soma total do CPF/CNPJ.")
        
        # Lista de autos em ambas
        if resultados['autos_em_ambas'] > 0:
            st.markdown("---")
            st.markdown("#### 📝 Lista de Autos de Infração em Ambas as Bases")
            autos_lista = sorted(list(resultados['autos_em_ambas_lista']))
            # Mostrar em colunas
            num_cols = 4
            cols = st.columns(num_cols)
            for idx, auto in enumerate(autos_lista[:100]):  # Limitar a 100 para performance
                with cols[idx % num_cols]:
                    st.text(auto)
            if len(autos_lista) > 100:
                st.caption(f"Mostrando 100 de {len(autos_lista)} autos. Use a exportação para ver a lista completa.")
    
    with tab2:
        st.markdown("### 📈 Agrupamento por CPF/CNPJ")
        st.info("💡 CPF/CNPJ ordenados do **MAIOR** para o **MENOR** número de autos de infração")
        
        if not resultados['agrupado_serasa'].empty:
            # Estatísticas gerais
            total_cpf = len(resultados['agrupado_serasa'])
            total_autos = resultados['agrupado_serasa']['QTD_AUTOS'].sum()
            maior_qtd = resultados['agrupado_serasa']['QTD_AUTOS'].max()
            menor_qtd = resultados['agrupado_serasa']['QTD_AUTOS'].min()
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total CPF/CNPJ", f"{total_cpf:,}")
            with col2:
                st.metric("Total Autos", f"{total_autos:,}")
            with col3:
                st.metric("Maior Quantidade", f"{int(maior_qtd)} autos")
            with col4:
                st.metric("Menor Quantidade", f"{int(menor_qtd)} autos")
            
            st.markdown("---")
            
            # Tabela ordenada (já está ordenada do maior para menor)
            st.markdown("#### 📊 Lista Completa (Ordenada: Maior → Menor)")
            st.success(f"✅ Ordenação: Do CPF/CNPJ com **{int(maior_qtd)} autos** até o com **{int(menor_qtd)} autos**")
            
            # Adicionar índice para mostrar a posição
            agrupado_display = resultados['agrupado_serasa'].copy()
            agrupado_display.insert(0, 'Posição', range(1, len(agrupado_display) + 1))
            
            # Formatar valores
            if 'VALOR_TOTAL' in agrupado_display.columns:
                agrupado_display['VALOR_TOTAL_FORMAT'] = agrupado_display['VALOR_TOTAL'].apply(
                    lambda x: f"R$ {x:,.2f}" if pd.notna(x) else "N/A"
                )
            
            # Visualização otimizada para grandes volumes (como Excel - um embaixo do outro)
            st.markdown("**💡 Dica:** Use a barra de rolagem para navegar por todos os registros. Os dados estão ordenados do maior para o menor número de autos, como no Excel.")
            
            # Configurar altura dinâmica baseada na quantidade (máximo 800px para performance)
            altura_tabela = min(800, max(400, total_cpf * 30))
            
            st.dataframe(
                agrupado_display, 
                use_container_width=True, 
                height=altura_tabela,
                hide_index=True
            )
            
            st.caption(f"📊 Exibindo {total_cpf:,} CPF/CNPJ ordenados do maior ({int(maior_qtd)} autos) para o menor ({int(menor_qtd)} autos)")
            st.info("💡 Para exportar os dados comparados, use a aba **Autos de Infração** onde está o download principal.")
        else:
            st.warning("⚠️ Não foi possível realizar o agrupamento. Verifique se a coluna de valor e CPF/CNPJ estão corretas.")
    
    with tab3:
        st.markdown("### Separação por Valores - Autos de Infração ANTT")
        
        # Área de Exportação - PRIMEIRO
        st.markdown("#### 📥 Exportar Planilhas por Valor")
        st.info("💡 Baixe as planilhas separadas por valor em formato Excel (.xlsx) com formatação completa")
        
        col_exp1, col_exp2 = st.columns(2)
        
        with col_exp1:
            st.markdown("##### 📥 Autos ≤ R$ 1.000,00")
            if not resultados['serasa_abaixo_1000_ind'].empty:
                dados_abaixo_1000 = resultados['serasa_abaixo_1000_ind'].copy()
                # Ordenar por quantidade de autos por CPF/CNPJ (maior para menor) e depois por CPF/CNPJ para agrupar autos do mesmo CNPJ
                if 'CPF_CNPJ_NORM' in dados_abaixo_1000.columns:
                    contagem_autos_abaixo = dados_abaixo_1000.groupby('CPF_CNPJ_NORM').size()
                    dados_abaixo_1000['_QTD_AUTOS'] = dados_abaixo_1000['CPF_CNPJ_NORM'].map(contagem_autos_abaixo).fillna(0)
                    dados_abaixo_1000 = dados_abaixo_1000.sort_values(['_QTD_AUTOS', 'CPF_CNPJ_NORM'], ascending=[False, True]).drop(columns=['_QTD_AUTOS'])
                # Preparar dados para exportação
                if coluna_auto in dados_abaixo_1000.columns and coluna_cpf_cnpj in dados_abaixo_1000.columns and coluna_valor in dados_abaixo_1000.columns:
                    df_export_abaixo = pd.DataFrame({
                        'Auto de Infração': dados_abaixo_1000[coluna_auto].fillna('').astype(str).str.strip(),
                        'CPF_CNPJ': dados_abaixo_1000[coluna_cpf_cnpj].fillna('').astype(str).str.strip().apply(formatar_cpf_cnpj_brasileiro),
                        'Valor': pd.to_numeric(dados_abaixo_1000[coluna_valor], errors='coerce')
                    })
                    # Adicionar coluna de Protocolo se existir no DataFrame
                    if coluna_protocolo in dados_abaixo_1000.columns:
                        df_export_abaixo.insert(1, 'Número de Protocolo', dados_abaixo_1000[coluna_protocolo].fillna('').astype(str).str.strip())
                    # Adicionar coluna de Data de Vencimento se existir no DataFrame
                    if coluna_vencimento in dados_abaixo_1000.columns:
                        try:
                            vencimento_dt = pd.to_datetime(
                                dados_abaixo_1000[coluna_vencimento],
                                errors='coerce',
                                dayfirst=True
                            )
                            data_vencimento = vencimento_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except:
                            data_vencimento = dados_abaixo_1000[coluna_vencimento].fillna('').astype(str).str.strip()
                        # Inserir após Protocolo (se existir) ou após Auto de Infração
                        posicao = 2 if coluna_protocolo in dados_abaixo_1000.columns else 1
                        df_export_abaixo.insert(posicao, 'Data de Vencimento', data_vencimento)
                    
                    # Gerar Excel com formatação completa
                    buffer_abaixo = io.BytesIO()
                    with pd.ExcelWriter(buffer_abaixo, engine='openpyxl') as writer:
                        df_export_abaixo.to_excel(writer, sheet_name='Autos_≤1000', index=False)
                        worksheet = writer.sheets['Autos_≤1000']
                        
                        # Aplicar formatação completa
                        num_colunas = len(df_export_abaixo.columns)
                        tem_protocolo = 'Número de Protocolo' in df_export_abaixo.columns
                        tem_data_venc = 'Data de Vencimento' in df_export_abaixo.columns
                        
                        if num_colunas == 5:  # Auto, Protocolo, Data Vencimento, CPF/CNPJ, Valor
                            worksheet.column_dimensions['A'].width = 25  # Auto de Infração
                            worksheet.column_dimensions['B'].width = 20  # Número de Protocolo
                            worksheet.column_dimensions['C'].width = 18  # Data de Vencimento
                            worksheet.column_dimensions['D'].width = 18  # CPF/CNPJ
                            worksheet.column_dimensions['E'].width = 15  # Valor
                        elif num_colunas == 4:  # Auto, Protocolo, CPF/CNPJ, Valor OU Auto, Data Vencimento, CPF/CNPJ, Valor
                            worksheet.column_dimensions['A'].width = 25  # Auto de Infração
                            worksheet.column_dimensions['B'].width = 20 if tem_protocolo else 18  # Protocolo ou Data
                            worksheet.column_dimensions['C'].width = 18  # CPF/CNPJ ou Data
                            worksheet.column_dimensions['D'].width = 15  # Valor
                        else:  # Auto, CPF/CNPJ, Valor (sem protocolo e sem data)
                            worksheet.column_dimensions['A'].width = 25
                            worksheet.column_dimensions['B'].width = 18
                            worksheet.column_dimensions['C'].width = 15
                        
                        header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF", size=11)
                        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                        
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        coluna_cpf_idx = 3 if num_colunas == 4 else 2
                        coluna_valor_idx = 4 if num_colunas == 4 else 3
                        
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                            for cell in row:
                                cell.border = thin_border
                                if cell.column == coluna_cpf_idx and cell.row > 1:  # CPF/CNPJ
                                    cell.number_format = '@'
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                elif cell.column == coluna_valor_idx and cell.row > 1:  # Valor
                                    if cell.value is not None:
                                        cell.number_format = '#,##0.00'
                                        cell.alignment = Alignment(horizontal="right", vertical="center")
                                elif cell.column == 1 and cell.row > 1:  # Auto de Infração
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                                elif cell.column == 2 and num_colunas == 4 and cell.row > 1:  # Protocolo (se existir)
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                        worksheet.freeze_panes = 'A2'
                    
                    buffer_abaixo.seek(0)
                    excel_abaixo = buffer_abaixo.getvalue()
                    buffer_abaixo.close()
                    
                    st.download_button(
                        label="📥 Download Autos ≤ R$ 1.000,00 (Excel)",
                        data=excel_abaixo,
                        file_name=f"Autos Ate 1000 {datetime.now().strftime('%d %m %Y %H:%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_abaixo_1000"
                    )
                    st.caption(f"📋 Total: {len(resultados['serasa_abaixo_1000_ind'])} autos de infração")
                else:
                    st.warning("⚠️ Colunas necessárias não encontradas")
            else:
                st.warning("Nenhum auto encontrado nesta faixa de valor")
        
        with col_exp2:
            st.markdown("##### 📥 Autos > R$ 1.000,00")
            if not resultados['serasa_acima_1000_ind'].empty:
                dados_acima_1000 = resultados['serasa_acima_1000_ind'].copy()
                # Ordenar por quantidade de autos por CPF/CNPJ (maior para menor) e depois por CPF/CNPJ para agrupar autos do mesmo CNPJ
                if 'CPF_CNPJ_NORM' in dados_acima_1000.columns:
                    contagem_autos_acima = dados_acima_1000.groupby('CPF_CNPJ_NORM').size()
                    dados_acima_1000['_QTD_AUTOS'] = dados_acima_1000['CPF_CNPJ_NORM'].map(contagem_autos_acima).fillna(0)
                    dados_acima_1000 = dados_acima_1000.sort_values(['_QTD_AUTOS', 'CPF_CNPJ_NORM'], ascending=[False, True]).drop(columns=['_QTD_AUTOS'])
                # Preparar dados para exportação
                if coluna_auto in dados_acima_1000.columns and coluna_cpf_cnpj in dados_acima_1000.columns and coluna_valor in dados_acima_1000.columns:
                    df_export_acima = pd.DataFrame({
                        'Auto de Infração': dados_acima_1000[coluna_auto].fillna('').astype(str).str.strip(),
                        'CPF_CNPJ': dados_acima_1000[coluna_cpf_cnpj].fillna('').astype(str).str.strip().apply(formatar_cpf_cnpj_brasileiro),
                        'Valor': pd.to_numeric(dados_acima_1000[coluna_valor], errors='coerce')
                    })
                    # Adicionar coluna de Protocolo se existir no DataFrame
                    if coluna_protocolo in dados_acima_1000.columns:
                        df_export_acima.insert(1, 'Número de Protocolo', dados_acima_1000[coluna_protocolo].fillna('').astype(str).str.strip())
                    # Adicionar coluna de Data de Vencimento se existir no DataFrame
                    if coluna_vencimento in dados_acima_1000.columns:
                        try:
                            vencimento_dt = pd.to_datetime(
                                dados_acima_1000[coluna_vencimento],
                                errors='coerce',
                                dayfirst=True
                            )
                            data_vencimento = vencimento_dt.dt.strftime('%d/%m/%Y').fillna('')
                        except:
                            data_vencimento = dados_acima_1000[coluna_vencimento].fillna('').astype(str).str.strip()
                        # Inserir após Protocolo (se existir) ou após Auto de Infração
                        posicao = 2 if coluna_protocolo in dados_acima_1000.columns else 1
                        df_export_acima.insert(posicao, 'Data de Vencimento', data_vencimento)
                    
                    # Gerar Excel com formatação completa
                    buffer_acima = io.BytesIO()
                    with pd.ExcelWriter(buffer_acima, engine='openpyxl') as writer:
                        df_export_acima.to_excel(writer, sheet_name='Autos_>1000', index=False)
                        worksheet = writer.sheets['Autos_>1000']
                        
                        # Aplicar formatação completa
                        num_colunas = len(df_export_acima.columns)
                        tem_protocolo = 'Número de Protocolo' in df_export_acima.columns
                        tem_data_venc = 'Data de Vencimento' in df_export_acima.columns
                        
                        if num_colunas == 5:  # Auto, Protocolo, Data Vencimento, CPF/CNPJ, Valor
                            worksheet.column_dimensions['A'].width = 25  # Auto de Infração
                            worksheet.column_dimensions['B'].width = 20  # Número de Protocolo
                            worksheet.column_dimensions['C'].width = 18  # Data de Vencimento
                            worksheet.column_dimensions['D'].width = 18  # CPF/CNPJ
                            worksheet.column_dimensions['E'].width = 15  # Valor
                        elif num_colunas == 4:  # Auto, Protocolo, CPF/CNPJ, Valor OU Auto, Data Vencimento, CPF/CNPJ, Valor
                            worksheet.column_dimensions['A'].width = 25  # Auto de Infração
                            worksheet.column_dimensions['B'].width = 20 if tem_protocolo else 18  # Protocolo ou Data
                            worksheet.column_dimensions['C'].width = 18  # CPF/CNPJ ou Data
                            worksheet.column_dimensions['D'].width = 15  # Valor
                        else:  # Auto, CPF/CNPJ, Valor (sem protocolo e sem data)
                            worksheet.column_dimensions['A'].width = 25
                            worksheet.column_dimensions['B'].width = 18
                            worksheet.column_dimensions['C'].width = 15
                        
                        header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF", size=11)
                        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                        
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        # Calcular índices das colunas (ordem: Auto, Protocolo, Data Venc, Modais, CPF/CNPJ, Valor)
                        tem_modais = 'Modais' in df_export_acima.columns
                        idx_auto = 1
                        idx_protocolo = 2 if tem_protocolo else None
                        idx_data_venc = None
                        idx_modais = None
                        idx_cpf = None
                        idx_valor = num_colunas
                        
                        # Calcular índices dinamicamente baseado nas colunas presentes
                        col_idx = 1
                        if tem_protocolo:
                            col_idx += 1
                        if tem_data_venc:
                            idx_data_venc = col_idx
                            col_idx += 1
                        if tem_modais:
                            idx_modais = col_idx
                            col_idx += 1
                        idx_cpf = col_idx
                        
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                            for cell in row:
                                cell.border = thin_border
                                if cell.column == idx_cpf and cell.row > 1:  # CPF/CNPJ
                                    cell.number_format = '@'
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                elif cell.column == idx_valor and cell.row > 1:  # Valor
                                    if cell.value is not None:
                                        cell.number_format = '#,##0.00'
                                        cell.alignment = Alignment(horizontal="right", vertical="center")
                                elif cell.column == idx_auto and cell.row > 1:  # Auto de Infração
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                                elif idx_protocolo and cell.column == idx_protocolo and cell.row > 1:  # Protocolo (se existir)
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                                elif idx_data_venc and cell.column == idx_data_venc and cell.row > 1:  # Data de Vencimento (se existir)
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                    cell.number_format = '@'  # Formato texto para manter formato DD/MM/YYYY
                                elif idx_modais and cell.column == idx_modais and cell.row > 1:  # Modais (se existir)
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                                    cell.number_format = '@'  # Formato texto
                        
                        worksheet.freeze_panes = 'A2'
                    
                    buffer_acima.seek(0)
                    excel_acima = buffer_acima.getvalue()
                    buffer_acima.close()
                    
                    st.download_button(
                        label="📥 Download Autos > R$ 1.000,00 (Excel)",
                        data=excel_acima,
                        file_name=f"Autos Acima 1000 {datetime.now().strftime('%d %m %Y %H:%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_acima_1000_tab3"
                    )
                    st.caption(f"📋 Total: {len(resultados['serasa_acima_1000_ind'])} autos de infração")
                else:
                    st.warning("⚠️ Colunas necessárias não encontradas")
            else:
                st.warning("Nenhum auto encontrado nesta faixa de valor")
        
        st.markdown("---")
        st.markdown("#### 📊 Visualização dos Dados")
        st.info("💡 Visualize os dados abaixo antes de exportar")
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**≤ R$ 1.000,00 (Individual)**")
            if not resultados['serasa_abaixo_1000_ind'].empty:
                st.dataframe(resultados['serasa_abaixo_1000_ind'], use_container_width=True, height=350)
                st.caption(f"📋 Total: {len(resultados['serasa_abaixo_1000_ind'])} autos de infração")
            else:
                st.warning("Nenhum auto encontrado nesta faixa de valor")
        
        with col2:
            st.markdown("**> R$ 1.000,00 (Individual)**")
            if not resultados['serasa_acima_1000_ind'].empty:
                st.dataframe(resultados['serasa_acima_1000_ind'], use_container_width=True, height=350)
                st.caption(f"📋 Total: {len(resultados['serasa_acima_1000_ind'])} autos de infração")
            else:
                st.warning("Nenhum auto encontrado nesta faixa de valor")
        
        st.markdown("---")
        st.markdown("#### 📊 SERASA - Valores Acumulativos (por CPF/CNPJ)")
        st.info("💡 A soma de todos os autos do mesmo CPF/CNPJ é considerada")
        
        col3, col4 = st.columns(2)
        
        with col3:
            st.markdown("**≤ R$ 1.000,00 (Acumulado)**")
            if not resultados['serasa_abaixo_1000_acum'].empty:
                st.markdown("**Resumo por CPF/CNPJ:**")
                st.dataframe(resultados['serasa_abaixo_1000_acum'], use_container_width=True, height=200)
                st.caption(f"📊 Total: {len(resultados['serasa_abaixo_1000_acum'])} CPF/CNPJ")
                
                st.markdown("**Autos de Infração Correspondentes:**")
                if not resultados['serasa_abaixo_1000_acum_autos'].empty:
                    st.dataframe(resultados['serasa_abaixo_1000_acum_autos'], use_container_width=True, height=300)
                    st.caption(f"📋 Total: {len(resultados['serasa_abaixo_1000_acum_autos'])} autos")
                else:
                    st.info("Nenhum auto encontrado")
            else:
                st.warning("Nenhum CPF/CNPJ encontrado nesta faixa de valor")
        
        with col4:
            st.markdown("**> R$ 1.000,00 (Acumulado)**")
            if not resultados['serasa_acima_1000_acum'].empty:
                st.markdown("**Resumo por CPF/CNPJ:**")
                st.dataframe(resultados['serasa_acima_1000_acum'], use_container_width=True, height=200)
                st.caption(f"📊 Total: {len(resultados['serasa_acima_1000_acum'])} CPF/CNPJ")
                
                st.markdown("**Autos de Infração Correspondentes:**")
                if not resultados['serasa_acima_1000_acum_autos'].empty:
                    st.dataframe(resultados['serasa_acima_1000_acum_autos'], use_container_width=True, height=300)
                    st.caption(f"📋 Total: {len(resultados['serasa_acima_1000_acum_autos'])} autos")
                else:
                    st.info("Nenhum auto encontrado")
            else:
                st.warning("Nenhum CPF/CNPJ encontrado nesta faixa de valor")
    
    with tab4:
        st.markdown("### ⚠️ Registros com Divergências")
        st.info("💡 Análise baseada em **Autos de Infração** (principal) e CPF/CNPJ (adicional)")
        
        st.markdown("#### 🔑 Divergências por Auto de Infração (PRINCIPAL)")
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("##### 🔴 Autos Apenas na Base SERASA")
            st.warning(f"Total de {resultados['autos_apenas_serasa']:,} autos de infração que não constam na Dívida Ativa")
            if resultados['autos_apenas_serasa'] > 0 and not resultados['df_autos_apenas_serasa'].empty:
                st.dataframe(resultados['df_autos_apenas_serasa'], use_container_width=True, height=400)
                st.caption(f"📋 {len(resultados['df_autos_apenas_serasa'])} registros")
            elif resultados['autos_apenas_serasa'] == 0:
                st.success("✅ Todos os autos da SERASA possuem correspondência na Dívida Ativa!")
        
        with col2:
            st.markdown("##### 🔴 Autos Apenas na Base Dívida Ativa")
            st.warning(f"Total de {resultados['autos_apenas_divida']:,} autos de infração que não constam na SERASA")
            if resultados['autos_apenas_divida'] > 0 and not resultados['df_autos_apenas_divida'].empty:
                st.dataframe(resultados['df_autos_apenas_divida'], use_container_width=True, height=400)
                st.caption(f"📋 {len(resultados['df_autos_apenas_divida'])} registros")
            elif resultados['autos_apenas_divida'] == 0:
                st.success("✅ Todos os autos da Dívida Ativa possuem correspondência na SERASA!")
        
        st.markdown("---")
        st.markdown("#### 📋 Divergências por CPF/CNPJ (ADICIONAL)")
        col3, col4 = st.columns(2)
        
        with col3:
            st.markdown("##### 🔴 CPF/CNPJ Apenas na Base SERASA")
            st.info(f"Total de {resultados['cpf_apenas_serasa']:,} CPF/CNPJ que não constam na Dívida Ativa")
            if resultados['cpf_apenas_serasa'] > 0 and not resultados['df_cpf_apenas_serasa'].empty:
                st.dataframe(resultados['df_cpf_apenas_serasa'], use_container_width=True, height=300)
            elif resultados['cpf_apenas_serasa'] == 0:
                st.success("✅ Todos os CPF/CNPJ da SERASA possuem correspondência na Dívida Ativa!")
        
        with col4:
            st.markdown("##### 🔴 CPF/CNPJ Apenas na Base Dívida Ativa")
            st.info(f"Total de {resultados['cpf_apenas_divida']:,} CPF/CNPJ que não constam na SERASA")
            if resultados['cpf_apenas_divida'] > 0 and not resultados['df_cpf_apenas_divida'].empty:
                st.dataframe(resultados['df_cpf_apenas_divida'], use_container_width=True, height=300)
            elif resultados['cpf_apenas_divida'] == 0:
                st.success("✅ Todos os CPF/CNPJ da Dívida Ativa possuem correspondência na SERASA!")
    

else:
    st.info("👆 Por favor, faça o upload das duas planilhas na barra lateral para iniciar a análise.")
    
    # Mostrar instruções
    with st.expander("ℹ️ Como usar o sistema"):
        st.markdown("""
        ### Instruções de Uso:
        
        1. **Upload de Arquivos**: Faça o upload das planilhas SERASA e Dívida Ativa na barra lateral
        2. **Configuração de Colunas**: Informe os nomes exatos das colunas:
           - ⚠️ **Coluna Auto de Infração** (OBRIGATÓRIA - ex: "Auto de Infração")
           - Coluna CPF/CNPJ (adicional)
           - Coluna Valor (adicional)
           - Coluna Vencimento (adicional)
        3. **Executar Análise**: Clique no botão "Executar Análise Completa"
        4. **Visualizar Resultados**: Explore as abas com os resultados detalhados
        5. **Exportar**: Baixe a planilha formatada na aba "Agrupamento por CPF/CNPJ"
        
        ### Funcionalidades:
        - ✅ **Análise principal por Auto de Infração** (chave de comparação)
        - ✅ Cruzamento automático entre bases baseado em autos
        - ✅ Filtro por ano de vencimento (2025 em diante)
        - ✅ Agrupamento por CPF/CNPJ (análise adicional)
        - ✅ Separação por valores SERASA (≤ R$ 1.000 e > R$ 1.000)
        - ✅ Visualização de autos de infração correspondentes
        - ✅ Análise individual e acumulativa
        - ✅ Identificação de divergências (autos e CPF/CNPJ)
        - ✅ Dashboard visual com gráficos
        - ✅ Exportação de resultados
        """)

